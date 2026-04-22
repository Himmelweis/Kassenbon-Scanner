# -*- coding: utf-8 -*-
"""
Created on Sun Aug 31 19:14:15 2025

@author: ONeum
"""

# =========================
# CONFIG
# =========================
import os

# Paddle / Paddlex leiser machen
os.environ["FLAGS_log_level"] = "3"   # 0=debug, 3=error only
os.environ["GLOG_minloglevel"] = "3"

# Model-Check deaktivieren (hast du teilweise schon)
os.environ["PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK"] = "True"

# oneDNN / MKL Spam reduzieren
os.environ["OMP_NUM_THREADS"] = "1"

import warnings

warnings.filterwarnings("ignore", category=UserWarning)
warnings.filterwarnings("ignore", category=FutureWarning)
warnings.filterwarnings("ignore", message=".*RequestsDependencyWarning.*")

import torch
import  re, cv2, pytesseract, pandas as pd, numpy as np
os.environ["PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK"] = "True"
from datetime import datetime
from PIL import Image
import re
from pathlib import Path

from paddle.device.cuda.cuda_graphed_layer import debug_print
from paddleocr import PaddleOCR
# ========= Debug Switches (standard: aus) =========

FAST = True
DEBUG_OCR_TYPES = False   # zeigt "DEBUG ocr_text_multi types: [...]"
DEBUG_PAYMENTS  = True   # zeigt "💳 Zahlungsart ..." Debug-Ausgaben
DEBUG_PRINTS = False  # global
DEBUG_HEAD = False  # bei Bedarf auf False setzen
DEBUG_FOOTER = False
DEBUG_ALL = False

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

SCAN_DIR = os.path.join(BASE_DIR, "Kassenbons")
EXCEL_PATH = os.path.join(BASE_DIR, "kassenbons.xlsx")

USE_PADDLE = True
_PADDLE_OCR = None

# =========================
# HELFER / BASISFUNKTIONEN
# =========================

def _print_clean_head(label: str, text: str, n: int = 20):
    if not DEBUG_HEAD:
        return
    lines = [ln for ln in (text or "").splitlines() if ln.strip()]
    print(f"\n--- CLEAN HEAD{(' ' + label) if label else ''} ---")
    for ln in lines[:n]:
        print(ln)


# --- Tesseract fest verdrahten ---
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
os.environ["TESSDATA_PREFIX"] = r"C:\Program Files\Tesseract-OCR\tessdata"

# ==== Debug-Helfer ====
def _print_clean_blocks(label: str, text: str, n_head: int = 15, n_tail: int = 25, multi: bool = False):
#    if not isinstance(text, str):
    if not DEBUG_HEAD:
        return
    tag = " (multi)" if multi else ""
    lines = text.splitlines()
    head = "\n".join(lines[:max(0, n_head)])
    tail = "\n".join(lines[-max(0, n_tail):])
    print(f"\n--- CLEAN HEAD{tag} ---\n{head}")
    print(f"\n--- CLEAN (letzte {n_tail} Zeilen{tag}) ---\n{tail}")

# ---------------------------
# Excel-Konfiguration
# ---------------------------
EXCEL_PATH = "kassenbons.xlsx"  # <- hier deinen neuen Dateinamen eintragen, z.B. "kassenbons_neu.xlsx"
STRICT_TOTAL = True  # ganz oben

COLUMNS = [
    "Datum","Uhrzeit","Laden","Liter","€/L","Betrag (€)","Gegeben (€)","Wechselgeld (€)",
    "MwSt %","MwSt (€)","Netto (€)","Säule","Kassierer","Beleg-Nr.",
    "Confidence","Prüfstatus","Rohtext"
]

# 1) Master-Spalten (Reihenfolge fix, Prüfstatus am Ende)
MASTER_COLUMNS = [
    "Datum","Uhrzeit","Laden",
    "Betrag (€)","Gegeben (€)","Wechselgeld (€)",
    "MwSt %","MwSt (€)","Netto (€)","Zahlung",
    "Belegtyp",  # Typ immer vor Prüfstatus
    "Prüfstatus" # << immer letzte Spalte
]

# 2) Zusätzliche Spalten je Belegtyp
EXTRA_BY_TYPE = {
    "fuel":        ["Kraftstoff","Säule","Liter","€/L","Kennzeichen","Beleg-Nr"],
    "grocery":     ["Artikelanzahl","Netto 7% (€)","MwSt 7% (€)","Netto 19% (€)","MwSt 19% (€)"],
    "pharmacy":    ["Rezeptpflichtig","Zuzahlung (€)","Netto 7% (€)","MwSt 7% (€)","Netto 19% (€)","MwSt 19% (€)"],
    "restaurant":  ["Tisch","Bedienung","Trinkgeld (€)","Netto 7% (€)","MwSt 7% (€)","Netto 19% (€)","MwSt 19% (€)"],
    "retail":      ["Artikelanzahl","Netto 7% (€)","MwSt 7% (€)","Netto 19% (€)","MwSt 19% (€)"],
    "generic":     [],  # Fallback
}

# 3) Blattnamen je Typ
SHEET_BY_TYPE = {
    "fuel": "Tankstelle",
    "grocery": "Discounter",
    "pharmacy": "Apotheke",
    "restaurant": "Restaurant",
    "retail": "Einzelhandel",
    "generic": "Einzelhandel",
}

def _columns_for_type(rtype: str) -> list[str]:
    """
    Liefert die Spaltenreihenfolge fürs Typ-Blatt:
    - Basis: MASTER_COLUMNS (ohne Belegtyp/Prüfstatus)
    - Dann alle EXTRA_BY_TYPE[rtype]
    - Dann 'Belegtyp' und 'Prüfstatus' (Prüfstatus ganz am Ende)
    """
    # 1) Basis: Master ohne die zwei Schluss-Spalten
    base = [c for c in MASTER_COLUMNS if c not in ("Belegtyp", "Prüfstatus")]

    # 2) Typ-Extras (nur solche, die noch nicht in base sind)
    extras = EXTRA_BY_TYPE.get(rtype, [])
    for c in extras:
        if c not in base:
            base.append(c)

    # 3) Schluss: Belegtyp, dann Prüfstatus
    base.append("Belegtyp")
    base.append("Prüfstatus")
    return base


def _order_row_for_columns(row: dict, columns: list) -> dict:
    """
    Baut ein dict passend zur Spaltenreihenfolge und füllt fehlende Keys mit None.
    """
    return {c: row.get(c) for c in columns}

# (optional) Für alte Aufrufer, die noch 'COLUMNS' verwenden:
COLUMNS = MASTER_COLUMNS[:]  # wird nicht fürs Schreiben genutzt, aber bricht nix

# =========================
# OCR / IO
# =========================
FAST = True  # ← auf True: schneller (1 Variante), auf False: 3 Varianten (besser)

CUSTOM_CONFIG_MAIN = r'--oem 3 --psm 6 -c preserve_interword_spaces=1'
CUSTOM_CONFIG_ALT  = r'--oem 3 --psm 4'
CUSTOM_CONFIG1 = r'--oem 3 --psm 6 -c preserve_interword_spaces=1'
CUSTOM_CONFIG2 = r'--oem 3 --psm 4'

MERCHANT_KEYWORDS = [
    # Tanken
    "bft", "aral", "shell", "esso", "jet", "total", "avanti", "q1", "statoil", "omv", "agip",
    # Supermärkte / Drogerie (für später)
    "rewe", "edeka", "aldi", "lidl", "kaufland", "penny", "dm", "rossmann", "müller"
]

def _pick_merchant(lines: list[str]) -> str | None:
    # 1) Keyword-Match bevorzugt
    for ln in lines[:12]:
        low = ln.lower()
        if any(k in low for k in MERCHANT_KEYWORDS):
            return ln.strip()
    # 2) erste sinnvolle Kopfzeile (keine Zahl/Total/MwSt etc.)
    for ln in lines[:8]:
        if re.match(r"^(#|\d|mw?st|netto|brutto|total|summe|gesamt)", ln, re.IGNORECASE):
            continue
        if len(ln.strip()) >= 3:
            return ln.strip()
    return None

MERCHANT_TO_KAT = {
    "bft": "Auto/Tanken", "shell": "Auto/Tanken", "aral": "Auto/Tanken",
    "rewe": "Lebensmittel", "aldi": "Lebensmittel", "lidl": "Lebensmittel",
    "dm": "Verbrauchsmaterial", "rossmann": "Verbrauchsmaterial",
}
def guess_category(name: str|None) -> str|None:
    if not name: return None
    low = name.lower()
    for k, cat in MERCHANT_TO_KAT.items():
        if k in low: return cat
    return None

# =================== Hilfsfunktionen ===================

import re
import re
_STORE_NOISE = re.compile(r"^[A-ZÄÖÜ\s]{1,6}$")
_STORE_BROKEN = re.compile(r"[A-ZÄÖÜ]{2,}\s+[A-ZÄÖÜ]{1,2}\s+[A-ZÄÖÜ]{2,}")
def _looks_like_noise(ln: str) -> bool:
    if not ln: return True
    s = ln.strip()
    if not s: return True
    letters = sum(c.isalpha() for c in s)
    if letters < 4: return True
    if _STORE_NOISE.match(s): return True
    if _STORE_BROKEN.search(s): return True
    low = sum(c.islower() for c in s)
    if len(s) > 24 and low == 0: return True
    return False

def coerce_type_by_store(store: str, fallback: str = "grocery") -> str:
    """Erzwingt anhand des Laden-Namens einen Belegtyp (z.B. grocery)."""
    s = (store or "").upper()
    if re.search(r"\b(NIELSEN|SCAN-?SHOP)\b", s):
        return "grocery"
    # weitere Ketten hier ergänzen (optional):
    # if re.search(r"\b(LIDL|ALDI|EDEKA|REWE|KAUFLAND|NORMA|PENNY|NETTO)\b", s): return "grocery"
    return None  # keine Meinung -> nichts erzwingen


def _reset_payment_log_once():
    # Attribut beim Funktionsobjekt zurücksetzen
    if hasattr(apply_payment_detection, "_did_log"):
        apply_payment_detection._did_log = False


def diagnose_image(path):
    import os, cv2, numpy as np
    print("🧪 Diagnose:", path)
    if not os.path.exists(path):
        print("❌ Datei existiert nicht.")
        return
    try:
        data = np.fromfile(path, dtype=np.uint8)
        img  = cv2.imdecode(data, cv2.IMREAD_UNCHANGED)
    except Exception as e:
        img = None
        print("❌ imdecode-Fehler:", e)
    if img is None:
        print("⚠️ imdecode None, versuche cv2.imread …")
        img = _imread_safe(path, cv2.IMREAD_UNCHANGED)
    if img is None:
        print("❌ Konnte nicht laden (Pfad/Datei beschädigt?).")
        return
    h, w = img.shape[:2]
    ch = 1 if len(img.shape)==2 else img.shape[2]
    print(f"✅ geladen: {w}x{h}px, Kanäle: {ch}")
    if ch == 4:
        print("ℹ️ Alphakanal vorhanden – in ocr_text_multi wird auf BGR reduziert.")
    if h > 8000:
        print("ℹ️ Sehr langes Bild – Kachelung wird verwendet.")

def diagnose_folder(folder):
    import os
    exts = (".png",".jpg",".jpeg",".tif",".tiff",".bmp",".webp")
    files = [os.path.join(folder, f) for f in os.listdir(folder)
             if f.lower().endswith(exts)]
    files.sort()
    print(f"📂 Ordner: {folder} | {len(files)} Dateien gefunden")
    for f in files:
        diagnose_image(f)


def _norm_money(s: str) -> float | None:
    if not s: return None
    s = s.strip().replace(" ", "")
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except:
        return None

def _find_first(patterns, text, flags=0):
    import re
    for p in patterns:
        m = re.search(p, text, flags)
        if m:
            return m
    return None


def _roi_lines(text: str, key_regex: str, window: int = 2) -> str:
    lines = [l for l in text.splitlines()]
    out = []
    for i, ln in enumerate(lines):
        if re.search(key_regex, ln, re.IGNORECASE):
            start = max(0, i - window)
            end   = min(len(lines), i + window + 1)
            out.extend(lines[start:end])
    return "\n".join(out) if out else text

import re

def _norm_ws(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())

def _is_amount_line(ln: str) -> bool:
    return bool(re.search(r"\d{1,3}(?:[.,]\d{3})*[.,]\d{2}\s*(?:€|EUR)?", ln))

def _looks_trash(ln: str) -> bool:
    # sehr kurze Splitter oder fast nur Sonderzeichen
    if len(ln.strip()) < 3:
        return True
    letters = sum(ch.isalpha() for ch in ln)
    return letters < 3

def _extract_store_from_footer(text: str) -> str | None:
    """
    Sucht im Bonfuß nach 'Einkauf getätigt in' und nutzt die nächsten 1–2 Zeilen,
    um einen Store-Name + Ort zu bauen. Speziell gut für Lidl-App-Bons.
    """
    lines = [ln.strip() for ln in (text or "").splitlines()]
    for i, ln in enumerate(lines):
        if re.search(r"(?i)Einkauf\s+getäti?gt\s+in", ln):
            # Nimm bis zu zwei sinnvolle Folgezeilen
            tail = []
            for j in range(i+1, min(i+4, len(lines))):
                s = _norm_ws(lines[j])
                if not s or _is_amount_line(s) or _looks_trash(s):
                    continue
                tail.append(s)
                if len(tail) >= 2:
                    break

            # Lidl-Heuristik: wenn im gesamten Text LIDL/Lidl Pay vorkommt → 'Lidl <Ort>'
            txt_up = (text or "").upper()
            if re.search(r"\bLIDL\b", txt_up) or re.search(r"\bLIDL\s*PAY\b", txt_up):
                city = None
                # suche eine Zeile mit PLZ + Ort oder etwas, das wie eine Ortszeile aussieht
                for z in tail:
                    m = re.search(r"\b(\d{5})\s+([A-ZÄÖÜ][\wÄÖÜäöüß\- ]+)", z)
                    if m:
                        city = m.group(2).strip()
                        break
                if not city and tail:
                    # fallback: erste „vernünftige“ tail-Zeile als Ort
                    city = tail[-1].split()[-1].strip()
                if city:
                    return f"Lidl {city}"
                else:
                    return "Lidl"

            # Allgemeiner Fallback: nimm erste sinnvolle Zeile als Laden
            if tail:
                # Wenn die erste Tail-Zeile wie eine Adresse wirkt, versuche zweite als Store
                addr_pat = re.compile(r"(?i)(straße|str\.|weg|allee|platz|gasse|\b\d{1,4}\b)")
                if len(tail) == 1:
                    return tail[0]
                if not addr_pat.search(tail[0]) and addr_pat.search(tail[1]):
                    return tail[0]
                # sonst nimm Kombination, aber ohne zu lang zu werden
                combo = _norm_ws(" ".join(tail[:2]))
                if len(combo) > 40:
                    return tail[0]
                return combo

    return None

def _guess_store_name_strict(text: str) -> str | None:
    """
    Robuste Erkennung aus dem Kopf: bekannte Ketten oder „volle“ Kopfzeile.
    """
    CHAINS = [
        r"\bLIDL\b", r"\bALDI\b", r"\bEDEKA\b", r"\bREWE\b", r"\bNORMA\b",
        r"\bNETTO\b", r"\bPENNY\b", r"\bDM\b", r"\bROSSMANN\b", r"\bAPOTHEKE\b",
        r"\bARAL\b", r"\bSHELL\b", r"\bJET\b", r"\bESSO\b", r"\bTOTAL\b",
        r"\bNIELSEN\b", r"\bSCAN-?SHOP\b",
        r"\bMELANCHTHONSTR(?:AßE|ASSE)?\b",  # Standort-Hinweis hilft dem Kontext
    ]

    lines = [ln.strip() for ln in (text or "").splitlines() if ln.strip()]
    head  = lines[:18]

    # 1) direkte Kettennamen
    head_join = "\n".join(head)
    for pat in CHAINS:
        m = re.search(pat, head_join, re.I)
        if m:
            return m.group(0).upper().strip()

    # 2) Zeilen mit viel Uppercase, aber ohne Beträge
    candidates = []
    for ln in head:
        if _is_amount_line(ln) or _looks_trash(ln):
            continue
        letters = [c for c in ln if c.isalpha()]
        if not letters:
            continue
        ups = sum(1 for c in letters if c.isupper())
        score = ups / max(1, len(letters))
        candidates.append((score, len(ln), ln))

    if candidates:
        candidates.sort(reverse=True)
        cand = _norm_ws(candidates[0][2])
        # Müll ausschließen (einzelne „Wörter“ mit Lücken)
        if re.search(r"^[A-ZÄÖÜ]{2,}(?:\s+[A-ZÄÖÜ]{2,})+$", cand) and len(cand) < 6:
            return None
        return cand

    return None

def _store_is_bad(s: str | None) -> bool:
    """
    Erkennung „schlechter“ Store-Namen (z.B. 'SHES Ste  FN', 'IN DER', zu kurz, fast leer).
    """
    if not s:
        return True
    t = _norm_ws(s)
    if len(t) < 4:
        return True
    if t.upper() in {"IN", "DER", "IN DER"}:
        return True
    # enthält viele Lücken mitten in Wörtern → OCR-Müll
    if re.search(r"[A-Za-zÄÖÜß]\s{2,}[A-Za-zÄÖÜß]", t):
        return True
    # nur Großbuchstabenblöcke mit vielen Einzelfragmenten
    parts = [p for p in t.split() if p]
    if len(parts) >= 3 and sum(len(p) <= 3 for p in parts) >= 2:
        return True
    return False


# =================== Vorverarbeitung ===================
# ---------- PDF-Helper ----------
def _is_pdf(path: str) -> bool:
    return str(path).lower().endswith(".pdf")

def pdf_extract_text_fitz(path: str, max_pages: int = 10) -> str:
    """Direkttext aus PDF lesen (A4-Rechnungen oft 1A als Text eingebettet)."""
    import fitz  # PyMuPDF
    try:
        doc = fitz.open(path)
        texts = []
        for i, page in enumerate(doc):
            if i >= max_pages:
                break
            # 'text' ist in der Regel schon sauber (keine OCR nötig)
            texts.append(page.get_text("text"))
        doc.close()
        return "\n".join(texts)
    except Exception:
        return ""

def pdf_render_pages_to_pngs(path: str, dpi: int = 300, max_pages: int = 5) -> list[str]:
    """PDF-Seiten als temporäre PNGs rendern (für OCR)."""
    import fitz, os, tempfile
    out_paths = []
    try:
        zoom = dpi / 72.0  # 72 dpi Basis in PDF
        doc = fitz.open(path)
        tmpdir = tempfile.mkdtemp(prefix="pdf_ocr_")
        for i, page in enumerate(doc):
            if i >= max_pages:
                break
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat, alpha=False)
            pth = os.path.join(tmpdir, f"page_{i+1}.png")
            pix.save(pth)
            out_paths.append(pth)
        doc.close()
    except Exception as e:
        print(f"⚠️ PDF-Rendering fehlgeschlagen: {e}")
    return out_paths

def _is_fuel_pdf(s: str) -> bool:
    """Erkennt typische Tankrechnungs-PDFs mit Tabellenkopf."""
    up = s.upper()
    hints = [
        "TANKRECHNUNG",
        "KARTEN-NR", "KARTENNR", "KUNDEN-NR", "RG.-NR", "RG.-DATUM",
        "ARTIKEL", "MENGE", "NETTOPREIS", "NETTOBETRAG", "STEUERBETRAG", "BRUTTOBETRAG",
        "ZG TANKSTELLE",  # dein Beispiel
        "MWST-SATZ", "MWST", "UST"
    ]
    return sum(1 for h in hints if h in up) >= 3


# ---------- Ende PDF-Helper ----------

def preprocess_image(image_path: str):
    import cv2
    img = _imread_safe(image_path)

    # 1) groß ziehen
    scale = 1.5
    img = cv2.resize(img, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)

    # 2) Graustufen
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    # 2a) Gelbstich rausziehen (Weißabgleich / Normalisierung)
    gray = cv2.normalize(gray, None, alpha=0, beta=255, norm_type=cv2.NORM_MINMAX)

    # 3) Rauschen filtern
    gray = cv2.fastNlMeansDenoising(gray, None, h=10, templateWindowSize=7, searchWindowSize=21)

    # 4) Kontrast anheben
    clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8,8))
    gray = clahe.apply(gray)

    # 5) Binärisieren
    _, bw = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

    # 6) kleine Flecken killen
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2,2))
    bw = cv2.morphologyEx(bw, cv2.MORPH_OPEN, kernel, iterations=1)

    return bw

def preprocess_yellow(img):
    # auf 1.5x skalieren
    scale = 1.5
    img2 = cv2.resize(img, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
    # Grünkanal
    g = img2[:,:,1]
    # normalisieren + CLAHE
    g = cv2.normalize(g, None, 0, 255, cv2.NORM_MINMAX)
    clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8,8))
    g = clahe.apply(g)
    # adaptiv + invert als Alternative testen
    bw1 = cv2.adaptiveThreshold(g, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 35, 8)
    bw2 = cv2.bitwise_not(bw1)
    return bw1, bw2

def infer_cash_from_amounts(best: dict, text: str) -> str | None:
    low = text.lower()
    given = best.get("Gegeben (€)")
    change = best.get("Wechselgeld (€)")
    if isinstance(given, (int, float)) and (isinstance(change, (int, float)) or "wechselgeld" in low or "zurück" in low):
        return "Bar"
    return None

def _float_de(s: str):
    """Deutsch formatierte Zahl -> float (auch 51,11 / 1.999,9 etc.)."""
    import re
    if s is None:
        return None
    s = s.strip()
    s = re.sub(r"[^\d,.\-]", "", s)
    # Tausenderpunkte entfernen, Komma -> Punkt
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None

def _find_value_after_label(lines, label_regex, window=3,
                            value_regex=r"\d{1,3}(?:[.,]\d{3})*[.,]\d{1,3}"):
    """
    Sucht eine Zeile mit Label (regex). Gibt erste passende Zahl auf derselben
    oder in den nächsten `window` Zeilen zurück.
    Return: (wert_str, index_der_wertzeile) oder (None, -1)
    """
    import re
    lab = re.compile(label_regex, re.IGNORECASE)
    val = re.compile(value_regex)
    for i, ln in enumerate(lines):
        if lab.search(ln):
            # gleiche Zeile zuerst
            m = val.search(ln)
            if m:
                return m.group(0), i
            # sonst Folgezeilen
            for j in range(1, window + 1):
                if i + j < len(lines):
                    m = val.search(lines[i + j])
                    if m:
                        return m.group(0), i + j
    return None, -1

import re

def _pick_total_candidate(text: str) -> float | None:
    money = r"(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})"
    kw = r"(zu\s*zahlen|endbetrag|gesamt|summe|betrag|total)"
    bad_line = re.compile(r"(?:TSE|UST|ID|TRANS|TERMINAL|000Z|T\d|Z$)", re.I)
    best = None
    for ln in text.splitlines():
        s = (ln or "").strip()
        if not s or bad_line.search(s): continue
        m = re.search(rf"(?i)\b{kw}\b[^\d]{{0,40}}{money}\s*(?:€|EUR)?", s)
        if not m:
            m = re.search(rf"{money}\s*(?:€|EUR)?\s*$", s)
        if m:
            val = _norm_money(m.group(m.lastindex))
            if 0 < val <= 2000:
                best = val
    return best

def _is_plausible_amount(v: float) -> bool:
    return (v is not None) and (0 < v <= 2000)
# =================== OCR Varianten ===================

# ======= Kachelung für sehr lange Bilder (AUTO) =======

# Optionales Debug: Kachelbereiche + kurze OCR-Vorschau ausgeben
DEBUG_TILES = False  # bei Bedarf True setzen

import re

NOISE_PATTERNS = [
    r"^-+K-?U-?N-?D-?E-?N-?B-?E-?L-?E-?G-+",
    r"^H+N+H+|^#\s*H+|^PAN\s#+",
    r"^[A-F0-9]{16,}$",
    r"^TSE|^Seriennr\.|^Prufwert|^SigZ|^TSE-",
    r"^\|+$",                          # nur senkrechte Balken
    r"^[\s\|\-_~=]{3,}$",              # reine Trennlinien
]


import re

def apply_payment_detection(best: dict, combo_text: str, debug: bool = False) -> None:
    # Nur setzen, wenn noch leer
    if not USE_PADDLE:
        if not best.get("Zahlung"):
            pay = detect_payment_method(combo_text)

            # Heuristik für Bar, wenn Gegeben/Wechselgeld plausibel
            if not pay:
                given  = best.get("Gegeben (€)")
                change = best.get("Wechselgeld (€)")
                low    = combo_text.lower()
                if isinstance(given, (int, float)) and (
                    isinstance(change, (int, float)) or "wechselgeld" in low or "zurück" in low
                ):
                    pay = "Bar"
        # alte Fallbacks / Merge-Korrekturen
            best["Zahlung"] = pay if pay else "Unbekannt"

    # Debug-Prints: exakt 1x pro Beleg
    if debug and DEBUG_PAYMENTS and not getattr(apply_payment_detection, "_did_log", False):
        apply_payment_detection._did_log = True  # Sperre setzen


def _fix_word_digit_confusion(s: str) -> str:
    """
    Korrigiert 0↔O, 1↔l/I kontextsensitiv.
    - In 'zahlenartigen' Tokens (mehr Ziffern als Buchstaben): O/o→0, I/l→1, °→0, •→., ‚’→,
    - In 'wortartigen' Tokens (>= ebenso viele Buchstaben): 0→o/O, 1→l/I (case-bewusst)
    """
    def fix_tok(tok: str) -> str:
        if not tok: return tok
        letters = sum(c.isalpha() for c in tok)
        digits  = sum(c.isdigit() for c in tok)
        is_numberish = digits >= letters and digits > 0

        if is_numberish:
            t = tok.replace("O","0").replace("o","0").replace("I","1").replace("l","1")
            t = t.replace("°","0").replace("•",".").replace("‚",",").replace("’",",")
            return t

        # Wortartig → Ziffern wieder zu Buchstaben, case-bewusst
        # Groß-/Kleinschreibung beibehalten
        def sub_word(m):
            ch = m.group(0)
            if ch == "0": return "O" if tok.isupper() else "o"
            if ch == "1": 
                # in Großwörtern ist meist 'I' korrekt, sonst 'l'
                return "I" if tok.isupper() else "l"
            return ch

        return re.sub(r"[01]", sub_word, tok)

    # auf Wort-/Zahl-Token anwenden (inkl. Umlaute)
    return re.sub(r"[A-Za-zÄÖÜäöüß0-9°•‚’]+", lambda m: fix_tok(m.group(0)), s)


def _fix_common_terms(s: str) -> str:
    import re
    # Gesellschaftsform / Schreibweisen
    s = re.sub(r"(?i)\b0HG\b", "OHG", s)
    s = re.sub(r"(?i)\bCHG\b", "OHG", s)

    # Umlaute/Orte/Typische Wörter
    s = re.sub(r"\b1N\b", "IN", s)  # "1N" -> "IN"
    s = re.sub(r"(?i)ap0theke", "Apotheke", s)
    s = re.sub(r"(?i)r0mer", "römer", s)
    s = s.replace("Rémer", "Römer")

    # Netto/Brutto-Fehllesungen
    s = re.sub(r"(?i)\bnett0\b", "Netto", s)
    s = re.sub(r"(?i)\bbrutt0\b", "Brutto", s)

    # MwSt-Fehllesungen
    s = re.sub(r"(?i)\bmust\s*y\b", "MwSt", s)
    s = re.sub(r"(?i)\bmust\b", "MwSt", s)

    # Datums-Fehllesung: 07.08,25 -> 07.08.25
    s = re.sub(r"\b(\d{2}\.\d{2}),(\d{2})\b", r"\1.\2", s)

    # OCR-Fehler rund um MwSt/Netto/Brutto
    repl = {
        r"\bmist\b": "mwst",
        r"\bnist\b": "mwst",
        r"\bmw5t\b": "mwst",
        r"\bmwst\b": "mwst",   # vereinheitlichen
        r"\bmetta\b": "netto",
        r"\bmetta?\b": "netto",
        r"\bnetto\b": "netto",
        r"\bbrutt0\b": "brutto",
        r"\bbrutt\b": "brutto",
    }
    t = s
    for pat, rep in repl.items():
        t = re.sub(pat, rep, t, flags=re.IGNORECASE)
    return t
    return s

def _guess_store_from_header(text: str) -> str | None:
    """
    Sucht im Kopf (erste ~30 Zeilen) nach 'SCAN-SHOP'/'SCAN SHOP' und
    kombiniert optional eine Besitzer-/Markenzeile darüber: 'Nielsen SCAN-SHOP'.
    """
    import re
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()][:30]
    lines = [_fix_spaced_letters(ln) for ln in lines]  # entgappen

    # 1) Direkt: 'Nielsen ... SCAN SHOP' in EINER Zeile
    head = "\n".join(lines)
    m = re.search(r'(?i)\b(nielsen)\b.*?\b(scan[\s\-]?shop)\b', head)
    if m:
        return "Nielsen SCAN-SHOP"

    # 2) 'SCAN SHOP' finden und Besitzername in den 3 Zeilen darüber abgreifen
    idx = None
    for i, ln in enumerate(lines):
        if re.search(r'(?i)\bscan[\s\-]?shop\b', ln):
            idx = i
            break
    if idx is None:
        return None

    # heuristische Besitzer-/Markenzeile 1–3 Zeilen darüber
    owner = None
    for j in range(max(0, idx-3), idx):
        cand = re.sub(r'[^A-Za-zÄÖÜäöüß0-9 .&/\-]', '', lines[j]).strip()
        # Ausschluss von typischen „technischen“/Adresse-Schlagworten
        if re.search(r'(?i)\b(ust|steuernr|ust\-id|straße|str\.|tel|fax|www|gmbh|ug|ohg|kg|ag)\b', cand):
            continue
        if sum(c.isalpha() for c in cand) >= 4 and 1 <= len(cand.split()) <= 3:
            owner = cand
            break

    if owner:
        return f"{owner} SCAN-SHOP"
    return "SCAN-SHOP"


def _normalize_footer_lines(s: str) -> str:
    # Pipes entschärfen & Trennlinien entfernen
    s = re.sub(r"[|]+", " | ", s)
    s = re.sub(r"(?m)^[\s\|\-_=~]{3,}$", "", s)
    s = re.sub(r"(?m)^\s*['‘`´]\s*", "", s)  # führendes Quote weg
    # Kopfbegriffe
    s = re.sub(r"(?i)\bNett0\b", "Netto", s)
    s = re.sub(r"(?i)\bBrutt0\b", "Brutto", s)
    s = re.sub(r"(?i)\bMust\s*y\b", "MwSt", s)
    s = re.sub(r"(?i)\bMust\b", "MwSt", s)
    # "Gesamt/Betrag/Summe ... | 8.85" -> "Gesamt 8,85 EUR"
    s = re.sub(r"(?mi)\b(Gesamt|Betrag|Summe)\b[^\d]{0,30}(?:EUR|€)?[^\d]{0,12}\|[^\d]{0,12}"
               r"(\d{1,3}(?:[.,]\d{3})*[.,]\d{1,2})",
               lambda m: f"{m.group(1)} {m.group(2).replace('.', ',')} EUR", s)
    # Falls ohne EUR
    s = re.sub(r"(?mi)\b(Gesamt|Betrag|Summe)\b[^\d]{0,20}"
               r"(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})(?!\s*(EUR|€))",
               lambda m: f"{m.group(1)} {m.group(2).replace('.', ',')} EUR", s)
    # Gegeben/Zurück ohne EUR -> EUR anhängen
    s = re.sub(r"(?mi)\b(Gegeben(?:\s+Bar)?|Zurück)\b[^\d]{0,24}"
               r"(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})(?!\s*(EUR|€))",
               lambda m: f"{m.group(1)} {m.group(2).replace('.', ',')} EUR", s)
    # generisch: Zahl.Zwei -> Zahl, Zwei
    s = re.sub(r"(\d+)\.(\d{2})(?=\D|$)", r"\1,\2", s)
    return s

# (optional) harter Fallback für Ausreißer im Fuß
def force_fix_footer_lines(s: str) -> str:
    out = []
    for ln in s.splitlines():
        L = ln
        L = re.sub(r"[|]+", " | ", L)
        L = re.sub(r"(?i)\bNett0\b", "Netto", L)
        L = re.sub(r"(?i)\bBrutt0\b", "Brutto", L)
        L = re.sub(r"(?i)\bMust\s*y\b", "MwSt", L)
        L = re.sub(r"(?i)\bMust\b", "MwSt", L)
        if re.search(r"(?i)\b(Gesamt|Betrag|Summe)\b", L):
            m = re.search(r"([0-9][0-9 .•°]*[.,][0-9]{1,2})", L)
            if m:
                amt = m.group(1).replace("•",".").replace("°","0")
                amt = re.sub(r"(?<=\d)\.(?=\d{2}\b)", ",", amt)
                key = re.search(r"(?i)\b(Gesamt|Betrag|Summe)\b", L).group(1).capitalize()
                L = f"{key} {amt} EUR"
        elif re.search(r"(?i)\b(Gegeben(?:\s+Bar)?|Zurück)\b", L):
            m = re.search(r"([0-9][0-9 .•°]*[.,][0-9]{1,2})(?!\s*(EUR|€))", L)
            if m:
                amt = m.group(1).replace("•",".").replace("°","0")
                amt = re.sub(r"(?<=\d)\.(?=\d{2}\b)", ",", amt)
                key = re.search(r"(?i)\b(Gegeben(?:\s+Bar)?|Zurück)\b", L).group(1)
                L = f"{key} {amt} EUR"
        out.append(L)
    return "\n".join(out)

def _drop_noise_lines(lines: list[str]) -> list[str]:
    NOISE_PATTERNS = [
        r"^-+K-?U-?N-?D-?E-?N-?B-?E-?L-?E-?G-+",
        r"^H+N+H+|^#\s*H+|^PAN\s#+",
        r"^[A-F0-9]{16,}$",
        r"^TSE|^Seriennr\.|^Prufwert|^SigZ|^TSE-",
        r"^\|+$",
        r"^[\s\|\-_~=]{3,}$",
    ]
    out = []
    for ln in lines:
        s = ln.strip()
        if not s:
            continue
        non_alnum_ratio = (sum(1 for c in s if not c.isalnum()) / max(1, len(s)))
        if non_alnum_ratio > 0.6 and not re.search(r"(EUR|€|\d[.,]\d{2})", s, re.I):
            continue
        if any(re.search(p, s, re.I) for p in NOISE_PATTERNS):
            continue
        out.append(ln)
    return out

from difflib import SequenceMatcher

def _dedupe_consecutive_lines(text: str, sim_threshold=0.94) -> str:
    lines = _drop_noise_lines([ln.rstrip() for ln in text.splitlines()])
    out, prev = [], ""
    for ln in lines:
        s1 = re.sub(r"\s+", " ", ln).strip()
        s2 = re.sub(r"\s+", " ", prev).strip()
        sim = SequenceMatcher(None, s1, s2).ratio() if s1 and s2 else 0.0
        if sim < sim_threshold:
            out.append(ln); prev = ln
    return "\n".join(out)

def _collapse_repeated_blocks(text: str, window=8, sim_threshold=0.96) -> str:
    lines = _drop_noise_lines([ln.rstrip() for ln in text.splitlines()])
    i, out = 0, []
    while i < len(lines):
        block1 = lines[i:i+window]
        if i + window >= len(lines):
            out.extend(block1); break
        block2 = lines[i+window:i+2*window]
        s1 = "\n".join(re.sub(r"\s+"," ",x).strip() for x in block1)
        s2 = "\n".join(re.sub(r"\s+"," ",x).strip() for x in block2)
        sim = SequenceMatcher(None, s1, s2).ratio()
        if sim >= sim_threshold:
            out.extend(block1); i += window*2
        else:
            out.append(lines[i]); i += 1
    return "\n".join(out)

def _normalize_dates(s: str) -> str:
    import re
    s = re.sub(r"\b(\d{2}),(\d{2}),(\d{2})\b", r"\1.\2.\3", s)      # 07,08,25 → 07.08.25
    s = re.sub(r"\b(\d{2}\.\d{2}),(\d{2})\b", r"\1.\2", s)          # 07.08,25 → 07.08.25
    return s

def _strip_line_leaders(s: str) -> str:
    import re
    # entfernt führende » | > • : und überflüssige Spaces je Zeile
    return "\n".join(re.sub(r"^[\s\|\>\u00BB\u2022»«·•:]+", "", ln) for ln in s.splitlines())


def post_ocr_cleanup(text: str) -> str:
    if text is None: return ""
    if not isinstance(text, str): text = str(text)
    t = _fix_word_digit_confusion(text)
    t = _fix_common_terms(t)
    t = _strip_line_leaders(t)    # <<< NEU: führt » | > etc. ab
    t = _normalize_dates(t)
    t = _soft_clean(t)
    t = _normalize_money_tokens(t)
    t = _normalize_footer_lines(t)
    t = force_fix_footer_lines(t)
    t = _dedupe_consecutive_lines(t)
    t = _collapse_repeated_blocks(t)
    return t

# ---------------------------------------------------------
# Hilfsfunktionen & Debug-Flag für ocr_text_tiled
# ---------------------------------------------------------

# globale Debug-Flag
DEBUG_TILES = False   # auf True setzen, um jede Kachel-Preview im Terminal zu sehen

import cv2, os


def _tile_image_vert(img, tile_h=2400, overlap=200):
    """
    Zerteilt ein Bild vertikal in überlappende Kacheln.
    Liefert Liste von (y0,y1)-Koordinaten.
    """
    h = img.shape[0]
    tiles = []
    y = 0
    while y < h:
        y0 = y
        y1 = min(y + tile_h, h)
        tiles.append((y0, y1))
        if y1 == h:
            break
        y = y1 - overlap  # überlappend weiterrücken
    return tiles


def ocr_top_header_only(img_bgr) -> str | None:
    """
    Nimmt oberste 18% als ROI und OCRt nur Buchstaben/Leerzeichen.
    Gut gegen „z n  n m“-Artefakte.
    """
    import cv2, pytesseract
    h, w = img_bgr.shape[:2]
    top = img_bgr[0:int(h*0.18), :]
    g = cv2.cvtColor(top, cv2.COLOR_BGR2GRAY)
    g = cv2.GaussianBlur(g, (3,3), 0)
    _, bw = cv2.threshold(g, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    cfg = r'--oem 1 --psm 7 -c preserve_interword_spaces=1 -c user_defined_dpi=350 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZÄÖÜabcdefghijklmnopqrstuvwxyzäöü .-/&'
    try:
        t = pytesseract.image_to_string(bw, lang="deu", config=cfg)
        return t.strip() or None
    except Exception:
        return None


def ocr_text_tiled(image_path: str, use_alt: bool = False) -> list[str]:
    """
    Kachelt sehr lange Belege und OCR't jede Kachel.
    use_alt=True: alternative, etwas aggressivere Vorverarbeitung.
    """
    import cv2, pytesseract, numpy as np

    img = _imread_safe(image_path, cv2.IMREAD_COLOR)
    if img is None:
        return []

    H, W = img.shape[:2]
    # Größere Kacheln für App-Lidl-Bons; mehr Überlapp
    tile_h = min(2600, max(1200, int(H * 0.35)))  # 35% der Höhe, gedeckelt
    stride = max(600, int(tile_h * 0.78))         # ~22% Überlapp

    texts = []
    y = 0
    while y < H:
        y2 = min(H, y + tile_h)
        tile = img[y:y2, :]

        gray = cv2.cvtColor(tile, cv2.COLOR_BGR2GRAY)
        if use_alt:
            gray = cv2.bilateralFilter(gray, d=7, sigmaColor=50, sigmaSpace=50)
            bw   = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                         cv2.THRESH_BINARY, 31, 8)
        else:
            gray = cv2.GaussianBlur(gray, (3,3), 0)
            _, bw = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

        cfg = r'--oem 3 --psm 6 -c preserve_interword_spaces=1'

        # >>> defensiv: tx immer initialisieren, dann bereinigen
        tx = ""
        try:
            tx = pytesseract.image_to_string(bw, lang="deu+eng", config=cfg)
        except Exception as te:
            print(f"⚠️ Tesseract-Fehler (Kachel): {te}")
        if isinstance(tx, bytes):
            tx = tx.decode("utf-8", "ignore")
        tx = (tx or "").strip()
        if tx:
            texts.append(safe_post_ocr_cleanup(tx))
        # <<<

        if y2 == H:
            break
        y = y + stride

    return texts


# ======= Ende Kachelung =======

import traceback, re

def safe_post_ocr_cleanup(text: str) -> str:
    try:
        return post_ocr_cleanup(text)
    except Exception as e:
        print("🔥 Cleanup-Fehler:", e)
        traceback.print_exc()
        # notfalls Rohtext zurückgeben, damit der Lauf weitergeht
        return text

def debug_cleanup_chain(sample: str):
    """Lokalisiert, in welcher Stufe ein Fehler passiert."""
    steps = [
        ("_fix_word_digit_confusion", _fix_word_digit_confusion),
        ("_fix_common_terms",         _fix_common_terms),
        ("_soft_clean",               _soft_clean),
        ("_normalize_money_tokens",   _normalize_money_tokens),
        ("_normalize_footer_lines",   _normalize_footer_lines),
        ("force_fix_footer_lines",    force_fix_footer_lines),
        ("_dedupe_consecutive_lines", _dedupe_consecutive_lines),
        ("_collapse_repeated_blocks", _collapse_repeated_blocks),
    ]
    t = sample
    for name, fn in steps:
        try:
            t = fn(t)
        except Exception as e:
            print(f"❌ Fehler in {name}: {e}")
            traceback.print_exc()
            return
    print("✅ cleanup chain OK")
    return t


def detect_payment_method(text: str) -> str | None:
    low = text.lower()
    for method, keys in PAYMENT_KEYWORDS.items():
        if any(k in low for k in keys):
            return method
    # Heuristik: Wenn explizit "Gegeben" und "Zurück" vorkommt, ist es meistens Bar
    if ("gegeben" in low or "gezahlt" in low) and ("zurück" in low or "wechselgeld" in low):
        return "Bar"
    return None

def _detect_payment(text: str) -> str | None:
    t = text.lower()
    # Klartexte ...
    if re.search(r"\bbar(?:zahlung)?\b|cash\b", t): return "Bar"
    if re.search(r"\b(girocard|ec-?karte|maestro|lastschrift)\b", t): return "Girocard"
    if re.search(r"\bkreditkarte\b|\bvisa\b|\bmastercard\b|\bamex\b", t): return "Kreditkarte"
    # Wallets / App
    if re.search(r"\blidl\s*pay\b", t): return "Lidl Pay"
    if re.search(r"\bapple\s*pay\b", t): return "Apple Pay"
    if re.search(r"\bgoogle\s*pay\b|\bgpay\b", t): return "Google Pay"
    if re.search(r"\bpaypal\b", t): return "PayPal"
    # Hinweise aus Terminal-Block:
    if re.search(r"\bkontaktlos\b|\bnfc\b|\bchip\b|emv", t) and ("bar" not in t):
        return "Karte"
        # Lidl Pay – auch falsche OCR-Varianten abfangen
    if re.search(r"\bL[I1]DL\s*PAY\b", t) or re.search(r"\bCN\s*PAY\b", t) or "LIDL PAY" in t:
        return "Lidl Pay"
    if re.search(r"\bGIRO?CARD\b", t) or "KARTENZAHLUNG" in t or "KARTE" in t:
        return "Karte"
    if "BAR" in t or "CASH" in t:
        return "Bar"
    if "LASTSCHRIFT" in t or "SEPA" in t or "MANDAT" in t:
        return "Lastschrift"

    return None

def majority_payment(texts: list[str]) -> str | None:
    from collections import Counter
    votes = []
    for tx in texts:
        pay = _detect_payment(tx)
        if pay:
            votes.append(pay)
    if not votes:
        return None
    best, cnt = Counter(votes).most_common(1)[0]
    # Schwelle: wenn mehrere Stimmen oder nur 1 eindeutige
    return best if cnt >= 1 else None


# ==== Helper: long-receipt-Heuristik ====
def _is_long_receipt(img) -> bool:
    h, w = img.shape[:2]
    return (h / max(1, w) >= 2.8) or (h >= 4800)

# ==== Helper: robustes Einlesen ====
def _imread_safe(path, flags=cv2.IMREAD_COLOR):
    import numpy as np, cv2
    # 1) imdecode (Pfade mit Umlauten)
    try:
        data = np.fromfile(path, dtype=np.uint8)
        img  = cv2.imdecode(data, flags)
        if img is not None:
            return img
    except:
        pass
    # 2) klassisches imread
    try:
        img = cv2.imread(path, flags)
        if img is not None:
            return img
    except:
        pass
    # 3) TIFF-Fallback via PIL (nutzt globalen Image-Import!)
    try:
        pil = Image.open(path)
        # nur erste Seite verwenden
        if getattr(pil, "n_frames", 1) > 1:
            pil.seek(0)
        if pil.mode not in ("RGB", "RGBA", "L"):
            pil = pil.convert("RGB")
        elif pil.mode == "RGBA":
            pil = pil.convert("RGB")
        arr = np.array(pil)  # RGB
        if arr.ndim == 2:
            arr = cv2.cvtColor(arr, cv2.COLOR_GRAY2BGR)
        img = cv2.cvtColor(arr, cv2.COLOR_RGB2BGR)
        return img
    except:
        return None

# ==== Helper: Orientierung ausgleichen (OSD) ====
def deskew_and_orient(img):
    import cv2, pytesseract
    try:
        osd = pytesseract.image_to_osd(img)
        rot = 0
        if "Rotate: 90" in osd:   rot = 90
        elif "Rotate: 180" in osd: rot = 180
        elif "Rotate: 270" in osd: rot = 270
        if rot == 90:
            img = cv2.rotate(img, cv2.ROTATE_90_CLOCKWISE)
        elif rot == 180:
            img = cv2.rotate(img, cv2.ROTATE_180)
        elif rot == 270:
            img = cv2.rotate(img, cv2.ROTATE_90_COUNTERCLOCKWISE)
    except Exception:
        pass
    return img

# ==== Helper: mehrere Footer-Zonen testen ====
def ocr_bottom_variants(image_path: str) -> list[str]:
    import cv2, pytesseract, numpy as np
    img = _imread_safe(image_path, cv2.IMREAD_COLOR)
    if img is None:
        return []
    h, w = img.shape[:2]
    # Bänder: auch knapp ÜBER dem eigentlichen Fuß
    bands = [
        (0.12, 0.22),
        (0.18, 0.22),
        (0.25, 0.22),
        (0.38, 0.24),
        (0.50, 0.26),  # Mitte
        (0.62, 0.28),
        (0.74, 0.24),
    ]

    out = []
    for start, height in bands:
        y0 = max(0, int(h * (1.0 - (start + height))))
        y1 = min(h, int(h * (1.0 - start)))
        crop = img[y0:y1, 0:w]
        if crop.size == 0:
            continue
        gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
        gray = cv2.GaussianBlur(gray, (3,3), 0)
        bw   = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                     cv2.THRESH_BINARY, 31, 8)
        bw   = 255 - bw
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2,2))
        bw    = cv2.morphologyEx(bw, cv2.MORPH_OPEN, kernel, iterations=1)
        try:
            txt = pytesseract.image_to_string(
                bw, lang="deu+eng",
                config=r'--oem 3 --psm 6 -c preserve_interword_spaces=1'
            )
            if txt and txt.strip():
                out.append(txt)
        except Exception:
            pass
    return out

def preprocess_app_receipt_strong(img_bgr):
    """Für App-/Screenshot-Bons: stark kontrastieren + 'dicker' machen."""
    import cv2
    g = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)

    # Unsharp mask
    blur = cv2.GaussianBlur(g, (0,0), sigmaX=1.2)
    sharp = cv2.addWeighted(g, 1.6, blur, -0.6, 0)

    # CLAHE für lokalen Kontrast
    clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8,8))
    eq = clahe.apply(sharp)

    # Binär + Otsu
    _, bw = cv2.threshold(eq, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

    # leicht verdicken
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2,2))
    bw_bold = cv2.morphologyEx(bw, cv2.MORPH_DILATE, kernel, iterations=1)
    return bw_bold

def _find_total_keyword(s: str) -> bool:
    import re
    # a) Schlüsselwörter
    if re.search(r"(?i)\b(betrag|summe|gesamt|zu\s*zahlen)\b", s):
        return True
    # b) irgendein Geldbetrag mit €/EUR – hilft bei Lidl, wenn nur "… EUR" steht
    if re.search(r"(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})\s*(?:€|eur)\b", s, re.I):
        return True
    return False

def _fix_spaced_letters(line: str) -> str:
    """
    Verbindet gesperrte Wörter wie 'N O R M A' → 'NORMA' oder 'L I D L' → 'LIDL'.
    Wir fassen NUR Sequenzen aus >=3 Ein-Buchstaben-Tokens zusammen.
    """
    import re
    def join_match(m: re.Match) -> str:
        s = m.group(0)
        return s.replace(" ", "")
    # Beispiel: "N O R M A" oder "L I D L"
    return re.sub(r'\b(?:[A-Za-zÄÖÜäöü]\s+){2,}[A-Za-zÄÖÜäöü]\b', join_match, line)

def _brand_from_anywhere(text: str) -> str | None:
    """
    Sucht bekannte Ketten im gesamten OCR-Text und gibt einen kanonischen Namen zurück.
    """
    low = text.lower()
    brand_map = {
        "lidl": "Lidl",
        "norma": "NORMA",
        "aldi": "ALDI",
        "rewe": "REWE",
        "edeka": "EDEKA",
        "kaufland": "Kaufland",
        "penny": "PENNY",
        "dm ": "dm", " dm": "dm", " dm-": "dm",
        "rossmann": "ROSSMANN",
        "hagebau": "hagebau",
        "obi": "OBI",
        "bauhaus": "BAUHAUS",
        "toom": "toom",
        "apotheke": "Apotheke",
        "aral": "ARAL", "shell": "Shell", "esso": "Esso", "jet": "JET",
        "bft": "bft", "total": "TOTAL",
        "scan-shop": "SCAN-SHOP", "scan shop": "SCAN-SHOP",
    }
    for key, canon in brand_map.items():
        if key in low:
            return canon
    return None

# ---- Laden aus dem Kopf (erste Zeilen) robust raten ----
STORE_KEYWORDS = [
    r"LIDL", r"ALDI", r"NORMA", r"EDEKA", r"KAUFLAND",
    r"REWE", r"DM\b", r"ROSSMANN", r"APOTHEKE",
    r"ARAL", r"SHELL", r"JET", r"TOTAL", r"ESSO",
    r"NIELSEN", r"SCAN-?SHOP"  # neu für deinen Fall
]
_store_pat = re.compile(r"(?i)(" + "|".join(STORE_KEYWORDS) + r")")

def _guess_store_name_head(text: str, top_n: int = 12) -> str | None:
    if not text: return None
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    head = lines[:top_n]
    CHAINS = [r"\bLIDL\b", r"\bALDI\b", r"\bEDEKA\b", r"\bREWE\b", r"\bNORMA\b",
              r"\bNETTO\b", r"\bPENNY\b", r"\bDM\b", r"\bROSSMANN\b",
              r"\bAPOTHEKE\b", r"\bNIELSEN\b", r"\bSCAN-?SHOP\b", r"\bKAUFLAND\b"]
    store_pat = re.compile("|".join(CHAINS), re.I)
    money_pat = re.compile(r"\b\d{1,3}(?:[.,]\d{3})*[.,]\d{2}\s*(?:€|EUR)?\b", re.I)
    for ln in head:
        if _looks_like_noise(ln): continue
        if store_pat.search(ln) and any(k in ln.upper() for k in ("SCAN","SHOP","NIELSEN","LIDL")):
            return ln.strip()
    for ln in head:
        if _looks_like_noise(ln): continue
        if store_pat.search(ln): return ln.strip()
    head_filtered = [ln for ln in head if not _looks_like_noise(ln)]
    if head_filtered:
        def _upper_score(s):
            letters=[c for c in s if c.isalpha()]
            if not letters: return 0.0
            return sum(1 for c in letters if c.isupper())/len(letters)
        cand = sorted(head_filtered, key=lambda s: (_upper_score(s), len(s)), reverse=True)[0]
        if not money_pat.search(cand): return cand.strip()
    return None

def _is_bad_store_name(name: str) -> bool:
    """Heuristik: offensichtlicher OCR-Müll wie 'SHES Ste FN' zurückweisen."""
    if not name:
        return True
    s = name.strip()
    if len(s) < 4:
        return True
    # Kein Geld/Keywords
    if re.search(r"\d|EUR|€", s, re.I):
        return True
    # Wenige Vokale -> oft zerpflückte Großbuchstaben
    vowels = len(re.findall(r"[AEIOUÄÖÜaeiouäöü]", s))
    letters = len(re.findall(r"[A-Za-zÄÖÜäöü]", s))
    if letters >= 6 and vowels <= 1:
        return True
    # Viele Ein-Buchstaben-Tokens
    tokens = [t for t in s.split() if t]
    short_tokens = sum(1 for t in tokens if len(t) == 1)
    if short_tokens >= 2 and len(tokens) <= 4:
        return True
    # Sehr viele doppelte Spaces (zerhackt)
    if re.search(r"[A-Za-z]\s{2,}[A-Za-z]", s):
        return True
    return False


def ocr_bruteforce_totals(image_path: str) -> list[str]:
    """
    Probiert verschiedene Skalierungen/Binarisierungen/PSM-Kombis und sammelt
    nur Texte, die typische Total-Schlüsselwörter enthalten.
    """
    import cv2, pytesseract, numpy as np
    out = []

    img = _imread_safe(image_path, cv2.IMREAD_COLOR)
    if img is None:
        return out

    # Kandidaten-Skalierungen (Breite)
    widths = [1200, 1600, 2000, 2400]

    # Binärisierungen als kleine Lambdas
    def bin_otsu(g):
        _, bw = cv2.threshold(g, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        return bw
    def bin_adapt(g):
        return cv2.adaptiveThreshold(g,255,cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                     cv2.THRESH_BINARY,31,9)
    bin_funcs = [bin_otsu, bin_adapt]

    # PSM-Configs
    cfgs = [
        r'--oem 1 --psm 6  -c preserve_interword_spaces=1 -c user_defined_dpi=350',
        r'--oem 1 --psm 4  -c preserve_interword_spaces=1 -c user_defined_dpi=350',
        r'--oem 1 --psm 3  -c preserve_interword_spaces=1 -c user_defined_dpi=350',  # NEU
        r'--oem 1 --psm 11 -c preserve_interword_spaces=1 -c user_defined_dpi=350',
        r'--oem 1 --psm 13 -c preserve_interword_spaces=1 -c user_defined_dpi=350',
    ]

    h0, w0 = img.shape[:2]
    for W in widths:
        scale = W / max(1, w0)
        im = cv2.resize(img, (int(w0*scale), int(h0*scale)),
                        interpolation=cv2.INTER_CUBIC if scale>1 else cv2.INTER_AREA)

        g = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
        # Schärfen + CLAHE
        blur  = cv2.GaussianBlur(g, (0,0), 1.2)
        sharp = cv2.addWeighted(g, 1.6, blur, -0.6, 0)
        clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8,8))
        eq    = clahe.apply(sharp)

        for bf in bin_funcs:
            bw = bf(eq)
            # Schrift etwas dicker
            kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2,2))
            bw_b  = cv2.morphologyEx(bw, cv2.MORPH_DILATE, kernel, iterations=1)

            for cfg in cfgs:
                # nur Deutsch – App-Bons sind deutsch
                try:
                    t = pytesseract.image_to_string(bw_b, lang="deu", config=cfg)
                    t = t.strip()
                    if t and _find_total_keyword(t):
                        out.append(t)
                except Exception:
                    pass

    # Deduplizieren, Reihenfolge beibehalten
    seen, uniq = set(), []
    for t in out:
        if t not in seen:
            seen.add(t); uniq.append(t)
    return uniq


def _dedupe_strs(items: list[str]) -> list[str]:
    """Dedupe bei Strings; ignoriert Nicht-Strings vorher (via _flatten_texts)."""
    flat = _flatten_texts(items)
    seen = set()
    uniq = []
    for s in flat:
        # s ist garantiert str
        if s in seen:
            continue
        seen.add(s)
        uniq.append(s)
    return uniq

def _money_regex() -> str:
    return r"(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})"

def _sanitize_store_name(store: str | None, head_text: str) -> str | None:
    import re
    if not store:
        return None
    s = store.strip()
    # harte Blacklist – häufige Fehlgriffe aus dem Fuß
    BAD = {
        "TOTAL","FATAL","SUMME","BETRAG","EINKAUFSWERT","GEGEBEN","ZURÜCK",
        "GIROCARD","BAR","NETTO","BRUTTO","MWST","UST"
    }
    if s.upper() in BAD or re.search(r"(€|EUR|\d)", s, re.I):
        # neu vom Kopf schätzen
        guess = _guess_store_name(head_text)
        return guess if guess else None
    return s

# =================== OCR-Varianten (DROP-IN) ===================
def ocr_text_multi(image_path: str) -> list[str]:
    """
    Liefert mehrere OCR-Textvarianten für einen Beleg:
    - Standard & alternative Vorverarbeitung
    - Kachel-OCR für lange Bons (standard + adaptiv)
    - Bold-Variante (dünne App-Schrift)
    - Globaler PSM-Sweep (11/13)
    - Gelbstich-Variante (Grünkanal, adaptiv, invert)
    - Mehrere Footer-ROIs
    """
    import cv2, pytesseract

    # Konfigs robust beziehen (mit Defaults, falls global nicht gesetzt)
    cfg_main = globals().get("CUSTOM_CONFIG_MAIN", r'--oem 3 --psm 6 -c preserve_interword_spaces=1')
    cfg_alt  = globals().get("CUSTOM_CONFIG_ALT",  r'--oem 3 --psm 4')
    fast     = bool(globals().get("FAST", True))

    txts: list[str] = []

    # === Bild laden ===
    img = _imread_safe(image_path, cv2.IMREAD_UNCHANGED)
    if img is None:
        print(f"⚠️ Bild konnte nicht geladen werden: {image_path}")
        return []

    # Alphakanal (BGRA) → BGR
    if len(img.shape) == 3 and img.shape[2] == 4:
        img = cv2.cvtColor(img, cv2.COLOR_BGRA2BGR)

    # Orientierung / Größe normalisieren
    img = deskew_and_orient(img)
    h, w = img.shape[:2]
    target_w = 1500
    if w < 1000 or w > 2200:
        scale = target_w / max(1, w)
        img = cv2.resize(img, (int(w*scale), int(h*scale)),
                         interpolation=cv2.INTER_AREA if scale < 1 else cv2.INTER_CUBIC)

    # 1) Lange Bons kacheln
    try:
        if _is_long_receipt(img):
            # Standard-Kachelung
            for tx in ocr_text_tiled(image_path, use_alt=False):
                if not tx:
                    continue
                if isinstance(tx, bytes):
                    tx = tx.decode("utf-8", "ignore")
                tx = tx.strip()
                if tx:
                    txts.append(safe_post_ocr_cleanup(tx))
            # Adaptive Kachelung
            if not fast:
                for tx in ocr_text_tiled(image_path, use_alt=True):
                    if not tx:
                        continue
                    if isinstance(tx, bytes):
                        tx = tx.decode("utf-8", "ignore")
                    tx = tx.strip()
                    if tx:
                        txts.append(safe_post_ocr_cleanup(tx))
    except Exception as e:
        print(f"⚠️ Kachel-OCR übersprungen: {e}")

    # 1b) Lidl-/App-Variante: starkes Preprocessing + nur Deutsch
    try:
        bw_app = preprocess_app_receipt_strong(img)
        t_app = pytesseract.image_to_string(
            bw_app, lang="deu",
            config=r'--oem 1 --psm 6 -c preserve_interword_spaces=1 -c user_defined_dpi=350'
        )
        if t_app.strip():
            txts.append(safe_post_ocr_cleanup(t_app))
    except Exception as e:
        print(f"⚠️ App-Variante fehlgeschlagen: {e}")
        
    # 2) Standard-Vorverarbeitung (ganzer Bon)
    try:
        p1 = preprocess_image(image_path)  # sollte ein BW-/Graubild liefern
        t1 = pytesseract.image_to_string(p1, lang="deu+eng", config=cfg_main)
        if t1.strip():
            txts.append(safe_post_ocr_cleanup(t1))
    except Exception as e:
        print(f"⚠️ Standard-OCR fehlgeschlagen: {e}")

    # 2a) Bold-Variante (dünne App-Schrift "verdicken")  <<< HIER kommt bw_bold >>>
    try:
        g = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        g = cv2.GaussianBlur(g, (3,3), 0)
        _, bw = cv2.threshold(g, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2,2))
        bw_bold = cv2.morphologyEx(bw, cv2.MORPH_DILATE, kernel, iterations=1)

        t_bold = pytesseract.image_to_string(
            bw_bold, lang="deu+eng",
            config=r'--oem 1 --psm 6 -c preserve_interword_spaces=1 -c user_defined_dpi=350'
        )
        if t_bold.strip():
            txts.append(safe_post_ocr_cleanup(t_bold))
    except Exception as e:
        print(f"⚠️ Bold-OCR fehlgeschlagen: {e}")

    # 2b) Globaler PSM-Sweep (holt oft Summen-/Betragszeilen)
    try:
        full = _imread_safe(image_path, cv2.IMREAD_GRAYSCALE)
        if full is not None:
            full = cv2.GaussianBlur(full, (3,3), 0)
            _, full_bw = cv2.threshold(full, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            for cfg in [r'--oem 3 --psm 6', r'--oem 3 --psm 4']:
                t = pytesseract.image_to_string(full_bw, lang="deu+eng", config=cfg)
                if t.strip():
                    txts.append(safe_post_ocr_cleanup(t))
            # Zahlen/Whitespace-freundlich
            cfg_nums = r'--oem 1 --psm 11 -c preserve_interword_spaces=1 -c user_defined_dpi=350 -c tessedit_char_whitelist=0123456789€.,:%- '
            t_nums = pytesseract.image_to_string(full_bw, lang="deu", config=cfg_nums)
            if t_nums.strip():
                txts.append(safe_post_ocr_cleanup(t_nums))

            cfg_sparse_de = r'--oem 1 --psm 13 -c preserve_interword_spaces=1 -c user_defined_dpi=350'
            t_sparse_de = pytesseract.image_to_string(full_bw, lang="deu", config=cfg_sparse_de)
            if t_sparse_de.strip():
                txts.append(safe_post_ocr_cleanup(t_sparse_de))

    except Exception as e:
        print(f"⚠️ Globaler PSM-Sweep übersprungen: {e}")

    # 3) Alternative Vorverarbeitung (nur wenn nicht FAST)
    if not fast:
        try:
            p2 = preprocess_image_alt(image_path)
            t2 = pytesseract.image_to_string(p2, lang="deu+eng", config=cfg_main)
            if t2.strip():
                txts.append(safe_post_ocr_cleanup(t2))
            t3 = pytesseract.image_to_string(p2, lang="deu+eng", config=cfg_alt)
            if t3.strip():
                txts.append(safe_post_ocr_cleanup(t3))
        except Exception as e:
            print(f"⚠️ Alternative OCR fehlgeschlagen: {e}")

    # 4) Gelbstich-Variante (Grünkanal + adaptiv + invert)
    try:
        if 'preprocess_yellow' in globals():
            bw1, bw2 = preprocess_yellow(img)
            tY1 = pytesseract.image_to_string(bw1, lang="deu+eng", config=cfg_main)
            if tY1.strip():
                txts.append(safe_post_ocr_cleanup(tY1))
            tY2 = pytesseract.image_to_string(bw2, lang="deu+eng", config=cfg_alt)
            if tY2.strip():
                txts.append(safe_post_ocr_cleanup(tY2))
    except Exception as e:
        print(f"⚠️ Gelb-Variante übersprungen: {e}")

    # 5) Belegfuß separat (mehrere Bänder)
    try:
        for bt in ocr_bottom_variants(image_path):
            if bt.strip():
                txts.append(safe_post_ocr_cleanup(bt))
    except Exception as e:
        print(f"⚠️ ROI unten übersprungen: {e}")

    # 5b) Brute-Force Totals-Jäger nur im Langsam-/Debug-Modus
    if not fast:
        try:
            for tx in ocr_bruteforce_totals(image_path):
                if tx.strip():
                    txts.append(safe_post_ocr_cleanup(tx))
        except Exception as e:
            print(f"⚠️ Totals-Bruteforce übersprungen: {e}")

    # 6) Flatten + Deduplizieren + Cleanup
    debug_types = bool(globals().get("DEBUG_OCR_TYPES", False))
    if debug_types:
        # Debug vor dem Deduplizieren: zeigt die Typen in txts
        types = [type(t).__name__ for t in (txts or [])]
        print(f"DEBUG ocr_text_multi types: {types}")

    # Header-Only OCR (hilft, Ladenamen sauber zu bekommen)
    try:
        hdr_txt = ocr_top_header_only(img)
        if hdr_txt:
            txts.append(safe_post_ocr_cleanup(hdr_txt))
    except Exception:
        pass
    try:
        _print_clean_head("(multi)",
                      safe_post_ocr_cleanup("\n".join(txts[:30])),
                      n=20)
    except Exception:
        pass

    uniq = _dedupe_strs(txts)

    # (optional) Debug nach dem Deduplizieren:
    # if debug_types:
    #     types2 = [type(t).__name__ for t in (uniq or [])]
    #     print(f"DEBUG ocr_text_multi uniq types: {types2}")

    return [safe_post_ocr_cleanup(u) for u in uniq]


def ocr_text_multi_many(paths: list[str]) -> list[str]:
    """
    Führt ocr_text_multi() über mehrere Bildteile aus und dedupliziert Texte.
    Reihenfolge der Teile bleibt erhalten.
    """
    all_texts: list[str] = []
    for p in paths:
        try:
            tlist = ocr_text_multi(p) or []
            all_texts.extend(tlist)
        except Exception as e:
            print(f"⚠️ OCR-Teil übersprungen ({p}): {e}")
    # Deduplizieren
    seen = set()
    uniq = []
    for tx in all_texts:
        key = (tx or "").strip()
        if not key or key in seen:
            continue
        seen.add(key)
        uniq.append(key)
    return uniq


def ocr_bottom_region(image_path: str, height_ratio: float = 0.65) -> str:
    """OCR nur im unteren Teil des Bons (z.B. MwSt/Netto/Brutto/Total)."""
    img = _imread_safe(image_path)
    if img is None:
        return ""
    h, w = img.shape[:2]
    y0 = int(h * (1 - height_ratio))
    crop = img[y0:h, 0:w]

    # leichte Vorverarbeitung
    gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
    gray = cv2.fastNlMeansDenoising(gray, None, h=10, templateWindowSize=7, searchWindowSize=21)
    _, bw = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

    txt = pytesseract.image_to_string(bw, lang="deu+eng", config=CUSTOM_CONFIG_MAIN)
    return post_ocr_cleanup(txt)


def _soft_clean(s: str) -> str:
    import re
    s = s.replace("€", " EUR")

    # BRUTT / BRUT -> BRUTTO
    s = re.sub(r"(?i)\bBRUTT\b", "BRUTTO", s)
    s = re.sub(r"(?i)\bBRUT\b(?!TO)", "BRUTTO", s)

    # "70,07 E" / "40.68 E" -> EUR
    s = re.sub(r"(?<=\d)\s*(?:E|£)(?![A-Z])", " EUR", s)

    # "EUR 40.68" / "EUR 40,68" -> "40,68 EUR"
    s = re.sub(
        r"(?i)\bEUR\s+(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})\b",
        lambda m: m.group(1).replace(".", ",") + " EUR",
        s
    )

    # "Ka1e" / "Kale" -> "KARTE"
    # (K + A/4 + 1/I/l + E), case-insensitive
    s = re.sub(r"(?i)\bK(?:A|4)(?:1|I|l)E\b", "KARTE", s)

    # Zeiten ":0%:" -> ":05:"
    s = re.sub(r":0%:", ":05:", s)

    # Dezimal mit EUR vereinheitlichen
    s = re.sub(r'(\d+)\.(\d{2})\s*(EUR|€)', r'\1,\2 EUR', s)

    # Zifferninseln: O/I -> 0/1
    def _fix_digits(m):
        t = m.group(0)
        return t.replace("O","0").replace("o","0").replace("I","1").replace("l","1")
    s = re.sub(r"\b[O0Il1\d]{2,}\b", _fix_digits, s)

    return s


def _normalize_money_tokens(s: str) -> str:
    import re
    # "EUR 10.00" -> "10,00 EUR"
    s = re.sub(r"(?i)\bEUR\s+(\d{1,3}(?:[.,]\d{3})*[.,]\d{1,2})\b",
               lambda m: m.group(1).replace(".", ",") + " EUR",
               s)

    # Reine Betragszeile: "8.85" / "8,85" -> "8,85 EUR"
    s = re.sub(r"(?m)^\s*([0-9][0-9 .•°]*[.,][0-9]{1,2})\s*$",
               lambda m: m.group(1).replace("•",".").replace("°","0").replace(".", ",") + " EUR",
               s)

    # Dezimalpunkt -> Dezimalkomma (allgemein)
    s = re.sub(r"(\d+)\.(\d{2})(?=\D|$)", r"\1,\2", s)
    
    # 1 Dezimalstelle → 2 Dezimalstellen („10,0 EUR“ -> „10,00 EUR“)
    s = re.sub(r"(\d+),(\d)(?=\s*(EUR|€))", r"\1,\g<2>0 \3", s)

    # Nach Schlüsselwörtern "EUR" ergänzen, falls fehlt
    s = re.sub(r"(?mi)\b(Gesamt|Gegeben(?:\s+Bar)?|Zurück|Betrag)\b[^\d]{0,24}"
               r"(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})(?!\s*(EUR|€))",
               lambda m: f"{m.group(1)} {m.group(2).replace('.', ',')} EUR",
               s)
    return s

def _is_filled(v):
    return v is not None and v != "" and v != 0

def _coerce_float(x):
    try:
        if isinstance(x, str):
            x = x.replace("€","").replace("EUR","").strip().replace(".","").replace(",",".")
        return float(x)
    except Exception:
        return None

def _majority(values):
    """Mehrheitswert (ohne Nullen/Leere); bei Gleichstand längster String."""
    vals = [v for v in values if _is_filled(v)]
    if not vals:
        return None
    from collections import Counter
    c = Counter(vals)
    most = c.most_common()
    if len(most) == 1 or most[0][1] > (most[1][1] if len(most)>1 else 0):
        return most[0][0]
    # Gleichstand → längeren/„reicheren“ String bevorzugen
    return max(vals, key=lambda v: (len(str(v)), str(v)))

def _median_float(values):
    nums = []
    for v in values:
        fv = _coerce_float(v)
        if fv is not None:
            nums.append(fv)
    if not nums:
        return None
    nums.sort()
    n = len(nums)
    return nums[n//2] if n % 2 == 1 else (nums[n//2-1] + nums[n//2]) / 2.0

def _score_variant(d):
    """Je mehr Kernfelder gefüllt, desto besser. Klein-Bonus für plausiblen Betrag/Datum/Zeit/Laden."""
    core = ["Betrag (€)","Datum","Uhrzeit","Laden","Netto (€)","MwSt (€)","MwSt %"]
    s = sum(1 for k in core if _is_filled(d.get(k)))
    if _is_filled(d.get("Betrag (€)")): s += 2
    if _is_filled(d.get("Datum")):      s += 1
    if _is_filled(d.get("Uhrzeit")):    s += 1
    if _is_filled(d.get("Laden")):      s += 1
    return s

def first_nonempty(key, dicts=None):
    """Kompatibilitäts-Helfer: erste nicht-leere Ausprägung eines Felds aus Varianten."""
    if dicts is None:
        return None
    for d in dicts:
        v = d.get(key)
        if v is not None and v != "":
            return v
    return None


# ================= Zahlungsarten =================

# ================= Zahlungsarten =================
PAYMENT_KEYWORDS = {
    "Bar": [
        "bar", "cash", "kontant", "contant", "gotówka"
    ],
    "Karte": [
        "girocard", "ec-karte", "ec ", "debit", "credit",
        "visa", "mastercard", "maestro", "karte", "card",
        "kort", "kreditkarte", "bankkarte"
    ],
    "Lidl Pay": ["lidl pay", "lidlpay"],
    "Apple Pay": ["apple pay", "applepay"],
    "Google Pay": ["google pay", "gpay", "googlepay"],
    "PayPal": ["paypal", "pay pal"],
    "Payback Pay": ["payback pay", "paybackpay"],
    "Klarna": ["klarna"],
}

# =================== Parsing ===================

def _guess_store_name(text: str) -> str | None:
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()][:18]
    head  = "\n".join(lines)
    CHAINS = [
        r"\bLIDL\b", r"\bALDI\b", r"\bEDEKA\b", r"\bREWE\b", r"\bNORMA\b",
        r"\bNETTO\b", r"\bPENNY\b", r"\bDM\b", r"\bROSSMANN\b",
        r"\bAPOTHEKE\b", r"\bARAL\b", r"\bSHELL\b", r"\bJET\b",
        r"\bESSO\b", r"\bTOTAL\b", r"\bHAGEBAU\b", r"\bOBI\b",
        r"\bBAUHAUS\b", r"\bHORNBACH\b", r"\bNIELSEN\b", r"\bSCAN-?SHOP\b"
    ]
    chain_pat = re.compile("|".join(CHAINS), re.I)
    MONEY  = re.compile(r"\d{1,3}(?:[.,]\d{3})*[.,]\d{2}\s*(?:€|EUR)?", re.I)
    FOOTKW = re.compile(r"(?i)\b(einkaufswert|summe|gesamt|betrag|zu\s*zahlen|gegeben|zurück|wechselgeld|mwst|ust|netto|brutto|girocard|ust-?id|ta-?nr|bnr|terminal|karte)\b")
    GENERIC= re.compile(r"(?i)^(apotheke|bäckerei|metzgerei|imbiss|shop)$")
    NAME   = re.compile(r"(?i)^[A-ZÄÖÜ0-9][A-ZÄÖÜ0-9\s\-\&\.]{2,40}$")
    STREET = re.compile(r"(?i)\b(str(?:aße|asse)\.?|weg|platz|allee|gasse)\b")
    ZIP    = re.compile(r"\b\d{4,5}\b")
    def _is_head_name(ln: str) -> bool:
        if MONEY.search(ln): return False
        if FOOTKW.search(ln): return False
        if _looks_like_noise(ln): return False
        letters = sum(ch.isalpha() for ch in ln)
        return letters >= 4
    m = chain_pat.search(head)
    if m: return m.group(0).upper().strip()
    for i, ln in enumerate(lines):
        if not _is_head_name(ln): continue
        if GENERIC.match(ln): continue
        if NAME.match(ln):
            near = lines[i+1:i+4]
            if any(_is_head_name(x) and (STREET.search(x) or ZIP.search(x)) for x in near):
                return ln.strip()
    for ln in lines:
        if _is_head_name(ln) and not GENERIC.match(ln):
            return ln.strip()
    return None

def _is_plausible_total(val: float) -> bool:
    """
    Sehr simple Schranke gegen Ausreißer:
    - Totale < 0.01 oder > 2000 verwerfen (kannst du anpassen)
    """
    try:
        if val is None:
            return False
        if not isinstance(val, (int, float)):
            return False
        if val < 0.01 or val > 2000:   # Deckel anpassen, falls du große Rechnungen hast
            return False
        return True
    except Exception:
        return False


def _sanitize_total(maybe_val, context_text: str = ""):
    """
    Repariert typische OCR-Patzer:
      - Zahlen ohne Dezimaltrenner, die aber wie 5-stellige Timestamps aussehen (z.B. '356.000Z' → 35600) → verwerfen
      - Beträge ohne Komma und >= 1000 → verwerfen
      - Nur Werte mit ,xx (oder .xx) bevorzugen
    """
    def _looks_like_garbage_nearby(s: str) -> bool:
        # Z / T / ms aus ISO-Zeitstempeln in der Nähe? => Mist
        return bool(re.search(r"[ZT]\s*$", s)) or "000Z" in s.upper()

    try:
        if maybe_val in (None, "", 0):
            return None
        v = float(maybe_val)
        # Ausreißer raus
        if not _is_plausible_total(v):
            return None
        # Zusatz-Heuristik: Wenn im Kontext häufig '000Z' / ISO-Stempel auftauchen und v sehr groß ist → No
        if v >= 1000 and re.search(r"\d{4}-\d{2}-\d{2}T\d{2}:\d{2}", context_text):
            return None
        return round(v, 2)
    except Exception:
        return None


def extract_data_from_text(text: str) -> dict:
    """
    Generischer Parser für Handels-/Apotheken-Belege.
    Greift speziell im Belegfuß nach Netto/Brutto/MwSt UND zusätzlich
    nach Gesamt/Betrag/Summe. Fällt bei Bedarf auf pick_total(...) zurück.
    Erwartet vorhandene Helfer:
      - _parse_date_loose, _parse_time_loose
      - _roi_lines(pattern, window)
      - _find_first(patterns, text, flags=...)
      - _norm_money(str) -> float
      - pick_total(text, liter=None, price_per_l=None)
    """
    import re
    # robustes Geld-Pattern (1.234,56 oder 47,66)
    money = _money_regex() if ' _money_regex' in globals() or '_money_regex' in dir() else r"(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})"

    data: dict = {}
    data = {}

    # Datum & Uhrzeit früh ziehen
    d = _parse_date_loose(text)
    if d: data["Datum"] = d
    t = _parse_time_loose(text)
    if t: data["Uhrzeit"] = t
    if not data.get("Uhrzeit"):
        m = re.search(r"(?<!\d)([01]?\d|2[0-3]):[0-5]\d(?!\d)", text)
        if m:
            data["Uhrzeit"] = m.group(0)

    # --- Laden (sehr einfache Heuristik: erste sinnvolle Zeile im Kopfbereich) ---
    head_lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    for ln in head_lines[:8]:
        low = ln.lower()
        if re.search(r"(apotheke|drogerie|markt|center|hagebau|aral|shell|bft|dm|rossmann|rewe|edeka|aldi|lidl|kajute|restaurant)", low):
            data["Laden"] = ln
            break
    if not USE_PADDLE:
        if "Laden" not in data or not data["Laden"]:
            store_now = (data.get("Laden") or "").strip()
            if (not store_now) or _is_bad_store_name(store_now) or store_now.upper() in {"IN","DER","IN DER","APOTHEKE"}:
                store_guess = _guess_store_name(text)
                if store_guess:
                    data["Laden"] = store_guess
    # --- Belegfuß-ROI: hier stehen meist MwSt/Netto/Brutto/Gesamt ---
    foot = _roi_lines(
        text,
        r"(NETTO|BRUTTO|MWST|UST|TOTAL|SUMME|GESAMT|BETRAG|EINKAUFSWERT|ZU\s*ZAHLEN|GIROCARD|KARTENZAHLUNG)",
        window=6
    )

    # Im Parser (extract_data_from_text), nach Bildung eines 'foot'-Abschnitts:
    m7  = _find_first([r"(?mi)\b7[,\.]?0{1,2}\s*%[^\d]{0,15}(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})[^\d]{0,15}(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})"], foot)
    m19 = _find_first([r"(?mi)\b19[,\.]?0{1,2}\s*%[^\d]{0,15}(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})[^\d]{0,15}(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})"], foot)
    # Gruppenreihenfolge an deinen Footer anpassen (manchmal Netto vor MwSt, manchmal umgekehrt)
    if m7:
        data["Netto 7% (€)"] = _norm_money(m7.group(1))
        data["MwSt 7% (€)"]  = _norm_money(m7.group(2))
    if m19:
        data["Netto 19% (€)"] = _norm_money(m19.group(1))
        data["MwSt 19% (€)"]  = _norm_money(m19.group(2))


    # MwSt % (z. B. 19,00 %)
    mw = _find_first(
        [r"MW?S?T[^0-9]{0,12}([0-9]{1,2}[.,]?[0-9]{0,2})\s*%"],
        foot,
        flags=re.IGNORECASE | re.DOTALL
    )
    if mw:
        try:
            data["MwSt %"] = _norm_money(mw.group(1))
        except Exception:
            pass

    # MwSt (€)
    m_mw_eur = re.search(
        r"(?:mwst|ust|steuer)[^\d]{0,15}([0-9]{1,3}(?:[.,]\d{3})*[.,]\d{2})",
        foot, re.IGNORECASE | re.DOTALL
    )
    if m_mw_eur:
        try:
            data["MwSt (€)"] = float(m_mw_eur.group(1).replace(".", "").replace(",", "."))
        except Exception:
            pass

    # Netto (€)
    m_net = re.search(
        r"\bnett[oa]?[^\d]{0,20}([0-9]{1,3}(?:[.,]\d{3})*[.,]\d{2})",
        foot, re.IGNORECASE | re.DOTALL
    )
    if m_net:
        try:
            data["Netto (€)"] = float(m_net.group(1).replace(".", "").replace(",", "."))
        except Exception:
            pass

    # Brutto (€) – optional, manchmal identisch zu „Gesamt“
    m_brut = re.search(
        r"\bbrutt[oa]?[^\d]{0,20}([0-9]{1,3}(?:[.,]\d{3})*[.,]\d{2})",
        foot, re.IGNORECASE | re.DOTALL
    )
    if m_brut:
        try:
            data["Brutto (€)"] = float(m_brut.group(1).replace(".", "").replace(",", "."))
        except Exception:
            pass

    # --- STRONG TOTAL FINDER (nimmt letzte "zu zahlen / Gesamt / Endbetrag"-Zeile) ---
    def _find_total_strong(s: str):
        money = r"(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})"
        # nimm die letzte passende Stelle (Bon-Ende)
        m_last = None
        for m in re.finditer(rf"(?im)\b(zu\s*zahlen|endbetrag|summe|gesamt|betrag)\b[^0-9]{{0,40}}{money}\s*(?:€|eur)?", s):
            m_last = m
        if not m_last:
            return None
        try:
            return _norm_money(m_last.group(2))
        except Exception:
            return None
    
    # 1) starker Versuch
    tot_val = _find_total_strong(text)
    # 2) falls das nichts liefert, evtl. steht EUR vorne ("EUR 25,01")
    if tot_val is None:
        for m in re.finditer(rf"(?im)\b(zu\s*zahlen|endbetrag|summe|gesamt|betrag)\b[^0-9]{{0,40}}\b(?:€|eur)\s*{money}", text):
            try:
                tot_val = _norm_money(m.group(2))
            except Exception:
                pass
    
    # 3) Sanity-Check: Ausreißer (OCR zusammengeklebt) abwehren
    if isinstance(tot_val, (int, float)):
        # Ein normaler Supermarktbon mit > 10.000 ist praktisch ausgeschlossen.
        # Lass es nur zu, wenn auch eine plausible EUR-Form daneben steht.
        if tot_val >= 10000:
            tot_val = None
    
    # 4) setzen, falls plausibel
    # --- BETRAG ---
    if not USE_PADDLE:
    # alte Fallbacks / Merge-Korrekturen
        # starker Fallback (holt z. B. „zu zahlen 25,01 EUR“)
        if "Betrag (€)" not in data or not data["Betrag (€)"]:
            def _is_plausible_amount(v: float) -> bool:
                # keine gigantischen Zahlen (OCR „356.000Z“ → 35600.0)
                if v <= 0 or v > 2000:  # Einkaufs-Bons >2000 € sind extrem selten
                    return False
                return True

    # ... beim Setzen von "Betrag (€)":
    tot = _pick_total_candidate(text)
    if tot is not None and _is_plausible_amount(tot):
        data["Betrag (€)"] = tot
    # --- Gegeben / Wechselgeld (optional) ---
    m = re.search(rf"(?i)\b(gegeben|bar(?:zahlung)?|gezahlt)\b[^\d]{{0,20}}{money}\s*(?:€|eur)?", text)
    if m:
        try: data["Gegeben (€)"] = _norm_money(m.group(2))
        except: pass

    m = re.search(rf"(?i)\b(zurück|wechselgeld|rückgeld)\b[^\d]{{0,20}}{money}\s*(?:€|eur)?", text)
    if m:
        try: data["Wechselgeld (€)"] = _norm_money(m.group(2))
        except: pass

    # Gegeben / Zurück optional erfassen
    m_given = _find_first(
        [r"(?mi)\bGegeben(?:\s+Bar)?\b[^\d]{0,12}(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})"],
        text
    )
    if m_given:
        try: data["Gegeben (€)"] = _norm_money(m_given.group(1))
        except: pass

    m_change = _find_first(
        [r"(?mi)\bZurück\b[^\d]{0,12}(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})"],
        text
    )
    if m_change:
        try: data["Wechselgeld (€)"] = _norm_money(m_change.group(1))
        except: pass

    # --- BETRAG (Belegfuß bevorzugen) ---
    # 1) Bevorzugt Zeilen mit Schlüsselwörtern im Fuß (letzte ~40 Zeilen)
    KEY = r"(?i)\b(zu\s*zahlen|endbetrag|gesamt(?:betrag)?|summe|betrag|total)\b"
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    tail  = lines[-40:] if len(lines) > 40 else lines

    candidates = []
    for ln in tail:
        if re.search(KEY, ln):
            for a in re.findall(money, ln):
                try:
                    v = _norm_money(a)
                    if v is not None:
                        candidates.append(v)
                except:
                    pass

    # 2) Falls in Keyword-Zeilen nichts gefunden: allgemein größte Beträge im Fuß
    if not candidates:
        for ln in tail:
            for a in re.findall(money, ln):
                try:
                    v = _norm_money(a)
                    if v is not None:
                        candidates.append(v)
                except:
                    pass

    if candidates:
        # Brutto ist fast immer der größte Wert im Fuß
        data["Betrag (€)"] = max(candidates)

    if not USE_PADDLE:
    # alte Fallbacks / Merge-Korrekturen
        # 3) Falls noch leer, letzter Fallback (ganzer Text, Keywords)
        if "Betrag (€)" not in data or not data["Betrag (€)"]:
            tot = _pick_total_candidate(text)
            if tot is not None:
                data["Betrag (€)"] = tot
            
    # --- Zahlungsart ---
    if "Zahlung" not in data or not data["Zahlung"]:
        pay = detect_payment_method(text)
        if pay:
            data["Zahlung"] = pay
        else:
            data["Zahlung"] = "Unbekannt"


    return data


# =================== Merger ===================

def score_and_flag(row: dict) -> tuple[int, str]:
    score = 0
    if row.get("Datum"): score += 20
    if row.get("Betrag (€)"): score += 40
    if row.get("Laden"): score += 10
    if row.get("Liter") is not None: score += 10
    if row.get("€/L") is not None: score += 10
    if row.get("Zahlung"): score += 5

    flag = "OK"
    if (row.get("Betrag (€)") is None) or (row.get("Datum") is None):
        flag = "prüfen"
    return score, flag

def choose_best(candidates: list[dict]) -> dict:
    keys = set().union(*[c.keys() for c in candidates])
    out = {}
    def has_amount(d): return d.get("Betrag (€)") is not None and d["Betrag (€)"] > 0

    def score(d):
        s = 0
        if d.get("Datum"): s += 20
        if has_amount(d): s += 40
        if d.get("Laden"): s += 10
        return s

    ranked = sorted(candidates, key=score, reverse=True)
    best_first = ranked[0] if ranked else {}

    for k in keys:
        val = best_first.get(k)
        if val is None:
            for d in ranked[1:]:
                v = d.get(k)
                if v: 
                    out[k] = v
                    break
        else:
            out[k] = val

    conf, flag = score_and_flag(out)
    out["Confidence"] = conf
    out["Prüfstatus"] = flag
    return out

def pick_total(text: str, liter: float | None = None, price_per_l: float | None = None) -> float | None:
    import re
    def _to_float(num: str): return float(num.replace(" ", "").replace(".", "").replace(",", "."))
    def amounts(X: str):
        vals = []
        for m in re.finditer(r"(?<!\d)(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})(?:\s*(?:EUR|€))?", X):
            try: vals.append(_to_float(m.group(1)))
            except: pass
        out=[]
        for v in sorted(vals, reverse=True):
            if not any(abs(v-u)<1e-3 for u in out): out.append(v)
        return out

    lines = text.splitlines()
    # Zahlungs-/Total-ROI erweitert
    keys = r"(RECHNUNGSBETRAG|TOTAL|SUMME|EINKAUFSWERT|ZU\s*ZAHLEN|BAR|EC|EC[-\s]?CHIP|KARTE|GIROCARD|BETRAG|BRUTTO|BRUTT)"
    roi=[]
    for i, ln in enumerate(lines):
        if re.search(keys, ln, re.IGNORECASE):
            roi.extend(lines[max(0,i-3):min(len(lines), i+6)])
    roi_text = "\n".join(roi) if roi else text

    roi_vals = [v for v in amounts(roi_text) if 0.5 <= v <= 5000.0]
    all_vals = [v for v in amounts(text)     if 0.5 <= v <= 5000.0]

    # Konsistenz Liter*€/L (±4 %)
    if liter and price_per_l:
        est = liter * price_per_l
        best, best_rel = None, 1e9
        for v in roi_vals + all_vals:
            rel = abs(v - est) / max(est, 1e-6)
            if rel < best_rel: best, best_rel = v, rel
        if best is not None and best_rel <= 0.04: return round(best, 2)

    # Zeilen mit Zahlungswörtern bevorzugen
    pay_lines = [ln for ln in lines if re.search(r"(girocard|karte|kartenzahlung|ec|betrag)", ln, re.IGNORECASE)]
    pay_text = "\n".join(pay_lines)
    pay_vals = [v for v in amounts(pay_text) if 0.5 <= v <= 5000.0]

    def freq_in(s, v):
        pat = f"{v:.2f}".replace(".", ",")
        return sum(1 for ln in s.splitlines() if (pat in ln) or (f"{v:.2f}" in ln))

    # Heuristik: „Eingefügte Ziffer“ (z. B. 57,81 vs 537,81) -> kleineren Wert bevorzugen, wenn er häufiger ist
    def penalize_insert_err(cands):
        if len(cands) < 2: return cands
        S = set(cands)
        for v in list(cands):
            for u in list(cands):
                if abs(u - v) > 300:  # nur bei plausiblen „+Ziffer“-Fehlern vergleichen
                    continue
                # Wenn u exakt eine Ziffer „mehr“ hat (stringbasiert), dann v bevorzugen
                sv = f"{v:.2f}".replace(".", ",")
                su = f"{u:.2f}".replace(".", ",")
                if len(su.replace(",","")) == len(sv.replace(",","")) + 1 and freq_in(text, v) >= freq_in(text, u):
                    S.discard(u)
        return sorted(S, reverse=True)

    # Kandidatenliste zusammenstellen & filtern
    cand = roi_vals or all_vals
    if not cand: return None
    cand = penalize_insert_err(cand)

    # Scoring: (Häufigkeit im ROI, Häufigkeit global, vorkommen in Zahlungszeilen, Wert)
    best = max(
        cand,
        key=lambda v: (
            freq_in(roi_text, v),
            freq_in(text, v),
            freq_in(pay_text, v),
            v
        )
    )
    return round(best, 2)


# =================== Excel ===================

def append_to_excel(row_dict, excel_path=EXCEL_PATH, sheet="Bons"):
    import pandas as pd
    # 1) existierende Datei/Sheet laden oder neu
    if os.path.exists(excel_path):
        try:
            df = pd.read_excel(excel_path, sheet_name=sheet)
        except Exception:
            df = pd.DataFrame(columns=COLUMNS)
    else:
        df = pd.DataFrame(columns=COLUMNS)

    # 2) fehlende Spalten ergänzen
    for c in COLUMNS:
        if c not in df.columns:
            df[c] = pd.Series(dtype="object")

    # 3) neue Zeile in fester Spaltenreihenfolge
    row = {c: row_dict.get(c) for c in COLUMNS}
    df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)

    # 4) speichern
    with pd.ExcelWriter(excel_path, engine="openpyxl", mode="w") as w:
        df[COLUMNS].to_excel(w, sheet_name=sheet, index=False)

def _ensure_cols(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    """Stellt sicher, dass alle gewünschten Spalten existieren (in richtiger Reihenfolge)."""
    for c in columns:
        if c not in df.columns:
            df[c] = None
    # nur gewünschte Spalten und in der gewünschten Reihenfolge
    return df.reindex(columns=columns)


def df_append_row(df: pd.DataFrame, row: dict, columns: list[str]) -> pd.DataFrame:
    """
    Hängt eine Zeile stabil an:
    - entfernt komplett leere Felder (None / "")
    - erzwingt Spaltenreihenfolge
    - vermeidet Pandas FutureWarning beim concat
    """
    # 1) leere Felder wegräumen (damit Pandas nicht „all-NA“ Spaltenratelei macht)
    row_clean = {k: v for k, v in row.items() if v is not None and v != ""}

    # 2) sicherstellen, dass alle Zielspalten existieren (und in richtiger Reihenfolge sind)
    df = _ensure_cols(df, columns)

    # 3) nur erlaubte Spalten in die neue Zeile, fehlende bleiben automatisch NaN/None
    row_aligned = {c: row_clean.get(c) for c in columns}

    # 4) ohne Warning anhängen
    to_add = pd.DataFrame([row_aligned], columns=columns, dtype=object)
    return pd.concat([df, to_add], ignore_index=True)

def _dedupe_append(df: pd.DataFrame, row: dict, key_cols: list[str]) -> pd.DataFrame:
    """Hängt row nur an, wenn keine Zeile mit identischem Schlüssel existiert."""
    if df.empty or not set(key_cols).issubset(df.columns):
        return pd.concat([df, pd.DataFrame([row])], ignore_index=True)
    mask = pd.Series([True] * len(df))
    for c in key_cols:
        mask &= (df[c].astype(str).fillna("") == str(row.get(c, "")))
    if mask.any():
        # Optional: vorhandene Zeile updaten – hier lassen wir's einfach so.
        return df
    return pd.concat([df, pd.DataFrame([row])], ignore_index=True)

# =========================
# SPEICHERN / EXCEL
# =========================

def build_receipt_signature(d: dict) -> str:
    def norm_text(x):
        return str(x or "").strip().lower()

    def norm_amount(x):
        if isinstance(x, (int, float)):
            return f"{x:.2f}"
        return norm_text(x)

    parts = [
        norm_text(d.get("Datum")),
        norm_text(d.get("Uhrzeit")),
        norm_text(d.get("Laden")),
        norm_amount(d.get("Betrag (€)")),
    ]
    return "|".join(parts)

def is_duplicate_receipt(d: dict, excel_path: str) -> bool:
    import os
    from openpyxl import load_workbook

    if not os.path.exists(excel_path):
        return False

    sig_new = build_receipt_signature(d)

    try:
        wb = load_workbook(excel_path)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        col_map = {name: idx + 1 for idx, name in enumerate(headers) if name}

        needed = ["Datum", "Uhrzeit", "Laden", "Betrag (€)"]
        if not all(k in col_map for k in needed):
            return False

        for row in range(2, ws.max_row + 1):
            existing = {
                "Datum": ws.cell(row=row, column=col_map["Datum"]).value,
                "Uhrzeit": ws.cell(row=row, column=col_map["Uhrzeit"]).value,
                "Laden": ws.cell(row=row, column=col_map["Laden"]).value,
                "Betrag (€)": ws.cell(row=row, column=col_map["Betrag (€)"]).value,
            }
            sig_old = build_receipt_signature(existing)
            if sig_old == sig_new:
                return True

    except Exception:
        return False

    return False

def append_to_excel_typed(row: dict, rtype: str, excel_path: str = "kassenbons.xlsx"):
    import os, pandas as pd

    def _ensure_cols(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
        # fehlende Spalten ergänzen, dann in gewünschter Reihenfolge anordnen
        for c in cols:
            if c not in df.columns:
                df[c] = None
        return df[cols]

    # 1) Bestehende Sheets laden (falls Datei schon existiert)
    sheets: dict[str, pd.DataFrame] = {}
    if os.path.exists(excel_path):
        try:
            xls = pd.ExcelFile(excel_path)
            for sh in xls.sheet_names:
                sheets[sh] = pd.read_excel(excel_path, sheet_name=sh)
        except Exception:
            sheets = {}

    # 2) Master aktualisieren
    master_cols = MASTER_COLUMNS
    master_df = sheets.get("Master", pd.DataFrame(columns=master_cols))
    row_master = {c: row.get(c) for c in master_cols}
    # Belegtyp sicher setzen
    row_master["Belegtyp"] = rtype

    # NEU: ohne concat anhängen
    # Master
    master_df  = sheets.get("Master", pd.DataFrame(columns=master_cols))
    row_master = {c: row.get(c) for c in master_cols}
    master_df  = df_append_row(master_df, row_master, master_cols)
    sheets["Master"] = master_df

    # 3) Typ-Blatt aktualisieren
    sheet_name = SHEET_BY_TYPE.get(rtype, "Einzelhandel")
    type_cols  = _columns_for_type(rtype)
    type_df    = sheets.get(sheet_name, pd.DataFrame(columns=type_cols))
    row_type   = _order_row_for_columns(row, type_cols)
    
    # Schlüssel fürs Dedupe (für Tankstellen robust):
    #key_cols = ["Datum","Uhrzeit","Laden","Betrag (€)"]
    
    # Typ-Blatt
    type_df  = sheets.get(sheet_name, pd.DataFrame(columns=type_cols))
    row_type = {c: row.get(c) for c in type_cols}
    type_df  = df_append_row(type_df, row_type, type_cols)
    sheets[sheet_name] = type_df

    # 4) Alles zusammen zurückschreiben (ein Writer-Durchgang)
    with pd.ExcelWriter(excel_path, engine="openpyxl", mode="w") as xls:
        for sh_name, df in sheets.items():
            df.to_excel(xls, sheet_name=sh_name, index=False)

# =========================
# OCR / DATEI-VERARBEITUNG
# =========================

def run_paddle_ocr(image_path: str) -> list[dict]:
    """
    Führt PaddleOCR auf einem Bild aus und gibt eine sortierte Liste von OCR-Zeilen zurück.
    Jede Zeile ist ein Dict mit text, score, x1, y1, x2, y2, cx, cy.
    """
    global _PADDLE_OCR

    if _PADDLE_OCR is None:
        kwargs = {
            "lang": "de",
            "use_doc_orientation_classify": False,
            "use_doc_unwarping": False,
        }

        try:
            _PADDLE_OCR = PaddleOCR(show_log=False, **kwargs)
        except Exception as e:
            if "show_log" in str(e):
                _PADDLE_OCR = PaddleOCR(**kwargs)
            else:
                raise

    result = _PADDLE_OCR.predict(str(image_path))
    if not result:
        return []

    page = result[0]
    texts = page.get("rec_texts", [])
    scores = page.get("rec_scores", [])
    boxes = page.get("rec_boxes", [])

    lines = []
    for i, text in enumerate(texts):
        if not text:
            continue

        score = float(scores[i]) if i < len(scores) else 0.0
        box = boxes[i] if i < len(boxes) else None
        if box is None:
            continue

        x1, y1, x2, y2 = map(int, box)

        lines.append({
            "text": str(text).strip(),
            "score": score,
            "x1": x1,
            "y1": y1,
            "x2": x2,
            "y2": y2,
            "cx": (x1 + x2) // 2,
            "cy": (y1 + y2) // 2,
        })

    lines.sort(key=lambda d: (d["cy"], d["x1"]))
    return lines

def normalize_tax_fields(best: dict) -> dict:
    """
    Vereinheitlicht MwSt/Netto-Felder.
    Bei gemischten Steuersätzen werden die globalen Felder geleert,
    damit keine falschen Einzelwerte in der Haupttabelle stehen.
    """
    import re

    rates = set()

    # 1) zuerst explizit gesammelte Steuersätze verwenden
    tax_rates = best.get("_tax_rates_found")
    if isinstance(tax_rates, (set, list, tuple)):
        for x in tax_rates:
            try:
                rates.add(int(x))
            except Exception:
                pass

    # 2) Fallback: direkt aus Rohtext erkennen
    raw = (best.get("Rohtext") or "").upper()

    # typische Muster: "A 7 %", "B 19 %", "7,00", "19,00", "7 %", "19 %"
    if re.search(r"\b7\s*%", raw) or re.search(r"\b7[.,]00\b", raw):
        rates.add(7)
    if re.search(r"\b19\s*%", raw) or re.search(r"\b19[.,]00\b", raw):
        rates.add(19)

    # 3) Entscheidung
    if len(rates) == 1:
        best["MwSt %"] = list(rates)[0]
    elif len(rates) > 1:
        best.pop("MwSt %", None)
        best.pop("MwSt (€)", None)
        best.pop("Netto (€)", None)

    # internes Hilfsfeld nicht nach Excel durchreichen
    best.pop("_tax_rates_found", None)

    return best
# =========================
# PARSER / EXTRAKTION
# =========================

def parse_receipt_blocks(lines: list[dict]) -> dict:
    """
    Parst PaddleOCR-Zeilen in ein Ergebnis-Dict.
    Fokus: Laden, Adresse, Datum/Uhrzeit, Betrag, Gegeben, Wechselgeld,
    MwSt, Netto, Zahlung.
    """
    import re

    out = {}
    if not lines:
        return out

    # -------------------------
    # Vorbereiten
    # -------------------------
    lines = sorted(lines, key=lambda d: (d["cy"], d["x1"]))
    max_y = max(ln["y2"] for ln in lines)
    header_lines = [ln for ln in lines if ln["cy"] <= max_y * 0.22]

    def _txt(ln):
        return (ln.get("text") or "").strip()

    def _money_matches(s: str, allow_negative: bool = True) -> list[tuple[str, float]]:
        patt = r"(?<!\d)-?\d{1,4}[.,]\d{2}(?!\d)" if allow_negative else r"(?<!\d)\d{1,4}[.,]\d{2}(?!\d)"
        found = []
        for m in re.findall(patt, s):
            try:
                val = float(m.replace(",", "."))
            except Exception:
                continue
            found.append((m, val))
        return found

    def _find_money_near(idx, window_before=2, window_after=2, allow_negative=True):
        vals = []
        start = max(0, idx - window_before)
        end = min(len(lines), idx + window_after + 1)
        for j in range(start, end):
            txt2 = _txt(lines[j])
            for raw, val in _money_matches(txt2, allow_negative=allow_negative):
                vals.append((j, raw, val, txt2))
        return vals

    def _has_card_hint(txt: str) -> bool:
        return any(k in txt for k in ["karte", "ec", "girocard", "mastercard", "visa", "karteec"])

    # -------------------------
    # Laden / Adresse
    # -------------------------
    bad_header_words = {
        "rechnung", "artikelname", "menge", "einzel", "gesamt",
        "summe", "gesamtsumme", "betrag", "datum", "uhrzeit"
    }

    store = None
    address = None

    for ln in header_lines:
        txt = _txt(ln)
        txt_low = txt.lower()

        if not txt:
            continue

        if not store:
            if (
                len(txt) > 8
                and any(c.isalpha() for c in txt)
                and txt_low not in bad_header_words
                and not re.fullmatch(r"[\d\s\-./]+", txt)
            ):
                store = txt
                continue

        if not address:
            if any(ch.isdigit() for ch in txt) and any(c.isalpha() for c in txt):
                address = txt

    if store:
        out["Laden"] = store
    if address:
        out["Adresse"] = address

    # -------------------------
    # Datum / Uhrzeit
    # -------------------------
    found_date = None
    found_time = None

    # 1) zuerst kombinierte Zeilen suchen
    for ln in lines:
        txt = _txt(ln)
        if not txt:
            continue

        m = re.search(r"(\d{2}\.\d{2}\.\d{2,4}).*?(\d{2}:\d{2})", txt)
        if m:
            found_date = m.group(1)
            found_time = m.group(2)
            break

    # 2) falls nicht gefunden: getrennte Zeilen suchen, bevorzugt im unteren Bereich
    if not found_date or not found_time:
        for i in range(len(lines) - 1, -1, -1):
            txt = _txt(lines[i])

            if not found_time:
                m_time = re.search(r"\b(\d{2}:\d{2})\b", txt)
                if m_time:
                    found_time = m_time.group(1)

            if not found_date:
                m_date = re.search(r"\b(\d{2}\.\d{2}\.\d{2,4})\b", txt)
                if m_date:
                    found_date = m_date.group(1)

            if found_date and found_time:
                break

    if found_date:
        m_date = re.match(r"(\d{2})\.(\d{2})\.(\d{2,4})", found_date)
        if m_date:
            dd, mm, yy = m_date.groups()
            if len(yy) == 2:
                yy = "20" + yy
            out["Datum"] = f"{yy}-{mm}-{dd}"

    if found_time:
        out["Uhrzeit"] = found_time

    # -------------------------
    # Zahlung / Gegeben / Rückgeld
    # -------------------------
    payment = None
    given_amount = None
    change_amount = None

    for i, ln in enumerate(lines):
        txt = _txt(ln).lower()

        # Explizite Bar-Zeilen zuerst
        if txt == "bar" or "barzahlung" in txt:
            payment = "Bar"

        # Rückgeld -> sicher Bar
        if any(k in txt for k in ["ruckgeld", "rückgeld", "zuruck", "zurück"]):
            payment = "Bar"
            nearby = _find_money_near(i, window_before=3, window_after=1, allow_negative=True)
            negs = [x for x in nearby if x[2] < 0]
            if negs:
                change_amount = negs[-1][1]

        # Karte nur dann, wenn nicht schon Bar sicher erkannt wurde
        if payment != "Bar" and _has_card_hint(txt):
            payment = "Karte"

        # "gegeben" nur dann als Bar interpretieren, wenn kein Kartenhinweis in derselben Zeile steht
        if ("gegeben" in txt) and not _has_card_hint(txt):
            payment = "Bar"
            nearby = _find_money_near(i, window_before=1, window_after=3, allow_negative=False)
            pos = [x for x in nearby if x[2] > 0]
            if pos:
                pos.sort(key=lambda x: x[2])
                given_amount = pos[-1][1]

    # -------------------------
    # Betrag
    # -------------------------
    total = None
    total_keywords = [
        "gesamtsumme", "summe", "summi", "gesamt",
        "endbetrag", "zu zahlen", "zahlbetrag", "betrag"
    ]

    # 1) Summe-Zeilen bevorzugen
    for i, ln in enumerate(lines):
        txt = _txt(ln).lower()
        if any(k in txt for k in total_keywords):
            nearby = _find_money_near(i, window_before=1, window_after=2, allow_negative=False)
            pos = [x for x in nearby if x[2] > 2.0]
            if pos:
                pos.sort(key=lambda x: x[2])  # kleineren plausiblen Wert nehmen
                total = pos[0][1]
                break

    # 2) Fallback: Brutto-Zeile
    if not total:
        for i, ln in enumerate(lines):
            txt = _txt(ln).lower()
            if "brutto" in txt:
                nearby = _find_money_near(i, window_before=0, window_after=3, allow_negative=False)
                pos = [x for x in nearby if x[2] > 2.0]
                if pos:
                    pos.sort(key=lambda x: x[2], reverse=True)
                    total = pos[0][1]
                    break

    # 3) Nur wenn Gegeben UND Wechselgeld da sind, daraus Betrag setzen
    if payment == "Bar" and given_amount and change_amount:
        try:
            g = float(str(given_amount).replace(",", "."))
            c = abs(float(str(change_amount).replace(",", ".")))
            total = f"{g - c:.2f}".replace(".", ",")
        except Exception:
            pass

    if total:
        try:
            out["Betrag (€)"] = float(str(total).replace(",", "."))
        except Exception:
            pass

    # -------------------------
    # Steuerblock
    # -------------------------
    tax_rates_found = set()
    tax_rate = None
    tax_amount = None
    net_amount = None
    gross_from_tax = None

    for i, ln in enumerate(lines):
        txt = _txt(ln).lower()

        if any(k in txt for k in ["mwst", "steuer", "steuersatz", "netto", "brutto"]):
            nearby = _find_money_near(i, window_before=0, window_after=6, allow_negative=False)
            vals = [(raw, val) for (_j, raw, val, _txt2) in nearby if val > 0]

            # Steuersatz
            for j in range(i, min(i + 6, len(lines))):
                t2 = _txt(lines[j])

                m_pct = re.search(r"\b(7|19)\s*%", t2)
                if m_pct:
                    tax_rates_found.add(int(m_pct.group(1)))

                m_pct2 = re.search(r"[AB]?\s*(7|19)[.,]00", t2)
                if m_pct2:
                    tax_rates_found.add(int(m_pct2.group(1)))

            if vals:
                # kleinster plausibler positiver Wert = Steuer
                if tax_amount is None:
                    smalls = [(r, v) for (r, v) in vals if 0 < v < 10]
                    if smalls:
                        smalls.sort(key=lambda x: x[1])
                        tax_amount = smalls[0][0]

                # Im Steuerblock liegen meist genau zwei große Werte:
                # Brutto und Netto. Der größere ist Brutto, der kleinere Netto.
                bigger = [(r, v) for (r, v) in vals if v >= 2]
                if len(bigger) >= 2:
                    bigger.sort(key=lambda x: x[1])
                    net_amount = bigger[-2][0]  # kleinerer der beiden großen Werte
                    gross_from_tax = bigger[-1][0]  # größerer der beiden großen Werte
                elif len(bigger) == 1:
                    # Fallback
                    if gross_from_tax is None:
                        gross_from_tax = bigger[0][0]

    out["_tax_rates_found"] = sorted(tax_rates_found)

    if len(tax_rates_found) == 1:
        tax_rate = list(tax_rates_found)[0]
        out["MwSt %"] = tax_rate
    else:
        tax_rate = None

    if tax_amount:
        try:
            out["MwSt (€)"] = float(str(tax_amount).replace(",", "."))
        except Exception:
            pass

    # Netto zunächst nur als Fallback übernehmen
    if net_amount:
        try:
            out["Netto (€)"] = float(str(net_amount).replace(",", "."))
        except Exception:
            pass

    if gross_from_tax:
        try:
            gross_tax_val = float(str(gross_from_tax).replace(",", "."))
            gross_now = out.get("Betrag (€)")
            if not isinstance(gross_now, (int, float)) or abs(gross_now - gross_tax_val) > 0.05:
                out["Betrag (€)"] = gross_tax_val
        except Exception:
            pass

    # -------------------------
    # Konsistenzkorrekturen
    # -------------------------
    gross = out.get("Betrag (€)")
    vat = out.get("MwSt (€)")
    net = out.get("Netto (€)")

    # Wenn Brutto und MwSt da sind, Netto IMMER daraus ableiten.
    if isinstance(gross, (int, float)) and isinstance(vat, (int, float)):
        calc_net = round(gross - vat, 2)
        if not isinstance(net, (int, float)) or abs(net - calc_net) > 0.05:
            out["Netto (€)"] = calc_net

    # Gegeben / Wechselgeld schreiben
    if given_amount:
        try:
            out["Gegeben (€)"] = float(str(given_amount).replace(",", "."))
        except Exception:
            pass

    if change_amount:
        try:
            out["Wechselgeld (€)"] = abs(float(str(change_amount).replace(",", ".")))
        except Exception:
            pass

    if payment:
        out["Zahlung"] = payment

    # Kartenzahlung: keine Gegeben/Wechselgeld-Felder mitschleppen
    if out.get("Zahlung") == "Karte":
        out.pop("Gegeben (€)", None)
        wechsel = out.get("Wechselgeld (€)")
        if wechsel in (0, 0.0):
            out.pop("Wechselgeld (€)", None)

    # Barzahlung: nur wenn Gegeben + Wechselgeld wirklich da sind, Betrag daraus setzen
    if out.get("Zahlung") == "Bar":
        gegeben = out.get("Gegeben (€)")
        wechsel = out.get("Wechselgeld (€)")
        if isinstance(gegeben, (int, float)) and isinstance(wechsel, (int, float)):
            out["Betrag (€)"] = round(gegeben - wechsel, 2)

    # Rohtext immer mitgeben
    out["Rohtext"] = "\n".join(_txt(ln) for ln in lines if _txt(ln))

    # Gemischte Steuersätze -> globale MwSt-Felder nicht setzen
    raw_up = out.get("Rohtext", "").upper()

    has_7 = bool(re.search(r"\b7\s*%", raw_up)) or "A 7" in raw_up or "7,00" in raw_up
    has_19 = bool(re.search(r"\b19\s*%", raw_up)) or "B 19" in raw_up or "19,00" in raw_up

    if has_7 and has_19:
        out.pop("MwSt %", None)
        out.pop("MwSt (€)", None)
        out.pop("Netto (€)", None)

    return out

def extract_data_with_paddle(image_path: str) -> dict:
    """
    Führt PaddleOCR + Blockparser aus und liefert ein Ergebnis-Dict.
    """
    lines = run_paddle_ocr(image_path)
    return parse_receipt_blocks(lines)
#-----------Paddle Code Ende ---------------------

def _clean_head_lines(text: str, n_lines: int = 15) -> list[str]:
    """
    Nimmt die ersten n Zeilen, filtert Müll (kurze Tokens, viel Sonderzeichen),
    normalisiert Whitespaces und Groß/Kleinschreibung.
    """
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()][:max(5, n_lines)]
    out = []
    for ln in lines:
        # Sonderzeichen-Blobs weg
        bad = sum(ch.isalpha() for ch in ln) < 3  # zu wenig Buchstaben
        too_many_sym = sum(not (ch.isalnum() or ch in " .,-/&()+*") for ch in ln) > 3
        if bad or too_many_sym:
            continue
        # Mehrere Spaces normalisieren
        ln = " ".join(ln.split())
        ln = _fix_spaced_letters(ln)
        out.append(ln)
    return out

def _best_store_from_head(head_lines: list[str]) -> str | None:
    """
    Pickt eine plausible Ladenzeile:
    - enthält mindestens 3 aufeinander folgende Buchstaben
    - ist nicht nur generische Wörter („KASSE“, „SUMME“, „EINKAUF“)
    - bevorzugt Zeilen mit bekannten Kettennamen
    """
    if not head_lines:
        return None

    generic_block = {"kasse","summe","einkauf","kundenbeleg","quittung","bon","rechnung",
                     "ust","mwst","tse","terminal","girocard","karte"}
    known_brands = ["aldi","lidl","norma","netto","rewe","edeka","kaufland","penny",
                    "rossmann","dm","bft","aral","shell","esso","jet","total","q1",
                    "apotheke","norma","hagebau","obi","bauhaus","toom","ikea"]

    def score(ln: str) -> int:
        s = 0
        low = ln.lower()
        # lange Wörter belohnen
        if any(len(tok) >= 4 for tok in low.split()):
            s += 2
        # Marke drin?
        if any(b in low for b in known_brands):
            s += 5
        # Adresse? (Straße/Str./Strasse/PLZ)
        if "str" in low or "straße" in low or any(ch.isdigit() for ch in low):
            s += 1
        # generische Wörter bestrafen
        if any(g in low for g in generic_block):
            s -= 2
        # mindestens 3 Buchstaben in Folge?
        if not any(tok.isalpha() and len(tok) >= 3 for tok in low.split()):
            s -= 3
        return s

    candidates = [(score(ln), ln) for ln in head_lines]
    candidates.sort(reverse=True)
    best = candidates[0][1] if candidates and candidates[0][0] > 0 else None

    # Apotheke-Spezialfall: „APOTHEKE“ + nächste Orts-/Centerzeile mergen
    if best and best.lower().startswith("apotheke") and len(head_lines) >= 2:
        nxt = head_lines[1]
        if nxt and nxt.lower() not in generic_block and len(nxt) >= 5:
            best = f"{best} {nxt}"

    # NORMA/LIDL/… gern upper lassen, sonst Titel-Case
    if best and best.isupper():
        return best
    return best.title() if best else None


import re

def _fmt_money(v):
    if v is None: return ""
    try:
        return f"{float(v):.2f}".replace(".", ",")
    except:
        return str(v)

def _parse_date_loose(text):
    """
    Sucht ein Datum in typischen Kassenbon-Formaten und gibt 'YYYY-MM-DD' zurück.
    Unterstützt u. a.:
      - 07.08.2025
      - 07.08.25   (→ 2025)
      - 2025-08-07 / 2025/08/07
      - 07,08,25   (→ 07.08.25 → 2025)
      - 07. Aug 2025 / 07. August 25 (dt. Monatsnamen)
    """
    import re

    if not text:
        return None

    # 1) leichte Normalisierung: Komma zwischen Ziffern als Punkt deuten
    s = re.sub(r'(?<=\d),(?=\d)', '.', str(text))

    # 2) ISO: YYYY-MM-DD oder YYYY/MM/DD
    m = re.search(r'\b(20\d{2})[.\-\/](0?[1-9]|1[0-2])[.\-\/](0?[1-9]|[12]\d|3[01])\b', s)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return f"{y:04d}-{mo:02d}-{d:02d}"

    # 3) Deutsch: DD.MM.YYYY
    m = re.search(r'\b(0?[1-9]|[12]\d|3[01])[.](0?[1-9]|1[0-2])[.](20\d{2})\b', s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return f"{y:04d}-{mo:02d}-{d:02d}"

    # 4) Deutsch: DD.MM.YY  -> 2000 + YY
    m = re.search(r'\b(0?[1-9]|[12]\d|3[01])[.](0?[1-9]|1[0-2])[.](\d{2})\b', s)
    if m:
        d, mo, y2 = int(m.group(1)), int(m.group(2)), int(m.group(3))
        y = 2000 + y2
        return f"{y:04d}-{mo:02d}-{d:02d}"

    # 5) Deutsch mit Monatsnamen: "07. Aug 2025", "7 August 25", "07.Aug.25"
    s_lower = s.lower()
    # einfache Umlaut-Normalisierung für Matching
    s_lower = s_lower.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")

    monats_map = {
        "jan":1, "januar":1,
        "feb":2, "februar":2,
        "maerz":3, "märz":3, "mrz":3, "mar":3, "maerz":3,
        "apr":4, "april":4,
        "mai":5,
        "jun":6, "juni":6,
        "jul":7, "juli":7,
        "aug":8, "august":8,
        "sep":9, "sept":9, "september":9,
        "okt":10, "oktober":10,
        "nov":11, "november":11,
        "dez":12, "dezember":12,
    }

    # DD. MON YYYY / DD MON YY / DD.MON.YY etc.
    m = re.search(r'\b(0?[1-9]|[12]\d|3[01])[.\s]+([a-z]{3,10})[.\s]+(20\d{2}|\d{2})\b', s_lower)
    if m:
        d = int(m.group(1))
        mon_raw = m.group(2)
        ytok = m.group(3)
        mo = monats_map.get(mon_raw, None)
        if mo:
            y = int(ytok)
            if y < 100:
                y = 2000 + y
            return f"{y:04d}-{mo:02d}-{d:02d}"

    return None

def _parse_time_loose(text: str) -> str | None:
    """
    Holt Uhrzeiten HH:MM, ignoriert Zeilen mit EUR/€, %, Dezimal-Komma,
    Datumsmuster UND Öffnungszeiten/Wochentage/„Uhr“/„bis“.
    """
    import re
    TIME = re.compile(r"\b([01]?\d|2[0-3]):([0-5]\d)\b")
    BADKW = re.compile(
        r"(?i)\b(EUR|€|betrag|summe|mwst|ust|netto|brutto|wechselgeld|gegeben|girocard|ust-?id|ta-?nr|bnr|filiale|service)\b"
    )
    DATE = re.compile(r"\b\d{1,2}[.,/-]\d{1,2}[.,/-]\d{2,4}\b")  # z.B. 07,08,25
    PCT  = re.compile(r"%")
    HOURS= re.compile(r"(?i)\b(öffnungs|uhr|bis|von|mo\.?|di\.?|mi\.?|do\.?|fr\.?|sa\.?|so\.?|montag|dienstag|mittwoch|donnerstag|freitag|samstag|sonntag)\b")

    for ln in (ln.strip() for ln in text.splitlines() if ln.strip()):
        # alles mit Geld, Datum, %, Dezimal-Komma, Öffnungszeiten raus
        if BADKW.search(ln) or DATE.search(ln) or PCT.search(ln) or re.search(r"\d,\d{2}", ln) or HOURS.search(ln):
            continue
        m = TIME.search(ln)
        if m:
            try:
                h = int(m.group(1))
                if 0 <= h <= 23:
                    return f"{h:02d}:{m.group(2)}"
            except:
                pass
    return None

def _parse_money_de(s):
    """
    Akzeptiert '109,75', '109,75 €', 'EUR 109,75', '1.234,56'.
    Gibt float oder None zurück.
    """
    if not s:
        return None
    s = s.strip()
    m = re.search(r"(\d{1,3}(?:[.\s]\d{3})*|\d+)[,\.]\d{2}", s)
    if not m:
        return None
    num = m.group(0)
    num = num.replace(" ", "").replace(".", "").replace(",", ".")
    try:
        return float(num)
    except:
        return None

# Falls noch nicht vorhanden:
def _all_money_candidates(text: str) -> list[float]:
    vals = []
    for m in re.finditer(r"(?<!\d)(\d{1,3}(?:[.,]\d{3})*[.,][0-9]{2})(?:\s*(?:EUR|€))?", text):
        v = _norm_money(m.group(1))
        if v is not None:
            vals.append(v)
    uniq = []
    for v in sorted(vals, reverse=True):
        if not any(abs(v - u) < 0.001 for u in uniq):
            uniq.append(v)
    return uniq

def review_and_correct(data: dict, text: str) -> dict | None:
    """Tkinter-Dialog: Felder prüfen/korrigieren. Gibt korrigiertes dict oder None zurück."""
    try:
        import tkinter as tk
        from tkinter import ttk
    except Exception as e:
        print("Hinweis: Tkinter nicht verfügbar, speichere ohne GUI. Grund:", e)
        return data

    money_options = _all_money_candidates(text)
    money_options_str = [_fmt_money(v) for v in money_options]
    current_amt = data.get("Betrag (€)")

    root = tk.Tk()
    root.title("Kassenbon prüfen & korrigieren")
    root.minsize(700, 600)
    root.geometry("760x640+120+80")
    root.attributes("-topmost", True)
    root.grid_rowconfigure(0, weight=1)
    root.grid_columnconfigure(0, weight=1)

    # Scrollbereich
    canvas = tk.Canvas(root, borderwidth=0)
    vsb = ttk.Scrollbar(root, orient="vertical", command=canvas.yview)
    frame = ttk.Frame(canvas, padding=14)
    frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0,0), window=frame, anchor="nw")
    canvas.configure(yscrollcommand=vsb.set)
    canvas.grid(row=0, column=0, sticky="nsew"); vsb.grid(row=0, column=1, sticky="ns")

    ttk.Label(frame, text="Erkannte Felder prüfen/korrigieren",
              font=("Segoe UI", 12, "bold")).grid(row=0, column=0, columnspan=3, pady=(0,12), sticky="w")

    # --- Felder + Vars definieren ---
    defs = [
        ("Datum",        data.get("Datum", "")),
        ("Uhrzeit",      data.get("Uhrzeit", "")),
        ("Laden",        data.get("Laden", "")),
        ("Liter",        "" if data.get("Liter") is None else _fmt_money(data.get("Liter"))),
        ("€/L",          "" if data.get("€/L") is None else _fmt_money(data.get("€/L"))),
        ("Betrag (€)",   "" if current_amt is None else _fmt_money(current_amt)),
        ("Zahlung",      data.get("Zahlung", "")),
        ("MwSt %",       "" if data.get("MwSt %") is None else _fmt_money(data.get("MwSt %"))),
        ("MwSt (€)",     "" if data.get("MwSt (€)") is None else _fmt_money(data.get("MwSt (€)"))),
        ("Netto (€)",    "" if data.get("Netto (€)") is None else _fmt_money(data.get("Netto (€)"))),
        ("Säule",        data.get("Säule", "")),
        ("Kassierer",    data.get("Kassierer", "")),
        ("Beleg-Nr.",    data.get("Beleg-Nr.", "")),
    ]
    vars_map: dict[str, tk.StringVar] = {}
    widgets: dict[str, tk.Widget] = {}

    r = 1
    for label, val in defs:
        ttk.Label(frame, text=label+":", width=16).grid(row=r, column=0, sticky="e", padx=(0,10), pady=4)
        var = tk.StringVar(value=str(val))
        vars_map[label] = var
        if label == "Betrag (€)":
            # editierbare Combobox mit Vorschlägen
            cb = ttk.Combobox(frame, textvariable=var, values=money_options_str, width=36, state="normal")
            cb.grid(row=r, column=1, sticky="w", pady=4)
            widgets[label] = cb
        else:
            ent = ttk.Entry(frame, textvariable=var, width=38)
            ent.grid(row=r, column=1, sticky="w", pady=4)
            widgets[label] = ent
        r += 1

    # Buttons unten
    btn_frame = ttk.Frame(root, padding=(12,8))
    btn_frame.grid(row=1, column=0, columnspan=2, sticky="e")
    result = {"ok": False, "values": None}

    def on_ok():
        # --- Werte VOR destroy einsammeln ---
        vals = {k: v.get().strip() for k, v in vars_map.items()}
        result["values"] = vals
        result["ok"] = True
        root.destroy()

    def on_cancel():
        result["ok"] = False
        result["values"] = None
        root.destroy()

    ttk.Button(btn_frame, text="Abbrechen", command=on_cancel).pack(side="right", padx=8)
    ttk.Button(btn_frame, text="OK & speichern", command=on_ok).pack(side="right")

    root.mainloop()
    if not result["ok"]:
        return None

    vals = result["values"] or {}

    # --- Normalisieren: Zahlenfelder parsen, Textfelder direkt übernehmen ---
    out = dict(data)

    # Textfelder 1:1 (leer => alten Wert behalten)
    for key in ["Datum","Uhrzeit","Laden","Zahlung","Säule","Kassierer","Beleg-Nr."]:
        val = vals.get(key, "")
        out[key] = val if val else data.get(key)

    # Zahlenfelder tolerant parsen (leere Eingaben => None)
    out["Liter"]       = _parse_money_de(vals.get("Liter",""))
    out["€/L"]         = _parse_money_de(vals.get("€/L",""))
    out["Betrag (€)"]  = _parse_money_de(vals.get("Betrag (€)",""))
    out["MwSt %"]      = _parse_money_de(vals.get("MwSt %",""))
    out["MwSt (€)"]    = _parse_money_de(vals.get("MwSt (€)",""))
    out["Netto (€)"]   = _parse_money_de(vals.get("Netto (€)",""))

    return out

def extract_data_restaurant(text: str) -> dict:
    import re
    data = {}

    # Datum/Zeit
    d = _parse_date_loose(text)
    if d: data["Datum"] = d
    t = _parse_time_loose(text)
    if t: data["Uhrzeit"] = t

    # Laden (Kajute/Restaurant)
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    for ln in lines[:10]:
        low = ln.lower()
        if any(k in low for k in ["restaurant","kajute","kajüte","imbiss","gaststätte","bistro","trattoria","ristorante","café","cafe","gast"]):
            data["Laden"] = ln; break
    if "Laden" not in data:
        for ln in lines[:6]:
            if re.match(r"^(#|\d|mw?st|ust|netto|brutto|summe|gesamt|total|tisch)", ln, re.IGNORECASE):
                continue
            if len(ln) >= 3: data["Laden"] = ln; break

    # MwSt/UST in % & €
    foot = "\n".join(lines[-30:])
    m_pct = re.search(r"(?:mwst|ust|steuer)[^\d]{0,12}([0-9]{1,2}[.,]?[0-9]{0,2})\s*%", foot, re.IGNORECASE)
    if m_pct:
        try: data["MwSt %"] = float(m_pct.group(1).replace(",", "."))
        except: pass
    m_tax = re.search(r"(?:mwst|ust|steuer)[^\d]{0,15}([0-9]{1,3}(?:[.,]\d{3})*[.,]\d{2})", foot, re.IGNORECASE)
    if m_tax:
        data["MwSt (€)"] = float(m_tax.group(1).replace(".", "").replace(",", "."))

    # Netto
    m_net = re.search(r"\bnett[oa]?[^\d]{0,20}([0-9]{1,3}(?:[.,]\d{3})*[.,]\d{2})", foot, re.IGNORECASE)
    if m_net:
        data["Netto (€)"] = float(m_net.group(1).replace(".", "").replace(",", "."))

    # Summe / Rechnungsbetrag / zu zahlen / Total / (… Karte)
    total_roi = _roi_lines(text, r"(rechnungsbetrag|summe|gesamt|zu\s*zahlen|total|bar|kar?te)", window=4)
    m_total = re.search(r"(?:rechnungsbetrag|summe|gesamt|zu\s*zahlen|total)[^\d]{0,20}([0-9]{1,3}(?:[.,]\d{3})*[.,]\d{2})",
                        total_roi, re.IGNORECASE | re.DOTALL)
    if m_total:
        data["Betrag (€)"] = float(m_total.group(1).replace(".", "").replace(",", "."))
    if not data.get("Betrag (€)"):
        data["Betrag (€)"] = pick_total(text, None, None)

    # Zahlungsart (KARTE/BAR/EC)
    pay = re.search(r"(?i)\b(girocard|ec[- ]?karte|ec|karte|kreditkarte|barzahlung|bar)\b", text)
    if pay:
        p = pay.group(1).lower()
        if "bar" in p:       data["Zahlung"] = "Bar"
        elif "giro" in p:    data["Zahlung"] = "girocard"
        elif "ec" in p:      data["Zahlung"] = "EC-Karte"
        elif "karte" in p:   data["Zahlung"] = "Karte"

    return data

def extract_data_fuel(text: str) -> dict:
    import re
    out = {}
    s = text
    lines = [ln.strip() for ln in s.splitlines() if ln.strip()]

    # --- Datum / Uhrzeit ---
    #if '._parse_date_loose' in dir():  # defensive, falls umbenannt
    #    d = _parse_date_loose(s)
    #    if d: out["Datum"] = d
    #if '._parse_time_loose' in dir():
    #    t = _parse_time_loose(s)
    #    if t: out["Uhrzeit"] = t
    #if "Uhrzeit" not in out:
    #    m = re.search(r'(?<!\d)([01]?\d|2[0-3])[:.;h ]([0-5]\d)(?::([0-5]\d))?', s)
    #    if m:
    #        out["Uhrzeit"] = f"{int(m.group(1)):02d}:{m.group(2)}"

    # --- Laden (Marke/Ort) ---
    head = lines[:15]
    brand_pat = re.compile(r'\b(ARAL|SHELL|ESSO|JET|TOTAL(?:E)?|AVIA|Q1|OMV|BFT|HEM|STAR|AGIP|ENI|TAMOIL|NILSEN|NIELSEN|SCAN-?SHOP)\b', re.I)
    city_pat  = re.compile(r'\b\d{4,5}\s+([A-ZÄÖÜ][A-Za-zÄÖÜäöüß\-]+)\b')

    brand_line, brand, city = None, None, None
    for ln in head:
        m = brand_pat.search(ln)
        if m:
            brand = m.group(1).upper()
            brand_line = ln
            break
    for ln in head:
        m = city_pat.search(ln)
        if m:
            city = m.group(1)
            break

    if brand_line:
        out["Laden"] = brand_line
    elif brand and city:
        out["Laden"] = f"{brand} {city}"
    elif city:
        out["Laden"] = f"Tankstelle {city}"

    # --- Liter (Menge) – spalten-/zeilenbasiert ---
    val, _ = _find_value_after_label(
        lines, r'\b(Menge|Abgabe|Abgabemenge|Volumen|Liter)\b', window=3,
        value_regex=r'\d{1,3}(?:[.,]\d{3})*[.,]\d{1,3}'
    )
    if val is None:
        m = re.search(r'(\d{1,4}(?:[.,]\d{1,3})?)\s*(?:L|Liter)\b', s, re.I)
        if m:
            val = m.group(1)
    if val is not None:
        v = _float_de(val)
        if v is not None:
            out["Liter"] = v

    # --- €/L – spalten-/zeilenbasiert ---
    val, _ = _find_value_after_label(
        lines, r'(€/L|EUR/L|Preis\s*(?:je|/)\s*(?:L|Liter)|Preis pro Liter|Einzelpreis)\b', window=3,
        value_regex=r'\d{1,2}[.,]\d{2,3}'
    )
    if val is None:
        m = (re.search(r'([0-9]{1,2}[.,][0-9]{2,3})\s*(?:€|EUR)?\s*/\s*(?:L|Liter)\b', s, re.I) or
             re.search(r'(?:EUR/?L|€/L)\s*[:=]?\s*([0-9]{1,2}[.,][0-9]{2,3})', s, re.I))
        if m:
            val = m.group(1)
    if val is not None:
        v = _float_de(val)
        if v is not None:
            out["€/L"] = v

    # --- Betrag (€) – nur echte Betrag-/Summe-/Brutto-Zeilen oder mit EUR/€ ---
    val, _ = _find_value_after_label(
        lines, r'\b(Betrag|Summe|Gesamt|Brutto|Rechnungsbetrag|Endbetrag)\b', window=3,
        value_regex=r'\d{1,3}(?:[.,]\d{3})*[.,]\d{2}'
    )
    if val is None:
        m = re.search(r'(?im)^\s*(?:.*?(Betrag|Summe|Gesamt|Brutto).{0,20})?'
                      r'(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})\s*(?:€|EUR)\b', s)
        if m:
            val = m.group(2)
    if val is not None:
        v = _float_de(val)
        if v is not None:
            out["Betrag (€)"] = round(v, 2)

    # --- Korrektur: Falls Betrag versehentlich die Liter-Menge ist ---
    if out.get("Betrag (€)") is not None and out.get("Liter") is not None:
        if abs(out["Betrag (€)"] - out["Liter"]) <= 0.02 and out.get("€/L"):
            est = round(out["Liter"] * out["€/L"] + 1e-6, 2)
            out["Betrag (€)"] = est

    # --- Wenn Betrag fehlt, aber Liter & €/L vorhanden: berechnen ---
    if out.get("Betrag (€)") in (None, "") and out.get("Liter") and out.get("€/L"):
        out["Betrag (€)"] = round(out["Liter"] * out["€/L"] + 1e-6, 2)

    # --- MwSt % / MwSt € / Netto (€) ---
    if "MwSt %".lower() not in (k.lower() for k in out.keys()):
        m = re.search(r'(?i)\b(19|7)[,\.]?\s*0{0,2}\s*%', s)
        if m:
            try:
                out["MwSt %"] = float(m.group(1).replace(',', '.'))
            except Exception:
                pass

    if out.get("Betrag (€)") and out.get("MwSt %"):
        r = out["MwSt %"] / 100.0
        net = round(out["Betrag (€)"] / (1.0 + r), 2)
        vat = round(out["Betrag (€)"] - net, 2)
        out.setdefault("Netto (€)", net)
        out.setdefault("MwSt (€)", vat)

    # --- Zahlung ---
    m = re.search(r'\b(bar|girocard|ec|karte|kreditkarte|visa|mastercard|amex|kontaktlos|lidl\s*pay|apple\s*pay|google\s*pay)\b', s, re.I)
    if m:
        p = m.group(1).lower()
        if p == "bar":
            out["Zahlung"] = "Bar"
        else:
            out["Zahlung"] = "Karte"
    # ===== PDF-/A4-Tabellenfall (Tankrechnung / Karten-Nr. / Tabellen-Header) =====
    # ===== PDF-/A4-Tabellen: Tankrechnung mit "St. Nettobetrag Mwst-Satz Steuerbetrag Bruttobetrag" =====
    try:
        # Nur aktivieren, wenn typische Tankrechnungs-Signale vorhanden sind
        if re.search(r"(?i)(Tankrechnung|Karten-?Nr|Artikel\s+Datum/?Uhrzeit\s+Standort\s+Menge|ZG\s+Tankstelle)", text):
            money = r"(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})"

            # 1) Laden / Standort
            m_site = re.search(r"(?i)\bZG\s+Tankstelle\s+([A-Za-zÄÖÜäöüß\- ]+)", text)
            if m_site:
                out["Laden"] = f"ZG Tankstelle {m_site.group(1).strip()}"

            # 2) Menge (Liter)
            m_qty = re.search(r"(?i)\bMenge\s+" + money + r"\b", text)
            if m_qty:
                try: out["Liter"] = _norm_money(m_qty.group(1))
                except: pass

            # 3) Tabellenzeile mit 5 Spalten: St. / Nettobetrag / Mwst-Satz / Steuerbetrag / Bruttobetrag
            #    Beispiel: "19  71,25  19,00  13,54  84,79"
            m_line = re.search(
                r"(?im)^\s*(?:St\.?\s+)?(\d{1,2})\s+{m}\s+(\d{1,2}(?:[.,]\d{1,2})?)\s+{m}\s+{m}\s*$"
                .format(m=money),
                text
            )
            if m_line:
                try:
                    # Gruppen:
                    # 1: MwSt %    2: Netto     3: MwSt €     4: Brutto €
                    mwst_pct   = _norm_money(m_line.group(2))  # Achtung Reihenfolge: s.u. (siehe Regex)
                    netto_eur  = _norm_money(m_line.group(1))  # <= Diese zwei sind vertauscht? -> Korrigieren:
                    # Korrektur der Gruppen-Zuordnung:
                    # Regex-Gruppen-Indexe zählen alle Klammern. Daher weisen wir per Namen neu zu:
                except Exception:
                    pass
                # Sauber mit benannten Gruppen:
                m_named = re.search(
                    r"(?im)^\s*(?:St\.?\s+)?(?P<mwst_pct>\d{1,2})\s+"
                    r"(?P<netto>{m})\s+(?P<mwst_pct2>\d{1,2}(?:[.,]\d{1,2})?)\s+"
                    r"(?P<mwst_eur>{m})\s+(?P<brutto>{m})\s*$".format(m=money),
                    text
                )
                if m_named:
                    try:
                        out["MwSt %"]    = _norm_money(m_named.group("mwst_pct2"))  # der Satz mit Dezimalen
                    except: pass
                    try:
                        out["Netto (€)"] = _norm_money(m_named.group("netto"))
                    except: pass
                    try:
                        out["MwSt (€)"]  = _norm_money(m_named.group("mwst_eur"))
                    except: pass
                    try:
                        out["Betrag (€)"] = _norm_money(m_named.group("brutto"))    # **wichtig**: Brutto als Betrag
                    except: pass

            if not USE_PADDLE:
            # alte Fallbacks / Merge-Korrekturen
                # 4) Endbetrag-Fallback (falls obige Zeile doch nicht greift)
                if "Betrag (€)" not in out or not out["Betrag (€)"]:
                    m_end = re.search(r"(?i)\bEndbetrag[^\d]{0,20}" + money, text)
                    if m_end:
                        try: out["Betrag (€)"] = _norm_money(m_end.group(1))
                        except: pass
                    else:
                        # Zeile mit drei Beträgen und "*Endbetrag" am Ende
                        m_end2 = re.search(
                            r"(?i)^{m}\s+{m}\s+({m})\s*\*?Endbetrag\s*$".format(m=money),
                            text, re.M
                        )
                        if m_end2:
                            try: out["Betrag (€)"] = _norm_money(m_end2.group(1))
                            except: pass

            # 5) Netto/MwSt-Überschriften (falls Variante mit Spaltentitel)
            if "Netto (€)" not in out:
                m_net = re.search(r"(?i)\bNettobetrag\s*€?\s*"+money, text)
                if m_net:
                    try: out["Netto (€)"] = _norm_money(m_net.group(1))
                    except: pass
            if "MwSt (€)" not in out:
                m_tax = re.search(r"(?i)\bSteuerbetrag\s*"+money, text)
                if m_tax:
                    try: out["MwSt (€)"] = _norm_money(m_tax.group(1))
                    except: pass

            # 6) €/L berechnen (falls Liter vorhanden)
            if (not out.get("€/L")) and out.get("Liter") and out.get("Betrag (€)"):
                try:
                    out["€/L"] = round(out["Betrag (€)"] / max(0.001, out["Liter"]), 3)
                except:
                    pass

            # 7) Zahlungsart (Karte)
            if not out.get("Zahlung"):
                if re.search(r"(?i)\b(Karten-?Nr|Kartenzahlung|girocard|EC|Kreditkarte|KARTE)\b", text):
                    out["Zahlung"] = "KARTE"
    except Exception:
        pass

    return out


def extract_data_fuel_pdf(text: str) -> dict:
    """
    Parser für Tankkarten-PDFs (z. B. Honeck-Waldschütz / ZG Tankstelle).
    Erwartet reinen Text aus extract_text_from_pdf(...).
    Liefert u. a.: Datum, Uhrzeit, Laden, Betrag (Brutto), MwSt %, MwSt (€),
                  Netto (€), Zahlung, Liter (Menge), €/L.
    """
    import re

    out = {}

    # ---------- kleine Helfer ----------
    def _nm(s: str) -> float | None:
        """Money-Parser mit globalem _norm_money-Fallback."""
        try:
            if "_norm_money" in globals():
                return globals()["_norm_money"](s)
            # Fallback
            s = (s or "").strip()
            s = s.replace(" ", "")
            s = s.replace(".", "").replace(",", ".")
            return float(s)
        except Exception:
            return None

    def _to_iso_date(d: str) -> str | None:
        # "15,09,2025" | "15.09.2025" | "15.9.25"
        m = re.match(r"^\s*(\d{1,2})[.,](\d{1,2})[.,](\d{2,4})\s*$", d)
        if not m:
            return None
        dd, mm, yy = m.groups()
        dd = int(dd); mm = int(mm); yy = int(yy)
        if yy < 100:
            yy += 2000
        if not (1 <= dd <= 31 and 1 <= mm <= 12 and 1900 <= yy <= 2100):
            return None
        return f"{yy:04d}-{mm:02d}-{dd:02d}"

    def _first(pattern, s, flags=0):
        return re.search(pattern, s, flags)

    # ---------- Laden / Standort ----------
    # z. B. "ZG Tankstelle Bretten"
    m_store = re.search(r"(?im)^\s*(ZG\s+Tankstelle\s+[^\n\r]+)\s*$", text)
    if m_store:
        out["Laden"] = m_store.group(1).strip()
    else:
        # Sonst: „Tankstelle <Ort>“ oder „ZG ...“
        m_store2 = re.search(r"(?im)^\s*(?:ZG\s+)?Tankstelle\s+[^\n\r]+$", text)
        if m_store2:
            out["Laden"] = m_store2.group(0).strip()

    # ---------- Datum & Uhrzeit ----------
    # Bevorzugt Tank-Zeitpunkt (z. B. "13,09,2025 16:01")
    m_dt = re.search(r"(?s)\b(\d{1,2}[.,]\d{1,2}[.,]\d{2,4})\s+(\d{1,2}:\d{2})\b", text)
    if m_dt:
        iso = _to_iso_date(m_dt.group(1))
        if iso:
            out["Datum"] = iso
        out["Uhrzeit"] = m_dt.group(2)

    # Falls nur Rechnungsdatum existiert (z. B. "Rg.-Datum 15,09,2025"), ggf. Datum setzen
    if "Datum" not in out:
        m_rg = _first(r"(?i)rg\.-?\s*datum[^\d]{0,10}(\d{1,2}[.,]\d{1,2}[.,]\d{2,4})", text)
        if m_rg:
            iso = _to_iso_date(m_rg.group(1))
            if iso:
                out["Datum"] = iso

    # ---------- Brutto / Netto / MwSt ----------
    # Bruttobetrag / Endbetrag
    m_brutto = _first(r"(?i)\b(Bruttobetrag|Endbetrag)\b[^\d]{0,15}([0-9]+(?:[.,][0-9]{2}))", text)
    if m_brutto:
        v = _nm(m_brutto.group(2))
        if v is not None:
            out["Betrag (€)"] = v

    # Netto
    m_netto = _first(r"(?i)\bNettobetrag\b[^\d]{0,15}([0-9]+(?:[.,][0-9]{2}))", text)
    if m_netto:
        v = _nm(m_netto.group(1))
        if v is not None:
            out["Netto (€)"] = v

    # MwSt-Satz
    m_mw = _first(r"(?i)\b(?:MwSt|UST|mwst-?satz)\b[^\d]{0,10}([0-9]{1,2}(?:[.,][0-9]{1,2})?)", text)
    if m_mw:
        v = _nm(m_mw.group(1))
        if v is not None:
            out["MwSt %"] = v

    # Steuerbetrag (MwSt €)
    m_tax = _first(r"(?i)\bSteuerbetrag\b[^\d]{0,15}([0-9]+(?:[.,][0-9]{2}))", text)
    if m_tax:
        v = _nm(m_tax.group(1))
        if v is not None:
            out["MwSt (€)"] = v

    # ---------- Menge (Liter) / Nettopreis je 100 ----------
    qty = None
    np100 = None  # Nettopreis je 100 L

    # (A) Primär: Zeile mit "ZG Tankstelle ..." – dort stehen i.d.R. drei Zahlen am Ende
    line_site = None
    m_line = re.search(r"(?im)^[^\n]*\bZG\s+Tankstelle\s+[^\n]*$", text)
    if m_line:
        line_site = m_line.group(0)
        nums = re.findall(r"([0-9]{1,3}(?:[.,][0-9]{3})*[.,][0-9]{2,3})", line_site)
        # Erwartet: [Menge, Nettopreis100, Nettobetrag]
        if len(nums) >= 2:
            q = _nm(nums[0])
            p100 = _nm(nums[1])
            if q is not None:
                qty = q
            if p100 is not None:
                np100 = p100

    # (B) Sekundär: „Menge <xx,xx>“ – auch wenn die Zahl in der nächsten Zeile steht
    if qty is None:
        m_qty_inline = re.search(r"(?i)\bMenge\b[^\n\r]*?([0-9]{1,3}(?:[.,][0-9]{2,3}))", text)
        if m_qty_inline:
            qty = _nm(m_qty_inline.group(1))
        else:
            # „Menge“ in einer Zeile, Zahl als nächste eigenständige Zahl in den folgenden 1–2 Zeilen
            m_qty_block = re.search(
                r"(?is)\bMenge\b.*?\n\s*([0-9]{1,3}(?:[.,][0-9]{2,3}))\s*(?:€|EUR)?\s*(?:\n|$)",
                text
            )
            if m_qty_block:
                qty = _nm(m_qty_block.group(1))

    # (C) Tertiär: 3-Zeilen-Muster (Menge / Nettopreis100 / Nettobetrag) – wie in deinem Dump
    if qty is None or np100 is None:
        m_triplet = re.search(
            r"(?m)^\s*([0-9]{1,3}(?:[.,][0-9]{2}))\s*(?:€|EUR)?\s*$\s*"
            r"^\s*([0-9]{1,3}(?:[.,][0-9]{3}))\s*$\s*"
            r"^\s*([0-9]{1,3}(?:[.,][0-9]{2}))\s*(?:€|EUR)?\s*$",
            text
        )
        if m_triplet:
            q = _nm(m_triplet.group(1))    # Menge
            p100 = _nm(m_triplet.group(2)) # Nettopreis je 100
            # m_triplet.group(3) wäre Nettobetrag (optional)
            if qty is None and q is not None:
                qty = q
            if np100 is None and p100 is not None:
                np100 = p100

    if qty is not None:
        out["Liter"] = round(float(qty), 3)

    # ---------- €/L bestimmen ----------
    # a) bevorzugt aus Brutto / Liter
    if out.get("Betrag (€)") and out.get("Liter"):
        try:
            out["€/L"] = round(float(out["Betrag (€)"]) / float(out["Liter"]), 3)
        except Exception:
            pass
    # b) sonst aus Nettopreis je 100 + MwSt
    elif np100 is not None:
        try:
            per_l_netto = float(np100) / 100.0
            mw = out.get("MwSt %")
            if isinstance(mw, (int, float)):
                out["€/L"] = round(per_l_netto * (1.0 + float(mw)/100.0), 3)
            else:
                out["€/L"] = round(per_l_netto, 3)
        except Exception:
            pass

    # ---------- Zahlungsart ----------
    # SEPA/Lastschrift-Hinweise
    if re.search(r"(?i)(SEPA|Lastschrift|wird.*abgebucht|Einzug)", text):
        out["Zahlung"] = "LASTSCHRIFT"
    elif re.search(r"(?i)\b(girocard|karte|kreditkarte|ec)\b", text):
        out["Zahlung"] = "KARTE"

    # ---------- Kraftstoff (optional) ----------
    # Nimm den ersten Artikel-Namen aus der Artikelzeile (z. B. "Super" / "Super E10" / "Diesel")
    m_fuel = re.search(r"(?im)^\s*([A-ZÄÖÜa-zäöü0-9 .+-/]+)\s+\d{1,2}[.,]\d{1,2}[.,]\d{2,4}\s+\d{1,2}:\d{2}\s+ZG\s+Tankstelle", text)
    if m_fuel:
        name = m_fuel.group(1).strip()
        # etwas putzen
        name = re.sub(r"\s{2,}", " ", name)
        out["Kraftstoff"] = name

    return out

# =================== Merge-Helfer ===================

def _keep_nonempty(dst: dict, src: dict, keys: list[str]):
    """
    Kopiert Werte aus src -> dst, überschreibt aber NIE mit None oder "".
    Bevorzugt die erste gefüllte Ausprägung über alle Varianten.
    """
    for k in keys:
        v_new = src.get(k)
        v_old = dst.get(k)
        if v_new not in (None, ""):
            dst[k] = v_new
        elif v_old not in (None, ""):
            dst[k] = v_old  # behalten

def _flatten_texts(seq) -> list[str]:
    out: list[str] = []
    def _add(x):
        if x is None: return
        if isinstance(x, (bytes, bytearray)):
            try: x = x.decode("utf-8", "ignore")
            except: return
        if isinstance(x, str):
            s = x.strip()
            if s: out.append(s)
        elif isinstance(x, (list, tuple)):
            for t in x: _add(t)
        # andere Typen ignorieren
    _add(seq)
    return out

def _sanitize_dict_values(d: dict) -> dict:
    """List/Tuple/Dict -> String, damit später keine unhashbaren Typen auftreten."""
    out = {}
    for k, v in d.items():
        if isinstance(v, (list, tuple)):
            out[k] = ", ".join(str(x) for x in v)
        elif isinstance(v, dict):
            out[k] = str(v)
        else:
            out[k] = v
    return out


def merge_variant_dicts(dicts: list[dict], receipt_type: str = "generic") -> dict:
    out: dict = {}

    # Rohtexte zusammenführen – nur Strings zulassen
    raw_items = [d.get("Rohtext","") for d in dicts if d.get("Rohtext") is not None]
    combined_text = "\n".join(_flatten_texts(raw_items))
    if combined_text:
        out["Rohtext"] = combined_text

    # Varianten-Dicts sanitizen (Listen -> Strings)
    clean = []
    for d in dicts:
        if isinstance(d, dict):
            clean.append(_sanitize_dict_values(d))
    dicts = clean

    # 1) Kernfelder: niemals mit None/"" überschreiben
    CORE_KEYS = ["Datum", "Uhrzeit", "Laden", "Betrag (€)", "Gegeben (€)", "Wechselgeld (€)", "Zahlung"]
    for d in dicts:
        _keep_nonempty(out, d, CORE_KEYS)

    # 2) Komponenten für Steuersätze einsammeln (erste sinnvolle Werte je Feld)
    def _num(x):
        return x if isinstance(x, (int, float)) else None

    comp_keys = ["Netto 7% (€)", "MwSt 7% (€)", "Netto 19% (€)", "MwSt 19% (€)"]
    comp = {k: None for k in comp_keys}
    for d in dicts:
        for k in comp_keys:
            if comp[k] is None:
                v = _num(d.get(k))
                if v is not None and v >= 0:
                    comp[k] = v

    # 3) Gesamt-Netto / Gesamt-MwSt defensiv bilden (nur aus Komponenten)
    sum_net, sum_vat = 0.0, 0.0
    has_net = False
    has_vat = False

    for k in ("Netto 7% (€)", "Netto 19% (€)"):
        v = _num(comp[k])
        if v is not None:
            sum_net += v
            has_net = True

    for k in ("MwSt 7% (€)", "MwSt 19% (€)"):
        v = _num(comp[k])
        if v is not None:
            sum_vat += v
            has_vat = True

    # nur setzen, wenn wirklich vorhanden – nichts aus "beliebigen" Zahlen basteln
    if has_net and out.get("Netto (€)") in (None, ""):
        out["Netto (€)"] = round(sum_net, 2)
    if has_vat and out.get("MwSt (€)") in (None, ""):
        out["MwSt (€)"] = round(sum_vat, 2)

    # 4) MwSt % nur setzen, wenn GENAU eine Rate erkennbar ist
    rates = []
    if _num(comp["Netto 7% (€)"]) is not None or _num(comp["MwSt 7% (€)"]) is not None:
        rates.append(7.0)
    if _num(comp["Netto 19% (€)"]) is not None or _num(comp["MwSt 19% (€)"]) is not None:
        rates.append(19.0)

    if len(rates) == 1:
        out["MwSt %"] = rates[0]
    elif "MwSt %" in out:  # <-- hier war vorher der Tippfehler ["MwSt %"]
        # Plausibilitätscheck: Zahl und im vernünftigen Bereich
        try:
            v = float(out["MwSt %"])
            if not (0 < v <= 25):
                out.pop("MwSt %", None)
        except Exception:
            out.pop("MwSt %", None)

    # 5) Sanfte Plausibilitätsbremse (gegen Ausreißer)
    total = _num(out.get("Betrag (€)"))
    def _cap(name, upper_factor):
        v = _num(out.get(name))
        if v is None:
            return
        if v < 0:
            out[name] = None
        elif total and v > total * upper_factor:
            out[name] = None

    # MwSt typischerweise < 20 % vom Brutto; Netto <= Brutto.
    _cap("MwSt (€)", 0.20)
    if total and _num(out.get("Netto (€)")) and out["Netto (€)"] > total:
        out["Netto (€)"] = None

    # 6) Belegtyp durchreichen (falls schon in Varianten vergeben)
    for d in dicts:
        if d.get("Belegtyp"):
            out["Belegtyp"] = d["Belegtyp"]
            break
    if "Belegtyp" not in out:
        out["Belegtyp"] = receipt_type or "generic"

    return out


def detect_receipt_type(text: str, store_hint: str | None = None) -> str:
    """
    Schätzt den Belegtyp ('grocery', 'fuel', 'pharmacy', 'restaurant', 'retail', 'generic')
    anhand von Schlagwörtern. Robust: alle Scores werden vorab initialisiert.
    Optionaler store_hint (z. B. 'LIDL', 'ZG Tankstelle ...') kann den Score boosten.
    """
    import re

    t = (text or "")
    up = t.upper()

    # --- Scores IMMER initialisieren ---
    score = {
        "grocery": 0,
        "fuel": 0,
        "pharmacy": 0,
        "restaurant": 0,
        "retail": 0,
        "generic": 0,
    }

    # --- Helper ---
    def count(pat: str, flags=re.IGNORECASE) -> int:
        try:
            return len(re.findall(pat, t, flags))
        except Exception:
            return 0

    def bump(cat: str, n: int = 1):
        score[cat] += n

    # --- Grocery / Discounter ---
    grocery_brand = r"\b(LIDL|ALDI|EDEKA|REWE|KAUFLAND|NORMA|PENNY|NETTO|NIELSEN|SCAN-?SHOP|ROSSMANN|DM)\b"
    grocery_tokens = r"(EINKAUFSWERT|ARTIKEL|PFAND|FILIALE|KUNDENBELEG|UST-?ID|BON[- ]?NR|SUMME|GESAMT)"
    bump("grocery", count(grocery_brand))
    bump("grocery", count(grocery_tokens))

    # --- Fuel / Tankstelle ---
    fuel_tokens = r"(ZAPF|SÄULE|ZAPFSÄULE|SUPER|DIESEL|E10\b|E5\b|€/L|EURO/?L|LITER\b|TANKSTELLE|TANKRECHNUNG)"
    bump("fuel", count(fuel_tokens))

    # --- Pharmacy ---
    pharmacy_tokens = r"(APOTHEKE|REZEPT|ZUZ(A|Ä)HLUNG|PZN\b|APOTHEKEN\s*OHG)"
    bump("pharmacy", count(pharmacy_tokens))

    # --- Restaurant ---
    restaurant_tokens = r"(TISCH|BEDIENUNG|TRINKGELD|SPEISEN|GETRÄNKE|RESTAURANT|SERVIEREND(E|IN))"
    bump("restaurant", count(restaurant_tokens))

    # --- Retail (Baumarkt / sonstiger Einzelhandel) ---
    retail_brand = r"(OBI|BAUHAUS|HAGEBAU|HORNBACH|MEDIA ?MARKT|SATURN|IKEA|CONRAD)"
    retail_tokens = r"(ARTIKELNR|STK\b|KASSE|KASSIERER|BON|WARENR|KUNDENNR)"
    bump("retail", count(retail_brand))
    bump("retail", count(retail_tokens))

    # --- Store-Hint (falls aus Kopfzeile erkannt) ---
    if store_hint:
        su = store_hint.upper()
        if re.search(grocery_brand, su):
            bump("grocery", 3)
        if re.search(r"\b(TANK|TANKSTELLE|ARAL|SHELL|JET|ESSO|TOTAL)\b", su):
            bump("fuel", 3)
        if re.search(r"\bAPOTHEK", su):
            bump("pharmacy", 3)
        if re.search(r"\b(RESTAURANT|RISTORANTE|TRATTORIA|GASTST[AÄ]TTE)\b", su):
            bump("restaurant", 3)

    # Wenn wirklich gar nichts passt, minimaler generic-Score (damit nicht alles 0 ist)
    if all(v == 0 for v in score.values()):
        score["generic"] = 1

    # Beste Kategorie wählen (bei Gleichstand einfache Präferenz)
    order = ["fuel", "grocery", "pharmacy", "restaurant", "retail", "generic"]
    best = max(order, key=lambda k: (score[k], -order.index(k)))

    return best

def extract_text_from_pdf(pdf_path: str) -> str:
    """
    Liest Text direkt aus einer PDF. Liefert leeren String, wenn nichts Brauchbares drin ist.
    """
    text_parts = []

    try:
        import fitz
        doc = fitz.open(pdf_path)
        for page in doc:
            txt = page.get_text("text") or ""
            if txt.strip():
                text_parts.append(txt)
        doc.close()
    except Exception:
        pass

    text = "\n".join(text_parts).strip()
    if len(text) < 40:
        return ""

    # Qualitätscheck: zu viele Ein-Zeichen-Zeilen -> unbrauchbarer Direkttext
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    if not lines:
        return ""

    short_lines = [ln for ln in lines if len(ln) <= 2]
    ratio_short = len(short_lines) / max(1, len(lines))

    # Wenn mehr als 60% der Zeilen nur 1-2 Zeichen haben: verwerfen
    if ratio_short > 0.6:
        return ""

    # Wenn fast keine normalen Wörter vorkommen: verwerfen
    long_word_lines = [ln for ln in lines if len(ln) >= 6 and any(c.isalpha() for c in ln)]
    if len(long_word_lines) < 3:
        return ""

    return text

def render_pdf_page_to_image(pdf_path: str, page_index: int = 0, zoom: float = 2.0) -> str | None:
    """
    Rendert die erste PDF-Seite als PNG und gibt den Bildpfad zurück.
    """
    try:
        import fitz
        import os

        doc = fitz.open(pdf_path)
        if page_index >= len(doc):
            doc.close()
            return None

        page = doc[page_index]
        mat = fitz.Matrix(zoom, zoom)
        pix = page.get_pixmap(matrix=mat, alpha=False)

        out_path = os.path.splitext(pdf_path)[0] + f"__page{page_index+1}.png"
        pix.save(out_path)
        doc.close()
        return out_path
    except Exception as e:
        print(f"⚠️ PDF-Render fehlgeschlagen: {e}")
        return None


def extract_data_from_text_only(raw_text: str) -> dict:
    """
    Einfacher, allgemeiner Textparser für direkt aus PDF gelesenen Text.
    Kein Laden-Hardcoding nötig.
    """
    import re

    out = {"Rohtext": raw_text}
    txt = raw_text or ""
    txt_up = txt.upper()

    lines = [ln.strip() for ln in txt.splitlines() if ln.strip()]

    # -------------------------
    # Brand zuerst aus Gesamtdokument
    # -------------------------
    brand = None
    brand_patterns = [
        (r"\bKAUFLAND\b", "KAUFLAND"),
        (r"\bLIDL\b", "LIDL"),
        (r"\bALDI\b", "ALDI"),
        (r"\bREWE\b", "REWE"),
        (r"\bEDEKA\b", "EDEKA"),
        (r"\bPENNY\b", "PENNY"),
        (r"\bHAGEBAU\b", "hagebaumarkt"),
        (r"\bBAUHAUS\b", "BAUHAUS"),
        (r"\bOBI\b", "OBI"),
        (r"\bHORNBACH\b", "HORNBACH"),
        (r"\bAPOTHEKE\b", "APOTHEKE"),
        (r"\bARAL\b", "ARAL"),
        (r"\bSHELL\b", "SHELL"),
        (r"\bJET\b", "JET"),
        (r"\bESSO\b", "ESSO"),
        (r"\bTOTAL\b", "TOTAL"),
        (r"\bBFT\b", "bft"),
    ]

    for pat, label in brand_patterns:
        if re.search(pat, txt_up):
            brand = label
            break

    # -------------------------
    # Adresse aus Kopf, Laden = Brand zuerst
    # -------------------------
    address = None

    def _looks_like_address(s: str) -> bool:
        s_up = s.upper()
        if re.search(r"\b\d{5}\b", s_up):
            return True
        if re.search(r"\b(STRASSE|STRAßE|STR\.|WEG|PLATZ|ALLEE|HÖHE|HOEHE)\b", s_up):
            return True
        if re.search(r"\bTEL\b", s_up):
            return True
        return False

    for ln in lines[:15]:
        if _looks_like_address(ln):
            address = ln
            break

    if brand:
        out["Laden"] = brand
    elif address:
        out["Laden"] = address
    if address:
        out["Adresse"] = address

    # -------------------------
    # Datum / Uhrzeit
    # -------------------------
    m = re.search(r"(\d{2}\.\d{2}\.\d{2,4}).*?(\d{2}:\d{2})(?::\d{2})?", txt)
    if m:
        dd, mm, yy = re.match(r"(\d{2})\.(\d{2})\.(\d{2,4})", m.group(1)).groups()
        if len(yy) == 2:
            yy = "20" + yy
        out["Datum"] = f"{yy}-{mm}-{dd}"
        out["Uhrzeit"] = m.group(2)
    else:
        m_date = re.search(r"\b(\d{2}\.\d{2}\.\d{2,4})\b", txt)
        m_time = re.search(r"\b(\d{2}:\d{2})(?::\d{2})?\b", txt)

        if m_date:
            dd, mm, yy = re.match(r"(\d{2})\.(\d{2})\.(\d{2,4})", m_date.group(1)).groups()
            if len(yy) == 2:
                yy = "20" + yy
            out["Datum"] = f"{yy}-{mm}-{dd}"

        if m_time:
            out["Uhrzeit"] = m_time.group(1)

    # -------------------------
    # Fuel-/Rechnungs-spezifische Erkennung
    # -------------------------
    is_fuel_like = False
    if re.search(r"\b(TANKRECHNUNG|TANKSTELLE|SUPER|DIESEL|EUR/L|LITER|KRAFTSTOFF)\b", txt_up):
        is_fuel_like = True

    if is_fuel_like:
        out["Belegtyp"] = "fuel"

        # Laden
        m_store = re.search(r"(ZG\s+TANKSTELLE[^\n\r]*)", txt, flags=re.I)
        if m_store:
            out["Laden"] = m_store.group(1).strip()
        elif re.search(r"HONECK-WALDSCHÜTZ ENERGIE", txt_up):
            out["Laden"] = "Honeck-Waldschütz Energie GmbH"

        # Kraftstoffart
        if re.search(r"\bSUPER(\s*E10|\s*E5| BLFR\.)?\b", txt_up):
            out["Kraftstoffart"] = "Super"
        elif re.search(r"\bDIESEL\b", txt_up):
            out["Kraftstoffart"] = "Diesel"

        # Datum / Uhrzeit aus Tankzeile bevorzugen
        m_dt = re.search(r"(\d{2}\.\d{2}\.\d{2,4})\s+(\d{2}:\d{2})", txt)
        if m_dt:
            dd, mm, yy = re.match(r"(\d{2})\.(\d{2})\.(\d{2,4})", m_dt.group(1)).groups()
            if len(yy) == 2:
                yy = "20" + yy
            out["Datum"] = f"{yy}-{mm}-{dd}"
            out["Uhrzeit"] = m_dt.group(2)

        # Zahlung
        if re.search(r"\bABGEBUCHT\b", txt_up) or re.search(r"\bIBAN\b", txt_up):
            out["Zahlung"] = "Lastschrift"

        # Betrag: Endbetrag / Bruttobetrag bevorzugen
        m_end = re.search(r"(?is)(ENDBETRAG|BRUTTOBETRAG)[^\d]{0,40}(\d{1,4}[.,]\d{2})", txt)
        if m_end:
            try:
                out["Betrag (€)"] = float(m_end.group(2).replace(",", "."))
            except Exception:
                pass
        else:
            m_sum = re.search(r"(?is)\bSUMME\b[^\d]{0,40}(\d{1,4}[.,]\d{2})", txt)
            if m_sum:
                try:
                    out["Betrag (€)"] = float(m_sum.group(1).replace(",", "."))
                except Exception:
                    pass

    # -------------------------
    # Zahlung
    # -------------------------
    # "Pay" / Karte hat Vorrang vor Rückgeld
    if re.search(r"\b(LIDL PAY|KAUFLAND PAY|BLUECODE|PAYBACK PAY|APPLE PAY|GOOGLE PAY)\b", txt_up):
        out["Zahlung"] = "Karte"
    elif re.search(r"\b(KARTE|EC|GIROCARD|VISA|MASTERCARD)\b", txt_up):
        out["Zahlung"] = "Karte"
    elif re.search(r"\b(BAR|BARZAHLUNG)\b", txt_up):
        out["Zahlung"] = "Bar"
    elif re.search(r"\b(RÜCKGELD|RUCKGELD)\b", txt_up):
        # nur dann Bar annehmen, wenn NICHT gleichzeitig Pay/Karte vorkommt
        out["Zahlung"] = "Bar"

    # Falls sowohl Pay/Karte als auch Rückgeld vorkommen: Karte gewinnt
    if re.search(r"\b(LIDL PAY|KAUFLAND PAY|BLUECODE|KARTE|EC|GIROCARD|VISA|MASTERCARD)\b", txt_up):
        out["Zahlung"] = "Karte"

    # -------------------------
    # Betrag
    # -------------------------
    money_pat = r"(?<!\d)\d{1,4}[.,]\d{2}(?!\d)"

    # bevorzugt Summe / zu zahlen / Pay-Zeile
    m_total = re.search(
        rf"(?is)(zu zahlen|endbetrag|summe|gesamt|betrag|lidl pay|kaufland pay)[^\d]{{0,40}}({money_pat})",
        txt
    )
    if m_total:
        try:
            out["Betrag (€)"] = float(m_total.group(2).replace(",", "."))
        except Exception:
            pass

    # -------------------------
    # Belegtyp generisch
    # -------------------------
    fuel_hits = 0
    if re.search(r"\bDIESEL\b", txt_up):
        fuel_hits += 1
    if re.search(r"\bSUPER(\s*E10|\s*E5)?\b", txt_up):
        fuel_hits += 1
    if re.search(r"\bKRAFTSTOFF\b", txt_up):
        fuel_hits += 1
    if re.search(r"(€/L|EUR/L|LITER)", txt_up):
        fuel_hits += 1
    if re.search(r"\bZAPFSÄULE\b|\bSAULENNUMMER\b|\bSÄULE\b", txt_up):
        fuel_hits += 1
    if re.search(r"\bTANK(RECHNUNG|STELLE|DATEN)?\b", txt_up):
        fuel_hits += 1

    if fuel_hits >= 2:
        out["Belegtyp"] = "fuel"
    elif re.search(r"\b(APOTHEKE|PHARMA|REZEPT)\b", txt_up):
        out["Belegtyp"] = "pharmacy"
    else:
        article_like = 0
        money_pat = r"(?<!\d)\d{1,4}[.,]\d{2}(?!\d)"
        for ln in lines[:120]:
            if re.search(rf"{money_pat}\s*[AB]?$", ln) or re.search(rf"\d+\s*\*\s*{money_pat}", ln):
                article_like += 1

        has_sum = bool(re.search(r"\b(SUMME|ZU ZAHLEN|BETRAG|TOTAL)\b", txt_up))
        has_tax = bool(re.search(r"\b(STEUER|MWST|BRUTTO|NETTO)\b", txt_up))

        if article_like >= 5 and has_sum:
            out["Belegtyp"] = "grocery" if has_tax else "retail"
        else:
            out["Belegtyp"] = "generic"

def extract_data_from_pdf(pdf_path: str) -> dict | None:
    """
    PDF-Einstieg:
    1) Erst versuchen, brauchbaren PDF-Text direkt zu lesen
    2) Falls das nicht reicht: erste Seite rendern und mit PaddleOCR parsen
    """
    pdf_text = extract_text_from_pdf(pdf_path)

    if pdf_text:
        print("📄 PDF-Text direkt gelesen.")
        return extract_data_from_text_only(pdf_text)

    print("📄 PDF-Text unbrauchbar -> Render zu Bild + PaddleOCR")
    image_path = render_pdf_page_to_image(pdf_path, page_index=0, zoom=2.0)
    if not image_path:
        return None

    if USE_PADDLE:
        lines = run_paddle_ocr(image_path)
        result = parse_receipt_blocks(lines)
        result["Rohtext"] = result.get("Rohtext") or ""
        return result

    # Fallback, falls Paddle aus ist
    texts = ocr_text_multi(image_path)
    if not texts:
        return None
    return extract_data_from_text_only("\n".join(texts))

def extract_data_grocery(text: str) -> dict:
    import re
    data = {}

    # Datum/Uhrzeit
    d = _parse_date_loose(text);  t = _parse_time_loose(text) if '_parse_time_loose' in globals() else None
    if d: data["Datum"] = d
    if t: data["Uhrzeit"] = t

    # Laden
    if not data.get("Laden") and re.search(r"(?i)\bLIDL\b", text):
        data["Laden"] = "Lidl"

    # Adresse (heuristisch)
    m_addr = re.search(r"(?i)\b([A-ZÄÖÜa-zäöüß\- ]+str(?:aße|asse)\.?\s*\d+)\s*\n\s*(\d{5}\s+[A-ZÄÖÜa-zäöüß\- ]+)", text)
    if m_addr:
        street = " ".join(m_addr.group(1).split())
        city   = " ".join(m_addr.group(2).split())
        if data.get("Laden"):
            data["Laden"] = f'{data["Laden"]} – {street}, {city}'


    money = r"(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})"
    m = re.search(rf"(?i)\b(gegeben|bar(?:zahlung)?|gezahlt)\b[^\d]{{0,20}}{money}\s*(?:€|eur)?", text)
    if m: 
        try: data["Gegeben (€)"] = _norm_money(m.group(2))
        except: pass
    m = re.search(rf"(?i)\b(zurück|wechselgeld|rückgeld)\b[^\d]{{0,20}}{money}\s*(?:€|eur)?", text)
    if m:
        try: data["Wechselgeld (€)"] = _norm_money(m.group(2))
        except: pass

    if not USE_PADDLE:
    # alte Fallbacks / Merge-Korrekturen
        # Lidl hat häufig genau "Betrag 216,02 EUR" – mit dem fangen wir zusätzlich ab:
        if "Betrag" in text and "Betrag (€)" not in data:
            m2 = re.search(rf"(?im)\bBetrag\b[^\d]{{0,20}}{money}", text)
            if m2:
                try: data["Betrag (€)"] = _norm_money(m2.group(1))
                except: pass

    # Fußbereich: Zeilen mit % sammeln
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    footer_lines = [ln for ln in lines if "%" in ln or re.search(r"(?i)(mwst|ust|netto|brutto)", ln)]

    netto_sum = 0.0
    mwst_sum  = 0.0

    # Hilfsparser: finde in einer Zeile zwei Geldbeträge → (netto, mwst)
    def _two_amounts(line: str):
        am = re.findall(money, line)
        if len(am) >= 2:
            vals = [_norm_money(x) for x in am[:2]]
            vals = [v for v in vals if isinstance(v,(int,float))]
            if len(vals) >= 2:
                big  = max(vals[0], vals[1])
                small= min(vals[0], vals[1])
                return big, small  # (netto, mwst)
        return None

    # Für 7% / 19% separat versuchen
    for rate, key_net, key_vat in [
        (r"(?i)\b7[,\.]?\s*0{1,2}\s*%",   "Netto 7% (€)",  "MwSt 7% (€)"),
        (r"(?i)\b19[,\.]?\s*0{1,2}\s*%",  "Netto 19% (€)", "MwSt 19% (€)"),
    ]:
        hit = next((ln for ln in footer_lines if re.search(rate, ln)), None)
        if hit:
            pair = _two_amounts(hit)
            if pair:
                n, v = pair
                data[key_net] = n
                data[key_vat] = v
                netto_sum += n
                mwst_sum  += v

    # Falls noch keine Ratenzeilen erkannt wurden: generischer Fallback
    if netto_sum == 0 and mwst_sum == 0:
        # Nimm die 2–3 letzten Fußzeilen, suche nach zwei Beträgen
        for ln in footer_lines[-5:]:
            pair = _two_amounts(ln)
            if pair:
                n, v = pair
                netto_sum += n
                mwst_sum  += v

    if netto_sum > 0 and "Netto (€)" not in data:
        data["Netto (€)"] = round(netto_sum, 2)
    if mwst_sum > 0 and "MwSt (€)" not in data:
        data["MwSt (€)"] = round(mwst_sum, 2)

    # Wenn Betrag & MwSt da → Netto = Betrag - MwSt
    if data.get("Betrag (€)") is not None and data.get("MwSt (€)") is not None and "Netto (€)" not in data:
        try:
            data["Netto (€)"] = round(float(data["Betrag (€)"]) - float(data["MwSt (€)"]), 2)
        except Exception:
            pass

    # --- MwSt/Netto aus Tabellenzeilen (NORMA/Lidl/Aldi) robust ziehen ---
    import re
    money = r"(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})"
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

    def _norm(x):
        try:
            return _norm_money(x)
        except Exception:
            return None

    def _is_vat_line(ln: str) -> bool:
        # echte Steuerzeilen: haben % und (A|B) ODER MWST/USt-Schlüsselwort
        if "%" not in ln:
            return False
        if re.search(r"(?i)\b(MWST|MWSt|USt|UST|MEHRWERTSTEUER)\b", ln):
            return True
        if re.search(r"(?i)^[AB]\b", ln):
            return True
        return False

    def _extract_pct(ln: str):
        m = re.search(r"(?i)\b(\d{1,2}[.,]\d{1,2})\s*%", ln)
        if m:
            try:
                return float(m.group(1).replace(",", "."))
            except:
                return None
        return None

    # Wir erwarten pro Zeile mind. zwei Geldbeträge: Netto und MwSt
    vat_found = {}
    for ln in lines:
        if not _is_vat_line(ln):
            continue
        pct = _extract_pct(ln)  # z.B. 7.00 oder 19.00
        amts = re.findall(money, ln)
        if len(amts) < 2 or pct is None:
            continue
        v1 = _norm(amts[0]); v2 = _norm(amts[1])
        if v1 is None or v2 is None:
            continue
        # Heuristik: Netto > MwSt
        net, vat = (v1, v2) if v1 >= v2 else (v2, v1)
        # Merke die beste Sicht pro Satz (7 oder 19)
        if 6.0 <= pct <= 8.0:
            if "Netto 7% (€)" not in data:
                data["Netto 7% (€)"] = net
            if "MwSt 7% (€)" not in data:
                data["MwSt 7% (€)"] = vat
            vat_found[7] = True
        elif 18.0 <= pct <= 20.0:
            if "Netto 19% (€)" not in data:
                data["Netto 19% (€)"] = net
            if "MwSt 19% (€)" not in data:
                data["MwSt 19% (€)"] = vat
            vat_found[19] = True

    # Gesamtwerte nur bilden, wenn Komponenten da sind
    sum_net = 0.0; has_net = False
    sum_vat = 0.0; has_vat = False
    for k in ("Netto 7% (€)", "Netto 19% (€)"):
        if isinstance(data.get(k), (int,float)):
            sum_net += data[k]; has_net = True
    for k in ("MwSt 7% (€)", "MwSt 19% (€)"):
        if isinstance(data.get(k), (int,float)):
            sum_vat += data[k]; has_vat = True

    if has_net: data["Netto (€)"] = round(sum_net, 2)
    if has_vat: data["MwSt (€)"]  = round(sum_vat, 2)

    # MwSt % nur setzen, wenn GENAU eine Rate vorkommt
    if 7 in vat_found and 19 not in vat_found:
        data["MwSt %"] = 7.0
    elif 19 in vat_found and 7 not in vat_found:
        data["MwSt %"] = 19.0

    # --- Zahlung & Gegeben/Wechselgeld ---
    # Zahlung
    pay = data.get("Zahlung")
    
    if re.search(r"(?i)\b(l[i1]dl\s*pay|cn\s*pay|lid[il]\s*pay)\b", text):
        pay = "Lidl Pay"
    elif re.search(r"(?i)\b(girocard|ec-?karte|bankkarte)\b", text):
        pay = "KARTE"
    elif re.search(r"(?i)\b(kreditkarte|visa|mastercard|amex)\b", text):
        pay = "KREDITKARTE"
    elif re.search(r"(?i)\b(bar|barzahlung)\b", text):
        pay = "BAR"
    
    if pay:
        data["Zahlung"] = pay

    # Gegeben (z.B. „Girocard 47,66 EUR“ oder „Gegeben …“)
    m_given = re.search(rf"(?im)^(?:gegeben|girocard|ec(?:-?karte)?)\b[^\d]{{0,20}}{money}\s*(?:€|eur)?", text)
    if m_given:
        val = _norm(m_given.group(1 if m_given.lastindex == 1 else 2))
        if isinstance(val, (int,float)):
            data["Gegeben (€)"] = val

    # Wechselgeld
    m_change = re.search(rf"(?i)\b(zurück|wechselgeld|rückgeld)\b[^\d]{{0,20}}{money}\s*(?:€|eur)?", text)
    if m_change:
        val = _norm(m_change.group(2))
        if isinstance(val, (int,float)):
            data["Wechselgeld (€)"] = val

def extract_fuel_details_from_text(raw_text: str) -> dict:
    """
    Extrahiert nur tankstellen-spezifische Zusatzfelder.
    Basisdaten wie Laden, Datum, Uhrzeit, Betrag, Zahlung kommen aus dem allgemeinen Parser.
    """
    import re

    out = {}
    txt = raw_text or ""
    txt_up = txt.upper()

    # Kraftstoffart
    fuel_patterns = [
        (r"\bSUPER\s*E10\b", "Super E10"),
        (r"\bSUPER\b", "Super"),
        (r"\bDIESEL\b", "Diesel"),
        (r"\bULTIMATE\s*102\b", "Ultimate 102"),
        (r"\bULTIMATE\s*DIESEL\b", "Ultimate Diesel"),
        (r"\bV-?POWER\s*DIESEL\b", "V-Power Diesel"),
        (r"\bV-?POWER\b", "V-Power"),
    ]
    for pat, label in fuel_patterns:
        if re.search(pat, txt_up):
            out["Kraftstoffart"] = label
            break

    # Liter
    m_liters = re.search(r"\b(\d{1,3}[.,]\d{2,3})\s*(L|LTR|LITER)\b", txt_up)
    if m_liters:
        try:
            out["Liter"] = float(m_liters.group(1).replace(",", "."))
        except Exception:
            pass

    # Preis pro Liter
    m_price_per_l = re.search(r"\b(\d[.,]\d{3})\s*(EUR/L|€/L|E/L)\b", txt_up)
    if m_price_per_l:
        try:
            out["Preis/L"] = float(m_price_per_l.group(1).replace(",", "."))
        except Exception:
            pass

    # Fallback: Betrag ≈ Liter * Preis/L plausibilisieren nur, nicht überschreiben
    if "Liter" in out and "Preis/L" in out:
        try:
            out["Tankbetrag_calc"] = round(out["Liter"] * out["Preis/L"], 2)
        except Exception:
            pass

    return out

def refine_receipt_type_from_text(raw_text: str, current_type: str = "generic") -> str:
    """
    Schärft den Belegtyp anhand klarer Textmuster nach.
    Fuel hat Vorrang, wenn mehrere starke Hinweise vorkommen.
    """
    import re

    txt = (raw_text or "").upper()
    fuel_hits = 0

    if re.search(r"\bDIESEL\b", txt):
        fuel_hits += 1
    if re.search(r"\bSUPER(\s*E10|\s*E5)?\b", txt):
        fuel_hits += 1
    if re.search(r"(€/L|EUR/L|LITER)", txt):
        fuel_hits += 1
    if re.search(r"\b(TANKRECHNUNG|TANKSTELLE|KRAFTSTOFF)\b", txt):
        fuel_hits += 1
    if re.search(r"\b(SÄULE|SAULENNUMMER|ZAPFSÄULE)\b", txt):
        fuel_hits += 1

    if fuel_hits >= 2:
        return "fuel"

    if re.search(r"\b(APOTHEKE|PHARMA|REZEPT)\b", txt):
        return "pharmacy"

    return current_type or "generic"

def _status_from(best: dict) -> str:
    reasons = []

    # Betrag
    betrag = best.get("Betrag (€)")
    if not isinstance(betrag, (int, float)) or betrag <= 0:
        reasons.append("Betrag fehlt/ungültig")

    # Datum
    if not best.get("Datum"):
        reasons.append("Datum fehlt")

    # Uhrzeit (optional, aber oft sinnvoll)
    if not best.get("Uhrzeit"):
        reasons.append("Uhrzeit fehlt")

    # Laden
    if not best.get("Laden"):
        reasons.append("Laden fehlt")

    # MwSt
    mwst = best.get("MwSt %")
    if mwst not in (7, 19, None):
        reasons.append("MwSt unklar")

    # Spezialfall: Fuel ohne MwSt
    if best.get("Belegtyp") == "fuel" and mwst not in (7, 19):
        reasons.append("MwSt fehlt (Fuel)")

    # Ergebnis
    if not reasons:
        return "OK"

    return "Prüfen: " + ", ".join(reasons)

def finalize_and_save_receipt(best: dict, rtype: str, excel_path: str):
    """
    Setzt Prüfstatus, säubert Werte, prüft auf Duplikate und schreibt nach Excel.
    Gibt das finale Dict zurück.
    """
    pruefstatus = _status_from(best)
    best["Prüfstatus"] = "✅ OK" if pruefstatus.upper().startswith("OK") else f"🔎 {pruefstatus}"

    reviewed = best

    try:
        reviewed = _sanitize_dict_values(reviewed)
    except Exception:
        pass

    reviewed.setdefault("Belegtyp", rtype)

    if is_duplicate_receipt(reviewed, excel_path):
        print("⚠️ Duplikat erkannt – nicht erneut gespeichert.")
        return reviewed

    try:
        append_to_excel_typed(reviewed, reviewed.get("Belegtyp", rtype), excel_path=excel_path)
        print(f"✅ Gespeichert nach: {excel_path}")
    except Exception as e:
        print(f"⚠️ Excel-Schreiben fehlgeschlagen: {e}")

    return reviewed
# =========================
# ENRICHER / NACHSCHÄRFUNG
# =========================

def enrich_fuel_data(best: dict):
    import re

    raw = best.get("Rohtext", "") or ""

    # 1) Spezialfelder
    try:
        fuel_spec = extract_fuel_details_from_text(raw) or {}
        for k, v in fuel_spec.items():
            if v not in (None, "", 0):
                if k not in best or best.get(k) in (None, "", 0):
                    best[k] = v
    except Exception as e:
        print(f"⚠️ Fuel-Spezialparser fehlgeschlagen: {e}")

    # 2) Datum/Uhrzeit-Fallback
    if not best.get("Datum") or not best.get("Uhrzeit"):
        m = re.search(r"\b(\d{2}\.\d{2}\.\d{2,4})\s+(\d{2}:\d{2})\b", raw)
        if m:
            raw_date, raw_time = m.groups()

            md = re.match(r"(\d{2})\.(\d{2})\.(\d{2,4})", raw_date)
            if md:
                dd, mm, yy = md.groups()
                if len(yy) == 2:
                    yy = "20" + yy
                best.setdefault("Datum", f"{yy}-{mm}-{dd}")

            best.setdefault("Uhrzeit", raw_time)

    # 3) Steuer/Netto bei Fuel plausibel ergänzen
    try:
        brutto = best.get("Betrag (€)")
        mwst_prozent = best.get("MwSt %")

        # Fallback: MwSt-Satz direkt aus Rohtext ableiten
        if mwst_prozent not in (7, 19):
            if re.search(r"\b19\s*%", raw) or re.search(r"\b19[.,]00\b", raw):
                mwst_prozent = 19
            elif re.search(r"\b7\s*%", raw) or re.search(r"\b7[.,]00\b", raw):
                mwst_prozent = 7

        if isinstance(brutto, (int, float)) and mwst_prozent == 19:
            netto = round(brutto / 1.19, 2)
            mwst = round(brutto - netto, 2)

            best.setdefault("MwSt %", 19)

            if not isinstance(best.get("Netto (€)"), (int, float)):
                best["Netto (€)"] = netto

            if not isinstance(best.get("MwSt (€)"), (int, float)):
                best["MwSt (€)"] = mwst

    except Exception:
        pass

    return best

def enrich_grocery_data(best: dict):
    import re
    raw = best.get("Rohtext", "") or ""

    # Betrag robuster sichern (falls nötig)
    if not best.get("Betrag (€)"):
        matches = re.findall(r"(?:SUMME|ZU ZAHLEN)[^\d]{0,40}(\d{1,4}[.,]\d{2})", raw, re.I)
        if matches:
            vals = [float(x.replace(",", ".")) for x in matches]
            best["Betrag (€)"] = max(vals)

    return best

ENRICHERS = {
    "fuel": enrich_fuel_data,
    "grocery": enrich_grocery_data,
}

def apply_enrichers(best: dict, rtype: str | None = None) -> tuple[dict, str]:
    rtype = rtype or best.get("Belegtyp", "generic")

    if rtype in ENRICHERS:
        best = ENRICHERS[rtype](best)

    best["Belegtyp"] = refine_receipt_type_from_text(
        best.get("Rohtext", ""),
        best.get("Belegtyp", rtype)
    )
    rtype = best.get("Belegtyp", rtype)

    return best, rtype

def extract_best_datetime(text: str) -> tuple[str | None, str | None]:
    """
    Extrahiert ein plausibles Datum + Uhrzeit aus OCR-Text.
    Akzeptiert z. B.:
    - 27.02.26 07:55
    - 27,02,26 07:55
    - 7.02.2607:55
    - 05.03.2026 16:38 Uhr
    """
    import re

    if not text:
        return None, None

    candidates = re.findall(
        r"\b(\d{1,2}[.,]\d{2}[.,]\d{2,4})\s*(\d{2}:\d{2})\b",
        text
    )

    found = []

    for raw_date, raw_time in candidates:
        try:
            hh, mi = map(int, raw_time.split(":"))
            if hh > 23 or mi > 59:
                continue

            raw_date = raw_date.replace(",", ".")
            md = re.match(r"(\d{1,2})\.(\d{2})\.(\d{2,4})", raw_date)
            if not md:
                continue

            dd, mm, yy = md.groups()
            if len(dd) == 1:
                dd = "0" + dd
            if len(yy) == 2:
                yy = "20" + yy

            # einfache Plausibilität
            dd_i = int(dd)
            mm_i = int(mm)
            if not (1 <= dd_i <= 31 and 1 <= mm_i <= 12):
                continue

            found.append((f"{yy}-{mm}-{dd}", raw_time))
        except Exception:
            continue

    if not found:
        return None, None

    # meist steht das relevante Bon-Datum eher weit unten → letzten plausiblen Treffer nehmen
    return found[-1]

def fix_truncated_day(date_str: str | None, raw: str) -> str | None:
    """
    Korrigiert OCR-Fälle wie 7.02.26 -> 27.02.26,
    wenn im Rohtext ein zweistelliger Tag mit gleichem Monat/Jahr vorkommt.
    Erkennt sowohl:
    - 27.02.26
    - 27,02,26
    - 270226
    """
    import re

    if not date_str or not raw:
        return date_str

    try:
        yy, mm, dd = date_str.split("-")
        dd_i = int(dd)

        # nur bei kleinem Tag prüfen
        if dd_i >= 10:
            return date_str

        yy2 = yy[-2:]

        # 1) normales OCR-Datum suchen
        m1 = re.search(rf"\b(\d{{2}})[.,]{mm}[.,]{yy2}\b", raw)
        if m1:
            dd2 = int(m1.group(1))
            if dd2 >= 10:
                return f"{yy}-{mm}-{dd2:02d}"

        # 2) kompaktes Muster DDMMYY suchen, z. B. 270226
        m2 = re.search(rf"\b(\d{{2}}){mm}{yy2}\b", raw)
        if m2:
            dd2 = int(m2.group(1))
            if dd2 >= 10:
                return f"{yy}-{mm}-{dd2:02d}"

    except Exception:
        pass

    return date_str
# =========================
# HAUPTABLÄUFE / SCAN
# =========================

def scan_kassenbon(
    image_path: str,
    excel_path: str = "kassenbons.xlsx",
    review_when: str = "prüfen",
    strict_total: bool = True,
    show_debug_footer: bool = False
):
    import os
    import re

    _reset_payment_log_once()

    if not os.path.isfile(image_path):
        print(f"❗ Bild nicht gefunden: {image_path}")
        return None

    # PDF separat behandeln
    if image_path.lower().endswith(".pdf"):
        best = extract_data_from_pdf(image_path)
        if not best:
            print("❗ PDF konnte nicht verarbeitet werden.")
            return None

        # Mindestkompatibilität für den Rest
        texts = [best.get("Rohtext", "")]
        combined_text = best.get("Rohtext", "")
        rtype = best.get("Belegtyp", "generic")

        #print("\n=== DEBUG PDF RAW ===")
        #print(best)
    else:
        # =========================
        # OCR + Parsing
        # =========================
        if USE_PADDLE:
            lines = run_paddle_ocr(image_path)
            texts = [ln["text"] for ln in lines]
            combined_text = "\n".join(texts)

            best = parse_receipt_blocks(lines)

            #print("\n=== DEBUG PADDLE RAW ===")
            #print(best)

            if not best:
                print("❗ PaddleOCR lieferte kein Ergebnis.")
                return None

            best["Rohtext"] = combined_text

            # Belegtyp nur ergänzen, nicht überschreiben
            if not best.get("Belegtyp"):
                txt_up = combined_text.upper()

                if re.search(r"\b(HAGEBAU|OBI|BAUHAUS|HORNBACH)\b", txt_up):
                    best["Belegtyp"] = "retail"
                elif re.search(r"\b(BÄCKEREI|BACKHAUS|BACKSTUBE|KONDITOREI)\b", txt_up):
                    best["Belegtyp"] = "retail"
                elif re.search(r"\b(LIDL|ALDI|EDEKA|REWE|NORMA|NETTO|PENNY|KAUFLAND)\b", txt_up):
                    best["Belegtyp"] = "grocery"
                elif re.search(r"\b(APOTHEKE)\b", txt_up):
                    best["Belegtyp"] = "pharmacy"
                elif re.search(r"\b(ARAL|SHELL|JET|ESSO|TOTAL|BFT)\b", txt_up):
                    best["Belegtyp"] = "fuel"
                else:
                    best["Belegtyp"] = "generic"

            rtype = best.get("Belegtyp", "generic")

        else:
            texts = ocr_text_multi(image_path)
            if not texts:
                print("❗ Keine OCR-Texte erhalten.")
                return None

            combined_text = "\n".join(texts)

            parsed = []
            for tx in texts:
                try:
                    parsed.append(extract_data_from_text(tx) | {"Rohtext": tx})
                except Exception:
                    parsed.append({"Rohtext": tx})

            rtype = detect_receipt_type(combined_text) if 'detect_receipt_type' in globals() else "generic"
            best = merge_variant_dicts(parsed, receipt_type=rtype)

            # Nur im Tesseract-Zweig alte Fallbacks
            if not best.get("Betrag (€)"):
                tot = _pick_total_candidate(combined_text)
                if tot is not None:
                    best["Betrag (€)"] = tot

            if not best.get("Laden"):
                store_guess = _guess_store_name(combined_text)
                if store_guess:
                    best["Laden"] = store_guess

            pay = majority_payment(texts)
            if pay and (not best.get("Zahlung") or best["Zahlung"] in ("Unbekannt", "")):
                best["Zahlung"] = pay

        if not texts:
            print("❗ Keine OCR-Texte erhalten.")
            return None

    # =========================
    # Gemeinsame Nachbearbeitung
    # =========================
    combo_raw = "\n".join(texts)
    combo_text = safe_post_ocr_cleanup(combo_raw)

    if DEBUG_HEAD:
        print("\n--- CLEAN HEAD ---")
        for ln in combo_text.splitlines()[:20]:
            print(ln)

        print("\n--- CLEAN (letzte 40 Zeilen) ---")
        for ln in combo_text.splitlines()[-40:]:
            print(ln)

    # Laden nur säubern, nicht hart überschreiben
    clean_head = "\n".join([ln for ln in combo_text.splitlines() if ln.strip()][:20])
    best["Laden"] = _sanitize_store_name(best.get("Laden"), clean_head) or best.get("Laden")

    # Store-basierten Typ nur ergänzen, nicht blind überschreiben
    forced = coerce_type_by_store(best.get("Laden") or "")
    if forced and (not best.get("Belegtyp") or best.get("Belegtyp") == "generic"):
        best["Belegtyp"] = forced
        rtype = forced

    # Keine unbekannte Zahlung erzwingen
    if not best.get("Zahlung"):
        best["Zahlung"] = ""

    # 0-Beträge verwerfen
    if best.get("Betrag (€)") in (0, 0.0, "0", "0,00"):
        best.pop("Betrag (€)", None)

    # Sanitizer nur auf vorhandene Beträge anwenden
    if best.get("Betrag (€)"):
        best["Betrag (€)"] = _sanitize_total(best["Betrag (€)"], combo_text)

    # Typ nach Rohtext nachschärfen
    best["Belegtyp"] = refine_receipt_type_from_text(
        best.get("Rohtext", ""),
        best.get("Belegtyp", "generic")
    )

    #print("\n=== DEBUG VOR STATUS ===")
    #print(best)

    best, rtype = apply_enrichers(best, best.get("Belegtyp"))
    best = normalize_tax_fields(best)

    # Letzte Stabilisierung vor dem Speichern
    raw = best.get("Rohtext", "") or ""
    txt_up = raw.upper()

    if re.search(r"\bLIDL\b", txt_up):
        best["Laden"] = "LIDL"
    elif re.search(r"\bKAUFLAND\b", txt_up):
        best["Laden"] = "KAUFLAND"

    # 2) Unplausible Uhrzeit verwerfen
    t = best.get("Uhrzeit")
    if t:
        m_t = re.match(r"^(\d{2}):(\d{2})$", str(t))
        if m_t:
            hh, mi = int(m_t.group(1)), int(m_t.group(2))
            if hh > 23 or mi > 59:
                best["Datum"] = None
                best["Uhrzeit"] = None

    # 3) Datum/Uhrzeit robust neu ziehen
    if not best.get("Datum") or not best.get("Uhrzeit"):
        dt_date, dt_time = extract_best_datetime(raw)
        if dt_date and dt_time:
            dt_date = fix_truncated_day(dt_date, raw)
            best["Datum"] = dt_date
            best["Uhrzeit"] = dt_time

    # 4) Korrektur: führende Ziffer beim Tag fehlt (z. B. 7 statt 27)
    try:
        datum = best.get("Datum")
        if datum:
            yy, mm, dd = datum.split("-")
            dd_int = int(dd)

            # nur wenn Tag sehr klein ist → typischer OCR-Verlust
            if dd_int < 10:
                # im Rohtext nach zweistelligem Datum suchen
                m2 = re.search(rf"\b(\d{{2}})[.,]{mm}[.,]{yy[-2:]}", raw)
                if m2:
                    dd2 = int(m2.group(1))
                    if dd2 > 9:
                        best["Datum"] = f"{yy}-{mm}-{dd2:02d}"
    except Exception:
        pass

    return finalize_and_save_receipt(best, rtype, excel_path)

def scan_kassenbon_group(
    image_paths: list[str],
    excel_path: str ="kassenbons.xlsx",
    review_when: str = "prüfen",
    strict_total: bool = True,
    show_debug_footer: bool = False
):
    """
    Wie scan_kassenbon(), aber für mehrere Bildteile eines Bons.
    Kombiniert OCR-Texte aller Teile, dann gleicher Parsing-/Merge-/Excel-Flow.
    """
    import re, os
    
    _reset_payment_log_once()

    if not image_paths:
        print("❗ Keine Teilbilder übergeben.")
        return None
    for p in image_paths:
        if not os.path.isfile(p):
            print(f"❗ Teilbild fehlt: {p}")
            return None

    # 1) OCR ALLER TEILE
    texts = ocr_text_multi_many(image_paths)
    if not texts:
        print("❗ Keine OCR-Texte aus Teilbildern erhalten.")
        return None

    # Roh + Clean für den Gruppentext
    combo_raw_multi = "\n".join(texts)
    combined_text   = safe_post_ocr_cleanup(combo_raw_multi)

    # 2) COMBO (wie bei single)
    combo_raw  = "\n".join(texts)
    combo_text = safe_post_ocr_cleanup(combo_raw)
    #print("\n--- CLEAN HEAD (for store) ---")
    #for ln in [l for l in combo_text.splitlines() if l.strip()][:12]:
    #    print(ln)
    #cand_store = _guess_store_name_head(combo_text)
    #print(f"STORE-CANDIDATE: {cand_store!r}")


    if DEBUG_HEAD:
        print("\n--- CLEAN (letzte 25 Zeilen, multi) ---")
        for ln in combo_text.splitlines()[-25:]:
            print(ln)

    # 3) Parsing + Merge
    parsed = []
    for tx in texts:
        try:
            parsed.append(extract_data_from_text(tx) | {"Rohtext": tx})
        except Exception:
            parsed.append({"Rohtext": tx})

    rtype = detect_receipt_type(combo_text) if 'detect_receipt_type' in globals() else "generic"
    best  = merge_variant_dicts(parsed, receipt_type=rtype)

    # --- Store-basierter Typ-Override ---
    forced = coerce_type_by_store(best.get("Laden") or "")
    if forced:
        rtype = forced
    best["Belegtyp"] = rtype

    forced = coerce_type_by_store(best.get("Laden"))
    if forced:
        rtype = forced

    reviewed = best
    reviewed.setdefault("Belegtyp", rtype)
    #append_to_excel_typed(reviewed, rtype, excel_path=excel_path)

    # --- Store-basierter Override ---
    forced = coerce_type_by_store(best.get("Laden"))
    if forced:
        rtype = forced

    # --- Store-basierter Override ---
    forced = coerce_type_by_store(best.get("Laden"))
    if forced:
        rtype = forced

    # --- Store-basierter Typ-Override ---
    store = (best.get("Laden") or "").upper()
    if re.search(r"\b(NIELSEN|SCAN-?SHOP)\b", store):
        rtype = "grocery"  # erzwingen

    # 4) Laden-/Datum-/Uhrzeit-Nachschärfung (deine bestehenden Helfer nutzen)
    # Laden aus Kopf säubern
    clean_head = "\n".join([ln for ln in combo_text.splitlines() if ln.strip()][:20])
    if ' _sanitize_store_name' in globals():
        best["Laden"] = _sanitize_store_name(best.get("Laden"), clean_head) or best.get("Laden")

    # Laden fallback, falls leer
    if not best.get("Laden") or best["Laden"].strip().upper() in {"IN","DER","IN DER","APOTHEKE"}:
        store = _guess_store_name(combo_text) or _guess_store_name(combo_raw)
        if store:
            best["Laden"] = store

    # Datum/Uhrzeit fallback & validierung
    if not best.get("Datum"):
        d = _parse_date_loose(combo_text) or _parse_date_loose(combo_raw)
        if d: best["Datum"] = d
    if not best.get("Uhrzeit"):
        t = _parse_time_loose(combo_text)
        if t: best["Uhrzeit"] = t
    if '_validate_time_in_context' in globals():
        best["Uhrzeit"] = _validate_time_in_context(best.get("Uhrzeit"), combo_text)


    # Laden aus Kopf priorisieren (z.B. 'Nielsen SCAN-SHOP')
    store_from_head = _guess_store_name_head(combo_text)
    if store_from_head:
       # Wenn „Laden“ noch leer/unsinnig ist oder schlechtere Heuristik hatte: überschreiben
       if (not best.get("Laden")) or best["Laden"].strip().upper() in {"IN","DER","IN DER","APOTHEKE"} \
           or len(best.get("Laden","")) < 5:
            best["Laden"] = store_from_head

    if cand_store:
        best["Laden"] = cand_store  # erzwingen

    # Betrag an Zahlungszeile festnageln (nutzt deine bestehenden Helfer)
    try:
        import re
        MNY = _money_regex()
        m = re.search(rf"(?i)\b(girocard|gegeben|bar(?:zahlung)?|gezahlt)\b[^\d]{{0,40}}{MNY}\s*(?:€|EUR)?", combo_text)
        pay_ref = _norm_money(m.group(2)) if m else None
        given   = best.get("Gegeben (€)")
        total   = best.get("Betrag (€)")
        ref     = given if isinstance(given, (int,float)) else pay_ref
        if isinstance(ref, (int,float)):
            if not isinstance(total, (int,float)) or abs(ref - float(total)) >= 0.01:
                best["Betrag (€)"] = round(float(ref), 2)
    except Exception:
        pass

    apply_payment_detection(best, combo_text, debug=True)

    #print("\n=== DEBUG VOR STATUS 3 ===")
    #print(best)
    status = _status_from(best)
    best["Prüfstatus"] = "✅ OK" if status.upper().startswith("OK") else f"🔎 {status}"

    # 6) Optionaler Review
    reviewed = best
    #if ('review_and_correct' in globals()):
    #    want_review = (
    #        review_when.lower() == "immer" or
    #        (review_when.lower().startswith("prüf") and not status.upper().startswith("OK"))
    #    )
    #    if want_review:
    #        try:
    #            reviewed = review_and_correct(best) or best
    #        except Exception as e:
    #            print(f"⚠️ Review-Dialog nicht verfügbar/abgebrochen: {e}")

    # 7) Excel
    #try:
    #    rtype = detect_receipt_type(combo_text)
    #    reviewed.setdefault("Belegtyp", rtype)
    #    append_to_excel_typed(reviewed, rtype, excel_path=excel_path)
    #    print(f"✅ Gespeichert nach: {excel_path}")
    #except Exception as e:
    #    print(f"⚠️ Excel-Schreiben fehlgeschlagen: {e}")

    # --- Typ bestimmen & ggf. per Laden überschreiben ---
    try:
        rtype = detect_receipt_type(combined_text) if 'detect_receipt_type' in globals() else 'generic'
        forced = coerce_type_by_store(best.get("Laden") or "")
        if forced:
            rtype = forced
        best["Belegtyp"] = rtype
    except Exception:
        rtype = best.get("Belegtyp") or "generic"

    # --- Excel nur HIER (einmal) schreiben ---
    try:
        #print("\n=== DEBUG VOR EXCEL 2 ===")
        #print(best)
        append_to_excel_typed(best, rtype, excel_path=excel_path)
        print(f"✅ Gespeichert nach: {excel_path}")
    except Exception as e:
        print(f"⚠️ Excel-Schreiben (group) fehlgeschlagen: {e}")

    return reviewed


def batch_scan(folder: str, excel_path=EXCEL_PATH, review_when="prüfen"):
    exts = (".jpg",".jpeg",".png",".tif",".tiff",".bmp",".webp")
    files = [os.path.join(folder,f) for f in os.listdir(folder) if f.lower().endswith(exts)]
    for f in sorted(files):
        res = scan_kassenbon(f, excel_path=excel_path)  # mit GUI
        # falls du ohne GUI willst: baue scan_* leicht um oder setze einen Flag
        # wenn Prüfstatus == 'prüfen': erneut mit Review öffnen:
        #if res and res.get("Prüfstatus") == review_when:
        #    res2 = review_and_correct(res, res.get("Rohtext",""))
        #    if res2:
        #        #print("\n=== DEBUG VOR EXCEL 3 ===")
        #        #print(best)
        #        append_to_excel(res2, excel_path=excel_path)

# =================== Batch-Scan (kompletter Block) ===================

from datetime import datetime, date

def _normalize_de_date_inplace(dct: dict, key="Datum"):
    """Datum in TT.MM.JJJJ normalisieren (in-place)."""
    d = dct.get(key)
    if isinstance(d, (datetime, date)):
        dct[key] = d.strftime("%d.%m.%Y")
    elif isinstance(d, str) and d.strip():
        for fmt in ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d"):
            try:
                dct[key] = datetime.strptime(d, fmt).strftime("%d.%m.%Y")
                break
            except:
                pass

def scan_kassenbon_batch(
    folder: str,
    excel_path: str = "kassenbons.xlsx",
    review_when: str = "prüfen",
    strict_total: bool = True,
    move_processed: bool = True,
    verbose: bool = True,
):
    """
    Verarbeitet alle Bilder im Ordner mit scan_kassenbon(...).
    Verschiebt Dateien nach _ok / _reviewed / _error (optional).
    Gibt eine kleine Zusammenfassung zurück.
    """
    import os, shutil

    exts = (".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp", ".webp")
    files = [os.path.join(folder, f) for f in os.listdir(folder) if f.lower().endswith(exts)]
    files.sort()

    if not files:
        print("📂 Keine passenden Bilddateien im Ordner gefunden.")
        return {"total": 0, "ok": 0, "pruefen": 0, "error": 0, "saved": 0}

    if verbose:
        print(f"📂 Starte Batch-Scan: {folder}  |  Dateien: {len(files)}")
        print(f"   review_when={review_when}  strict_total={strict_total}  move_processed={move_processed}")

    # Zielordner optional anlegen
    ok_dir = os.path.join(folder, "_ok")
    review_dir = os.path.join(folder, "_reviewed")
    error_dir = os.path.join(folder, "_error")
    if move_processed:
        for d in (ok_dir, review_dir, error_dir):
            os.makedirs(d, exist_ok=True)

    total = saved = okc = rc = ec = 0

    for image_path in files:
        fname = os.path.basename(image_path)
        if verbose:
            print(f"\n📄 Verarbeite: {fname}")

        try:
            # *** EINZIGER OCR/Parse/Excel-Aufruf im Batch ***
            result = scan_kassenbon(image_path, excel_path=excel_path)
            total += 1

            if not result:
                raise RuntimeError("scan_kassenbon lieferte kein Ergebnis")

            # Prüfstatus interpretieren
            status = (result.get("Prüfstatus") or "").lower()
            is_ok = status == "ok" or "✅" in status
            needs_review = ("prüf" in status) or ("review" in status)

            # Datei verschieben
            if move_processed:
                try:
                    if is_ok:
                        shutil.move(image_path, os.path.join(ok_dir, fname))
                    elif needs_review:
                        shutil.move(image_path, os.path.join(review_dir, fname))
                    else:
                        shutil.move(image_path, os.path.join(error_dir, fname))
                except Exception as mv_e:
                    print(f"⚠️ Verschieben fehlgeschlagen: {mv_e}")

            # Zähler
            saved += 1
            if is_ok:
                okc += 1
            elif needs_review:
                rc += 1
            else:
                ec += 1

        except Exception as e:
            total += 1
            ec += 1
            print(f"❌ Fehler bei {image_path}: {e}")
            if move_processed:
                try:
                    shutil.move(image_path, os.path.join(error_dir, fname))
                except Exception as mv_e:
                    print(f"⚠️ Verschieben in _error fehlgeschlagen: {mv_e}")

    # Zusammenfassung
    print("\n📊 Batch-Summary")
    print(f"  Total:   {total}")
    print(f"  Saved:   {saved}")
    print(f"  OK:      {okc}")
    print(f"  Prüfen:  {rc}")
    print(f"  Error:   {ec}")

    return {"total": total, "ok": okc, "pruefen": rc, "error": ec, "saved": saved}

#===================Mehrteilige Bons zusammensetzen=============================
import re
from collections import defaultdict

_PART_SUFFIX_RE = re.compile(
    r"""(?ix)           # case-insensitive, verbose
    (?:                 # akzeptierte Suffix-Muster vor der Extension:
        _[a-z]          #  _a, _b, ...
      | _\d+            #  _1, _2, ...
      | [\-\.\s]?part\d+#  -part1, .part2,  part3
    )$
    """
)

def _splitext_lower(path: str):
    import os
    root, ext = os.path.splitext(path)
    return root, ext.lower()

def _normalize_basename_for_grouping(filepath: str) -> str:
    """Nimmt den Dateinamen ohne Extension und entfernt Teil-Suffixe (_a/_1/-part2)."""
    import os
    root, ext = _splitext_lower(filepath)
    base = os.path.basename(root)
    # Suffix nur am ENDE entfernen
    base = _PART_SUFFIX_RE.sub("", base)
    # vollständiger Basisschlüssel inkl. Ordner (damit gleiche Namen in anderen Ordnern separat bleiben)
    return os.path.join(os.path.dirname(filepath), base)

def _part_order_key(filepath: str):
    """Sortierschlüssel: a,b,c… dann Zahlen; sonst Originalname."""
    root, _ = _splitext_lower(filepath)
    m_alpha = re.search(r"_([a-z])$", root, re.I)
    if m_alpha:
        return (0, ord(m_alpha.group(1).lower()) - ord('a'))
    m_num = re.search(r"_(\d+)$", root)
    if m_num:
        return (1, int(m_num.group(1)))
    m_part = re.search(r"(?:^|[\-\.\s])part(\d+)$", root, re.I)
    if m_part:
        return (2, int(m_part.group(1)))
    # Fallback: Name sortieren
    import os
    return (9, os.path.basename(filepath).lower())

def group_receipt_parts(filepaths: list[str]) -> dict[str, list[str]]:
    """Gruppiert Dateien mit Teil-Suffixen zu einem Bon. Rückgabe: {base: [teile...]}."""
    groups = defaultdict(list)
    for f in filepaths:
        base = _normalize_basename_for_grouping(f)
        groups[base].append(f)
    # Teile je Gruppe sortieren (a<b<c<...<1<2<...<-part1<…)
    for k in list(groups.keys()):
        groups[k].sort(key=_part_order_key)
    return dict(groups)

def parse_pdf_fuel_block(best, text, txt_up):
    import re

    # Laden
    m_store = re.search(r"(ZG\s+TANKSTELLE[^\n\r]*)", text, flags=re.I)
    if m_store:
        best["Laden"] = m_store.group(1).strip()
    elif re.search(r"HONECK-WALDSCHÜTZ ENERGIE", txt_up):
        best["Laden"] = "Honeck-Waldschütz Energie GmbH"

    # Datum / Uhrzeit
    m_dt = re.search(r"(\d{2}\.\d{2}\.\d{2,4})\s+(\d{2}:\d{2})", text)
    if m_dt:
        m_date = re.match(r"(\d{2})\.(\d{2})\.(\d{2,4})", m_dt.group(1))
        if m_date:
            dd, mm, yy = m_date.groups()
            if len(yy) == 2:
                yy = "20" + yy
            best["Datum"] = f"{yy}-{mm}-{dd}"
            best["Uhrzeit"] = m_dt.group(2)

    # Kraftstoffart
    if re.search(r"\bSUPER(\s*E10|\s*E5| BLFR\.)?\b", txt_up):
        best["Kraftstoffart"] = "Super"
    elif re.search(r"\bDIESEL\b", txt_up):
        best["Kraftstoffart"] = "Diesel"

    # Zahlung
    if re.search(r"\bABGEBUCHT\b", txt_up) or re.search(r"\bIBAN\b", txt_up):
        best["Zahlung"] = "Lastschrift"

    # Betrag
    amount_candidates = []

    for m in re.finditer(r"(?is)(\d{1,4}[.,]\d{2})\s*\*?\s*ENDBETRAG", text):
        amount_candidates.append(m.group(1))

    for m in re.finditer(r"(?is)ENDBETRAG[^\d]{0,40}(\d{1,4}[.,]\d{2})", text):
        amount_candidates.append(m.group(1))

    for m in re.finditer(r"(?is)BRUTTOBETRAG[^\d]{0,60}(\d{1,4}[.,]\d{2})", text):
        amount_candidates.append(m.group(1))

    for m in re.finditer(r"(?is)\bSUMME\b[^\d]{0,40}(\d{1,4}[.,]\d{2})", text):
        amount_candidates.append(m.group(1))

    if amount_candidates:
        try:
            vals = [float(x.replace(",", ".")) for x in amount_candidates]
            best["Betrag (€)"] = max(vals)
        except Exception:
            pass

    return best

def parse_pdf_grocery_block(best, text, txt_up):
    import re

    # Laden
    for brand in ["KAUFLAND", "LIDL", "ALDI", "REWE", "EDEKA", "NETTO", "PENNY"]:
        if brand in txt_up:
            best["Laden"] = brand
            break

    # Datum / Uhrzeit
    m_dt = re.search(r"(\d{2}[.,]\d{2}[.,]\d{2,4}).*?(\d{2}:\d{2})(?::\d{2})?", text)
    if m_dt:
        raw_date = m_dt.group(1).replace(",", ".")
        m_date = re.match(r"(\d{2})\.(\d{2})\.(\d{2,4})", raw_date)
        if m_date:
            dd, mm, yy = m_date.groups()
            if len(yy) == 2:
                yy = "20" + yy
            best["Datum"] = f"{yy}-{mm}-{dd}"
        best["Uhrzeit"] = m_dt.group(2)

    # Zahlung
    if re.search(r"\b(LIDL PAY|KAUFLAND PAY|BLUECODE|EC|GIROCARD|VISA|MASTERCARD|KARTE)\b", txt_up):
        best["Zahlung"] = "Karte"
    elif re.search(r"\bBAR\b", txt_up):
        best["Zahlung"] = "Bar"

    # Betrag
    amount_candidates = []
    for m in re.finditer(r"(?is)(KAUFLAND PAY|LIDL PAY|ZU ZAHLEN|SUMME)[^\d]{0,40}(\d{1,4}[.,]\d{2})", text):
        amount_candidates.append(m.group(2))

    if amount_candidates:
        try:
            vals = [float(x.replace(",", ".")) for x in amount_candidates]
            best["Betrag (€)"] = max(vals)
        except Exception:
            pass

    return best

def fix_store_from_known_brands(best: dict, txt_up: str) -> dict:
    import re

    if not best.get("Laden"):
        return best

    cur_store = str(best["Laden"]).upper()

    looks_bad_store = (
        re.search(r"\b\d{5}\b", cur_store) is not None or
        re.search(r"\b(STRASSE|STRAßE|STR\.|WEG|PLATZ|ALLEE|HÖHE|HOEHE)\b", cur_store) is not None or
        re.fullmatch(r"DE\d{6,}", cur_store) is not None or
        re.fullmatch(r"[\d\s\-./]+", cur_store) is not None
    )

    if looks_bad_store:
        for brand in ["KAUFLAND", "LIDL", "ALDI", "REWE", "EDEKA", "NETTO", "PENNY"]:
            if brand in txt_up:
                best["Laden"] = brand
                break

    return best

#===============================================================================
def scan_pdf_receipt(
    pdf_path: str,
    excel_path: str = "kassenbons.xlsx",
    review_when: str = "prüfen",
    strict_total: bool = True,
    show_debug_footer: bool = False,
    debug_print: bool = False,
    **kwargs
):
    """
    PDF-Beleg einlesen (z. B. Tankkarten-Abrechnung),
    Text extrahieren, parsen und nach Excel speichern.
    """
    import os
    if not os.path.isfile(pdf_path):
        print(f"❗ PDF nicht gefunden: {pdf_path}")
        return None

    try:
        # 1) PDF-Text direkt lesen
        pdf_text = extract_text_from_pdf(pdf_path)
        if not pdf_text:
            print("❗ Kein brauchbarer Text aus PDF erhalten.")
            return None

        # 2) Direkt verwenden
        combined_text = pdf_text

        # 3) Debug-Ausgaben (analog zu scan_kassenbon)
        if show_debug_footer or debug_print:
            print("\n--- CLEAN HEAD (PDF) ---")
            print("\n".join(combined_text.splitlines()[:25]))
            print("\n--- CLEAN (letzte 25 Zeilen, PDF) ---")
            print("\n".join(combined_text.splitlines()[-25:]))

        import re

        txt_up = combined_text.upper()
        best = {"Rohtext": combined_text}

        # PDF direkt mit dem allgemeinen Textparser versuchen
        base = extract_data_from_text_only(combined_text) or {}
        best.update(base)

        # Fuel-Rechnung gezielt nachschärfen
        if re.search(r"\b(TANKRECHNUNG|TANKSTELLE|SUPER|DIESEL|EUR/L|LITER|KRAFTSTOFF)\b", txt_up):
            best["Belegtyp"] = "fuel"
            best = parse_pdf_fuel_block(best, combined_text, txt_up)
            rtype = "fuel"
        elif re.search(r"\b(KAUFLAND|LIDL|ALDI|REWE|EDEKA|NETTO|PENNY)\b", txt_up):
            best["Belegtyp"] = "grocery"
            best = parse_pdf_grocery_block(best, combined_text, txt_up)
            rtype = "grocery"
        else:
            rtype = best.get("Belegtyp", "generic")

        rtype = best.get("Belegtyp", "generic")

        # Wenn im ganzen PDF klar eine Marke vorkommt, aber der Laden wie eine Adresse aussieht:
        best = fix_store_from_known_brands(best, txt_up)

        best, rtype = apply_enrichers(best, rtype)
        best = normalize_tax_fields(best)
        #print("\n=== DEBUG VOR STATUS 5 ===")
        #print(best)
        return finalize_and_save_receipt(best, rtype, excel_path)

    except Exception as e:
        print(f"❌ Fehler bei {pdf_path}: {e}")
        return None


def batch_scan_folder(folder: str,
                      excel_path: str,
                      review_when: str = "prüfen",
                      strict_total: bool = True,
                      move_processed: bool = True,
                      verbose: bool = True):
    import os, shutil

    exts = (".jpg",".jpeg",".png",".tif",".tiff",".bmp",".webp",".pdf")
    files = [os.path.join(folder, f) for f in os.listdir(folder) if f.lower().endswith(exts)]
    files.sort()

    if not files:
        print("📂 Keine passenden Bilddateien im Ordner gefunden.")
        return {"total": 0, "ok": 0, "pruefen": 0, "error": 0, "saved": 0}

    if verbose:
        print(f"📂 Starte Batch-Scan: {folder}  |  Dateien: {len(files)}")
        print(f"   review_when={review_when}  strict_total={strict_total}  move_processed={move_processed}")

    # Zielordner für Sortierung
    if move_processed:
        ok_dir     = os.path.join(folder, "_ok")
        review_dir = os.path.join(folder, "_reviewed")
        error_dir  = os.path.join(folder, "_error")
        for d in (ok_dir, review_dir, error_dir):
            os.makedirs(d, exist_ok=True)

    # === NEU: Gruppen bilden
    groups = group_receipt_parts(files)

    total = saved = okc = rc = ec = 0

    for base, paths in groups.items():
        # Hübscher Name für die Logzeile
        fname = (os.path.basename(paths[0])
                 if len(paths) == 1
                 else os.path.basename(base) + f" (+{len(paths) - 1} Teile)")
        print(f"\n📄 Verarbeite: {fname}")
    
        try:
            # --- Auswahl: PDF-Einzeldatei / Bild-Einzeldatei / Mehrteiler ---
            if len(paths) == 1 and _is_pdf(paths[0]):
                # Einzel-PDF
                result = scan_pdf_receipt(
                    paths[0],
                    excel_path=excel_path,
                    review_when=review_when,
                    strict_total=strict_total,
                    show_debug_footer=DEBUG_PRINTS,
                    debug_print=DEBUG_PRINTS,          # <<<< sorgt für CLEAN-Ausgabe
                )
            elif len(paths) == 1:
                # Einzelbild
                result = scan_kassenbon(
                    paths[0],
                    excel_path=excel_path,
                    review_when=review_when,
                    strict_total=strict_total,
                    show_debug_footer=False,
                )
            else:
                # Mehrteiliger Bon (mehrere Bilder)
                print(f"\n=== scan_kassenbon_group AKTIV: {image_paths} | USE_PADDLE={USE_PADDLE} ===")
                result = scan_kassenbon_group(
                    paths,
                    excel_path=excel_path,
                    review_when=review_when,
                    strict_total=strict_total,
                    show_debug_footer=False,
                )
    
            # Ergebnis prüfen
            if not result:
                raise RuntimeError("scan_* lieferte kein Ergebnis")
    
            total += 1
            saved += 1
    
            status = (result.get("Prüfstatus") or "").lower()
            is_ok = ("✅" in status) or (status == "ok")
            needs_review = ("prüf" in status) or ("review" in status)
    
            if is_ok:
                okc += 1
            elif needs_review:
                rc += 1
            else:
                ec += 1
    
            # Optional: Dateien verschieben (alle Teile eines Bons gemeinsam)
            if move_processed:
                try:
                    dest_dir = ok_dir if is_ok else (review_dir if needs_review else error_dir)
                    for p in paths:
                        shutil.move(p, os.path.join(dest_dir, os.path.basename(p)))
                except Exception as mv_e:
                    print(f"⚠️ Verschieben fehlgeschlagen: {mv_e}")
    
        except Exception as e:
            total += 1
            ec += 1
            print(f"❌ Fehler bei {paths[0]}: {e}")
            if move_processed:
                try:
                    for p in paths:
                        shutil.move(p, os.path.join(error_dir, os.path.basename(p)))
                except Exception as mv_e:
                    print(f"⚠️ Verschieben in _error fehlgeschlagen: {mv_e}")

    print("\n📊 Batch-Summary")
    print(f"  Total:   {total}")
    print(f"  Saved:   {saved}")
    print(f"  OK:      {okc}")
    print(f"  Prüfen:  {rc}")
    print(f"  Error:   {ec}")

    return {"total": total, "ok": okc, "pruefen": rc, "error": ec, "saved": saved}


# =========================
# STARTPUNKT / MANUELLER LAUF
# =========================

if __name__ == "__main__":
    TARGET_PATH     = r"C:\Users\ONeum\Documents\Projekte\Kassenbon-Scanner\Kassenbons"  # Datei ODER Ordner
    EXCEL_PATH      = "kassenbons.xlsx"
    REVIEW_WHEN     = "prüfen"
    STRICT_TOTAL    = True
    MOVE_PROCESSED  = True
    VERBOSE         = True
    SHOW_FOOTER_DBG = False

    import os
    if os.path.isdir(TARGET_PATH):
        batch_scan_folder(
            TARGET_PATH,
            excel_path=EXCEL_PATH,
            review_when=REVIEW_WHEN,
            strict_total=STRICT_TOTAL,
            move_processed=MOVE_PROCESSED,
            verbose=VERBOSE,
        )
    elif os.path.isfile(TARGET_PATH):
        print(f"🖼️ Einzeldatei erkannt: {os.path.basename(TARGET_PATH)}")
        scan_kassenbon(
            TARGET_PATH,
            excel_path=EXCEL_PATH,
            review_when=REVIEW_WHEN,
            strict_total=STRICT_TOTAL,
            show_debug_footer=SHOW_FOOTER_DBG,
        )
    else:
        print(f"❗ Pfad existiert nicht: {TARGET_PATH}")
